import re
import unicodedata
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import load_workbook


MONEY_ZERO = Decimal("0.00")
IMPORTED_SALARY_FIELDS = (
    "total_salary_ron",
    "salary_advance_ron",
    "salary_remainder_ron",
    "meal_vouchers_ron",
)


class SalarySourceError(ValueError):
    pass


@dataclass(frozen=True)
class SalarySourceRow:
    source_file: str
    sheet: str
    row_number: int
    company: str
    raw_name: str
    employee_name: str
    identifiers: tuple[str, ...]
    total: Decimal
    advance: Decimal
    remainder: Decimal
    meal_vouchers: Decimal
    garnishment: Decimal = MONEY_ZERO
    ignored_difference: Decimal = MONEY_ZERO

    @property
    def location(self):
        return f"{self.source_file} / {self.sheet} / rândul {self.row_number}"


def normalize_text(value):
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(character for character in text if not unicodedata.combining(character))
    return " ".join(re.findall(r"[a-z0-9]+", text.casefold()))


def normalize_identifier(value):
    return re.sub(r"[^A-Z0-9]", "", unicodedata.normalize("NFKD", str(value or "")).upper())


def _name_aliases(value):
    tokens = normalize_text(value).split()
    aliases = {"dtru": "dumitru", "gh": "gheorghe"}
    return " ".join(aliases.get(token, token) for token in tokens)


def clean_source_name(value):
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    text = re.split(r"_", text, maxsplit=1)[0]
    text = re.split(
        r",\s*(?:\d{6,}|CIM\b|CNP\b)|\s+CNP\b|\s+CIM\b|\s*/\s*AA\b|\s+salar\b",
        text,
        maxsplit=1,
        flags=re.IGNORECASE,
    )[0]
    return re.sub(r"\s+", " ", text).strip(" ,._-")


def clean_employee_name(value):
    text = str(value or "").split("//", 1)[0]
    text = re.sub(r"\s+-\s+Rara\b.*$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\b[A-Z]{1,3}\d{5,9}\b", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+\d+$", "", text)
    return re.sub(r"\s+", " ", text).strip(" ,._-")


def extract_identifiers(value):
    text = unicodedata.normalize("NFKD", str(value or "")).upper()
    found = []
    patterns = (
        r"(?<!\d)\d{6,14}(?!\d)",
        r"\b[A-Z]{1,3}\s?\d{5,9}\b",
    )
    for pattern in patterns:
        for match in re.findall(pattern, text):
            normalized = normalize_identifier(match)
            if normalized and normalized not in found:
                found.append(normalized)
    return tuple(found)


def money(value, *, location, field):
    if value is None or str(value).strip() in {"", "-"}:
        return None
    if isinstance(value, bool):
        raise SalarySourceError(f"{location}: valoare invalidă pentru {field}: {value!r}")
    raw = str(value).strip().replace(" ", "").replace(",", ".")
    try:
        result = Decimal(raw).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError) as exc:
        raise SalarySourceError(
            f"{location}: valoare invalidă pentru {field}: {value!r}"
        ) from exc
    if result < 0:
        raise SalarySourceError(f"{location}: {field} nu poate fi negativ: {value!r}")
    return result


def _header_field(value):
    header = normalize_text(value)
    if not header:
        return None
    if "nume" in header and ("angajat" in header or "prenume" in header):
        return "name"
    if "poprire" in header:
        return "garnishment"
    if "lichidare" in header or header == "rest" or header.startswith("rest salari"):
        return "remainder"
    if "total stat plata" in header or "salar stat plata" in header or "net tichete stat plata" in header:
        return "total"
    if "tichet" in header or "bonuri" in header:
        return "meal_vouchers"
    if "avans" in header:
        return "advance"
    return None


def _decimal_or_zero(value, *, location, field):
    return money(value, location=location, field=field) or MONEY_ZERO


def _read_xlsx(path):
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise SalarySourceError(f"Nu pot deschide fișierul XLSX '{path}': {exc}") from exc
    try:
        return [
            (sheet.title, [tuple(row) for row in sheet.iter_rows(values_only=True)])
            for sheet in workbook.worksheets
        ]
    finally:
        workbook.close()


def _read_xls(path):
    try:
        import xlrd
    except ImportError as exc:
        raise SalarySourceError(
            "Pentru fișierele .xls lipsește pachetul xlrd. "
            "Rulează mai întâi: python -m pip install -r requirements.txt"
        ) from exc
    try:
        workbook = xlrd.open_workbook(path, on_demand=True)
    except Exception as exc:
        raise SalarySourceError(f"Nu pot deschide fișierul XLS '{path}': {exc}") from exc
    try:
        return [
            (sheet_name, [tuple(workbook.sheet_by_name(sheet_name).row_values(index)) for index in range(workbook.sheet_by_name(sheet_name).nrows)])
            for sheet_name in workbook.sheet_names()
        ]
    finally:
        workbook.release_resources()


def read_salary_workbook(path):
    path = Path(path)
    if not path.is_file():
        raise SalarySourceError(f"Fișierul nu există: {path}")
    if path.suffix.casefold() == ".xlsx":
        return _read_xlsx(path)
    if path.suffix.casefold() == ".xls":
        return _read_xls(path)
    raise SalarySourceError(f"Format neacceptat pentru '{path}'; sunt acceptate .xls și .xlsx.")


def _cell(row, index):
    if index is None or index >= len(row):
        return None
    return row[index]


def _company_from_header(row):
    for value in row:
        if value in (None, "") or _header_field(value):
            continue
        text = str(value).strip()
        normalized = normalize_text(text)
        if not normalized or normalized.isdigit():
            continue
        if normalized in {"nr", "nr crt"}:
            continue
        if "salarii" in normalized or re.search(r"\b(?:ianuarie|februarie|martie|aprilie|mai|iunie|iulie|august|septembrie|octombrie|noiembrie|decembrie)\b", normalized):
            continue
        return text
    return ""


def _parse_generic_sheet(path, sheet_name, rows):
    columns = {
        "name": None,
        "advance": None,
        "garnishment": None,
        "remainder": None,
        "meal_vouchers": None,
        "total": None,
    }
    company = ""
    parsed = []

    for row_number, row in enumerate(rows, start=1):
        header_cells = [(index, _header_field(value)) for index, value in enumerate(row)]
        header_cells = [(index, field) for index, field in header_cells if field]
        is_money_header = any(field in {"advance", "garnishment", "remainder", "meal_vouchers", "total"} for _, field in header_cells)
        if header_cells:
            if is_money_header:
                for field in ("advance", "garnishment", "remainder", "meal_vouchers", "total"):
                    columns[field] = None
            for index, field in header_cells:
                columns[field] = index
            header_company = _company_from_header(row)
            if header_company:
                company = header_company
            continue

        non_empty = [value for value in row if value not in (None, "")]
        if non_empty and all(not isinstance(value, (int, float, Decimal)) for value in non_empty):
            section_company = _company_from_header(row)
            if section_company:
                company = section_company
            continue

        raw_name = _cell(row, columns["name"])
        if raw_name in (None, ""):
            continue
        location = f"{Path(path).name} / {sheet_name} / rândul {row_number}"
        raw_values = {
            field: money(_cell(row, columns[field]), location=location, field=field)
            for field in ("advance", "garnishment", "remainder", "meal_vouchers", "total")
        }
        if all(value is None for value in raw_values.values()):
            continue
        if raw_values["total"] is None:
            raise SalarySourceError(f"{location}: lipsește salariul total.")

        garnishment = raw_values["garnishment"] or MONEY_ZERO
        base_remainder = raw_values["remainder"] or MONEY_ZERO
        raw_name = str(raw_name).strip()
        parsed.append(SalarySourceRow(
            source_file=Path(path).name,
            sheet=sheet_name,
            row_number=row_number,
            company=company,
            raw_name=raw_name,
            employee_name=clean_source_name(raw_name),
            identifiers=extract_identifiers(raw_name),
            total=raw_values["total"],
            advance=raw_values["advance"] or MONEY_ZERO,
            remainder=(base_remainder + garnishment).quantize(Decimal("0.01")),
            meal_vouchers=raw_values["meal_vouchers"] or MONEY_ZERO,
            garnishment=garnishment,
        ))
    return parsed


def _parse_xmeg_sheet(path, sheet_name, rows):
    parsed = []
    # Contractul acestei foi este intenționat explicit: F este coloana etichetată
    # „AVANS STAT PLATA”. E (fără antet) și G („DIF”) nu sunt confundate cu avansul.
    for row_number, row in enumerate(rows[2:], start=3):
        raw_name = _cell(row, 3)
        if raw_name in (None, ""):
            continue
        location = f"{Path(path).name} / {sheet_name} / rândul {row_number}"
        total = money(_cell(row, 10), location=location, field="total")
        if total is None:
            continue
        raw_name = str(raw_name).strip()
        parsed.append(SalarySourceRow(
            source_file=Path(path).name,
            sheet=sheet_name,
            row_number=row_number,
            company="XMEG",
            raw_name=raw_name,
            employee_name=clean_source_name(raw_name),
            identifiers=extract_identifiers(raw_name),
            total=total,
            advance=_decimal_or_zero(_cell(row, 5), location=location, field="advance"),
            remainder=_decimal_or_zero(_cell(row, 8), location=location, field="remainder"),
            meal_vouchers=_decimal_or_zero(_cell(row, 9), location=location, field="meal_vouchers"),
            ignored_difference=_decimal_or_zero(_cell(row, 6), location=location, field="difference"),
        ))
    return parsed


def parse_salary_files(paths):
    parsed = []
    for raw_path in paths:
        path = Path(raw_path).expanduser().resolve()
        for sheet_name, rows in read_salary_workbook(path):
            if normalize_text(sheet_name) == "xmeg":
                parsed.extend(_parse_xmeg_sheet(path, sheet_name, rows))
            else:
                parsed.extend(_parse_generic_sheet(path, sheet_name, rows))
    if not parsed:
        raise SalarySourceError("Nu am găsit niciun rând salarial valid în fișiere.")
    return parsed


def _company_key(value):
    normalized = normalize_text(value)
    aliases = (
        ("xux", "xux"),
        ("servicex", "servicex"),
        ("dmx", "dmx"),
        ("victor steel", "victor steel"),
        ("munteanu", "munteanu"),
        ("omnia", "omnia"),
        ("rnx", "rnx"),
        ("vb rom", "vb rom"),
        ("xmeg", "xmeg"),
    )
    for needle, key in aliases:
        if needle in normalized:
            return key
    return normalized


def _name_score(source_name, employee_name):
    source = _name_aliases(source_name)
    employee = _name_aliases(clean_employee_name(employee_name))
    if not source or not employee:
        return 0.0
    if source == employee:
        return 100.0
    source_tokens = source.split()
    employee_tokens = employee.split()
    if sorted(source_tokens) == sorted(employee_tokens):
        return 98.0
    source_set = set(source_tokens)
    employee_set = set(employee_tokens)
    smaller = min(len(source_set), len(employee_set))
    larger = max(len(source_set), len(employee_set))
    if smaller >= 2 and (source_set <= employee_set or employee_set <= source_set):
        return 90.0 + (5.0 * smaller / larger)
    ratio = SequenceMatcher(None, source, employee).ratio()
    intersection = len(source_set & employee_set)
    union = len(source_set | employee_set)
    jaccard = intersection / union if union else 0
    # Potrivirea caracter-cu-caracter este permisă numai când primul cuvânt este
    # identic (de ex. VASI/VASILE). Altfel DALVIR nu trebuie asociat cu DAVINDER.
    if ratio >= 0.84 and source_tokens[0] == employee_tokens[0]:
        return 80.0 + ratio * 8.0
    if intersection >= 2 and jaccard >= 0.6:
        return 80.0 + jaccard * 5.0
    return 0.0


def match_salary_row(row, employees):
    identifier_candidates = []
    for employee in employees:
        employee_identifier = normalize_identifier(employee.UserSerie)
        if any(len(identifier) >= 6 and identifier in employee_identifier for identifier in row.identifiers):
            identifier_candidates.append(employee)
    if len(identifier_candidates) == 1:
        return identifier_candidates[0], []

    candidates = identifier_candidates or employees
    scored = []
    source_company = _company_key(row.company)
    for employee in candidates:
        score = _name_score(row.employee_name, employee.UserName)
        if not score:
            continue
        if source_company and source_company == _company_key(employee.Company):
            score += 4.0
        scored.append((score, employee))
    scored.sort(key=lambda item: (-item[0], item[1].UserId))
    if not scored or scored[0][0] < 82.0:
        return None, []
    top_score = scored[0][0]
    tied = [employee for score, employee in scored if top_score - score < 1.0]
    if len(tied) > 1:
        return None, tied
    return scored[0][1], []


def aggregate_salary_rows(matches):
    aggregated = {}
    for row, employee in matches:
        item = aggregated.setdefault(employee.pk, {
            "employee": employee,
            "rows": [],
            "total_salary_ron": MONEY_ZERO,
            "salary_advance_ron": MONEY_ZERO,
            "salary_remainder_ron": MONEY_ZERO,
            "meal_vouchers_ron": MONEY_ZERO,
            "garnishment": MONEY_ZERO,
        })
        item["rows"].append(row)
        item["total_salary_ron"] += row.total
        item["salary_advance_ron"] += row.advance
        item["salary_remainder_ron"] += row.remainder
        item["meal_vouchers_ron"] += row.meal_vouchers
        item["garnishment"] += row.garnishment
    return aggregated
