import re
import unicodedata
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from ToolApp.models import Tools, Users


EXPECTED_HEADERS = {
    "id intern": "internal_id",
    "nr inventar sursa": "source_inventory_number",
    "denumire scula echipament": "tool_name",
    "categorie": "category",
    "marca": "brand",
    "model": "model",
    "serie sn": "serial_number",
    "cantitate": "pieces",
    "stare": "source_status",
    "locatie": "location",
    "responsabil": "responsible",
    "observatii": "detail",
    "necesita verificare": "requires_verification",
    "poza sursa": "source_photo",
}

REQUIRED_FIELDS = {"internal_id", "tool_name"}

IMPORTED_FIELDS = [
    "ToolName",
    "User",
    "Detail",
    "Pieces",
    "MainLocation",
    "AssignedTo",
    "IsSSM",
    "Status",
    "IsReturned",
    "IsLost",
    "DateReturned",
    "DateLost",
    "SourceInventoryNumber",
    "Category",
    "Brand",
    "Model",
    "SerialNumber",
    "SourceStatus",
    "RequiresVerification",
    "SourcePhoto",
]

FIELD_LIMITS = {
    "internal_id": 100,
    "tool_name": 100,
    "category": 200,
    "brand": 100,
    "model": 100,
    "serial_number": 200,
    "source_status": 100,
    "detail": 500,
    "location": 500,
    "responsible": 100,
    "source_photo": 255,
}


def clean_text(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def normalize_text(value):
    text = unicodedata.normalize("NFD", str(value or ""))
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    text = re.sub(r"[^a-zA-Z0-9]+", " ", text).strip().lower()
    return text


def parse_positive_integer(value, *, default=None):
    if value in (None, ""):
        return default
    if isinstance(value, bool):
        raise ValueError(f"valoare invalidă: {value!r}")
    if isinstance(value, (int, float)) and float(value).is_integer():
        result = int(value)
    else:
        match = re.search(r"\d+", str(value))
        if not match:
            raise ValueError(f"nu conține un număr: {value!r}")
        result = int(match.group(0))
    if result < 1:
        raise ValueError(f"trebuie să fie cel puțin 1: {value!r}")
    return result


def is_yes(value):
    return normalize_text(value) in {"da", "yes", "true", "1", "x"}


def is_ssm_category(value):
    normalized = normalize_text(value)
    return normalized.startswith("siguranta protectie") or normalized == "ssm"


class Command(BaseCommand):
    help = "Importă idempotent uneltele din foaia 'Inventar extras' a unui fișier XLSX."

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", help="Calea către fișierul .xlsx de importat.")
        parser.add_argument(
            "--sheet",
            default="Inventar extras",
            help="Numele foii de import (implicit: Inventar extras).",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Validează și afișează rezultatul fără să modifice baza de date.",
        )
        parser.add_argument(
            "--update-existing",
            action="store_true",
            help="Actualizează uneltele care au deja același ID intern / ToolSerie.",
        )

    def handle(self, *args, **options):
        path = Path(options["xlsx_path"]).expanduser()
        sheet_name = options["sheet"]
        dry_run = options["dry_run"]
        update_existing = options["update_existing"]

        if not path.is_file():
            raise CommandError(f"Fișierul nu există: {path}")
        if path.suffix.lower() != ".xlsx":
            raise CommandError("Fișierul trebuie să aibă extensia .xlsx.")

        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            raise CommandError(f"Nu pot deschide fișierul XLSX: {exc}") from exc

        try:
            if sheet_name not in workbook.sheetnames:
                available = ", ".join(workbook.sheetnames)
                raise CommandError(
                    f"Foaia '{sheet_name}' nu există. Foi disponibile: {available}"
                )

            sheet = workbook[sheet_name]
            rows = sheet.iter_rows(values_only=True)
            try:
                raw_headers = next(rows)
            except StopIteration as exc:
                raise CommandError("Foaia de import este goală.") from exc

            columns = self._map_headers(raw_headers)
            parsed_rows, warnings = self._parse_rows(rows, columns)
        finally:
            workbook.close()

        if not parsed_rows:
            raise CommandError("Nu am găsit niciun rând de importat.")

        series = [item["ToolSerie"] for item in parsed_rows]
        duplicate_series = sorted({value for value in series if series.count(value) > 1})
        if duplicate_series:
            preview = ", ".join(duplicate_series[:10])
            raise CommandError(f"ID-uri interne duplicate în Excel: {preview}")

        user_map = self._build_user_map()
        existing = {
            tool.ToolSerie: tool
            for tool in Tools.objects.filter(ToolSerie__in=series).select_related("AssignedTo")
        }

        to_create = []
        to_update = []
        skipped = 0
        total_pieces = 0

        for item in parsed_rows:
            responsible = item.pop("_responsible")
            assigned_user = user_map.get(normalize_text(responsible)) if responsible else None
            if responsible and assigned_user is None:
                warnings.append(
                    f"{item['ToolSerie']}: responsabilul '{responsible}' nu există în Users; "
                    "numele a fost păstrat doar în câmpul User."
                )

            item["AssignedTo"] = assigned_user
            item["User"] = assigned_user.UserName if assigned_user else responsible
            if assigned_user and item["Status"] != Tools.ToolStatus.STRICATA:
                item["Status"] = Tools.ToolStatus.IN_LUCRU
                if item["MainLocation"] == "Magazie":
                    item["MainLocation"] = assigned_user.UserName

            total_pieces += item["Pieces"]
            current = existing.get(item["ToolSerie"])
            if current is None:
                to_create.append(Tools(**item))
            elif update_existing:
                for field, value in item.items():
                    if field != "ToolSerie":
                        setattr(current, field, value)
                to_update.append(current)
            else:
                skipped += 1

        summary = (
            f"Rânduri valide: {len(parsed_rows)} | bucăți: {total_pieces} | "
            f"de creat: {len(to_create)} | de actualizat: {len(to_update)} | "
            f"existente omise: {skipped}"
        )
        self.stdout.write(summary)

        if warnings:
            self.stdout.write(self.style.WARNING(f"Avertismente: {len(warnings)}"))
            for warning in warnings[:20]:
                self.stdout.write(self.style.WARNING(f"- {warning}"))
            if len(warnings) > 20:
                self.stdout.write(self.style.WARNING(f"- ... încă {len(warnings) - 20}"))

        if dry_run:
            self.stdout.write(self.style.SUCCESS("DRY-RUN finalizat: baza de date nu a fost modificată."))
            return

        with transaction.atomic():
            if to_create:
                Tools.objects.bulk_create(to_create, batch_size=400)
            if to_update:
                Tools.objects.bulk_update(to_update, IMPORTED_FIELDS, batch_size=400)

        self.stdout.write(
            self.style.SUCCESS(
                f"Import finalizat: {len(to_create)} create, {len(to_update)} actualizate, "
                f"{skipped} omise."
            )
        )

    def _map_headers(self, raw_headers):
        columns = {}
        for index, header in enumerate(raw_headers):
            field = EXPECTED_HEADERS.get(normalize_text(header))
            if field:
                columns[field] = index

        missing = sorted(REQUIRED_FIELDS - set(columns))
        if missing:
            raise CommandError(
                "Lipsesc coloanele obligatorii: " + ", ".join(missing)
            )
        return columns

    def _parse_rows(self, rows, columns):
        parsed = []
        warnings = []

        for excel_row_number, row in enumerate(rows, start=2):
            if not any(value not in (None, "") for value in row):
                continue

            values = {
                field: clean_text(row[index]) if index < len(row) else None
                for field, index in columns.items()
            }
            internal_id = values.get("internal_id")
            tool_name = values.get("tool_name")
            if not internal_id or not tool_name:
                raise CommandError(
                    f"Rândul {excel_row_number}: ID intern și denumirea sunt obligatorii."
                )

            for field, limit in FIELD_LIMITS.items():
                value = values.get(field)
                if value and len(value) > limit:
                    raise CommandError(
                        f"Rândul {excel_row_number}, {field}: depășește {limit} caractere."
                    )

            try:
                pieces = parse_positive_integer(values.get("pieces"), default=1)
                source_number = parse_positive_integer(
                    values.get("source_inventory_number"), default=None
                )
            except ValueError as exc:
                raise CommandError(f"Rândul {excel_row_number}: {exc}") from exc

            source_status = values.get("source_status")
            normalized_status = normalize_text(source_status)
            requires_verification = is_yes(values.get("requires_verification"))
            if "verificat" in normalized_status:
                requires_verification = True

            status = (
                Tools.ToolStatus.STRICATA
                if "defect" in normalized_status
                else Tools.ToolStatus.MAGAZIE
            )
            location = values.get("location") or "Magazie"

            parsed.append(
                {
                    "ToolSerie": internal_id,
                    "ToolName": tool_name,
                    "User": None,
                    "Detail": values.get("detail"),
                    "Pieces": pieces,
                    "MainLocation": location,
                    "AssignedTo": None,
                    "IsSSM": is_ssm_category(values.get("category")),
                    "Status": status,
                    "IsReturned": False,
                    "IsLost": False,
                    "DateReturned": None,
                    "DateLost": None,
                    "SourceInventoryNumber": source_number,
                    "Category": values.get("category"),
                    "Brand": values.get("brand"),
                    "Model": values.get("model"),
                    "SerialNumber": values.get("serial_number"),
                    "SourceStatus": source_status,
                    "RequiresVerification": requires_verification,
                    "SourcePhoto": values.get("source_photo"),
                    "_responsible": values.get("responsible"),
                }
            )

        return parsed, warnings

    def _build_user_map(self):
        result = {}
        duplicates = set()
        for user in Users.objects.all():
            key = normalize_text(user.UserName)
            if key in result:
                duplicates.add(key)
            else:
                result[key] = user
        for key in duplicates:
            result.pop(key, None)
        return result
