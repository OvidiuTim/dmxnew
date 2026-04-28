# --- Standard library ---
import csv, io, re
import json, logging, math
import hashlib
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock
from queue import Queue, Empty
from collections import deque
from datetime import datetime, timedelta, date as _date
from decimal import Decimal, ROUND_HALF_UP
from typing import Optional

# --- Third-party ---
from rest_framework.parsers import JSONParser

# --- Django ---
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse, StreamingHttpResponse, HttpResponseNotAllowed, HttpResponseBadRequest
from django.utils import timezone
from django.utils.timezone import localtime, localdate
from django.db import transaction
from django.db.models import Sum, Min, Max, OuterRef, Subquery, Q
from django.core.cache import cache
from django.contrib.auth.hashers import check_password

# --- App (ToolApp) ---
from ToolApp.models import (
    Users, AttendanceSession, PresenceEvent, Tools, Histories,
    Shed, Unfunctional, Materials, Consumables, WorkField,
    CofrajMetalics, CofrajtTipDokas, Popis, SchelaUsoaras, SchelaFatadas,
    SchelaFatadaModularas, Combustibils, HistorieScheles, MijloaceFixes, DailyPay, LeaveDay,
    PinAttemptLog
)
from ToolApp.serializers import (
    ConsumableSerializer, ShedSerializer, UnfunctionalSerializer, UserSerializer, ToolSerializer,
    HistorySerializer, MaterialSerializer, WorkFieldSerializer, CofrajMetalicSerializer,
    CofrajtTipDokaSerializer, PopiSerializer, SchelaUsoaraSerializer, SchelaFatadaSerializer,
    SchelaFatadaModularaSerializer, CombustibilSerializer, HistorieScheleSerializer, MijloaceFixeSerializer, 
)
# --- API: /api/pay/day și /api/pay/month ---
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.utils.timezone import localdate
from django.db.models import Sum
from decimal import Decimal
# --- PONTAJ API ---
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.utils import timezone
from django.utils.timezone import localtime, localdate
from datetime import date, datetime
from django.db.models import Sum, Min, Max
from ToolApp.models import Users, AttendanceSession
from ToolApp.security import TOKEN_AGE, check_admin_token, get_token_from_request, make_admin_token

from datetime import datetime
from django.utils import timezone

# === AUTH SIMPLU PENTRU ANGULAR ===
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.conf import settings
import os, hmac, json

# === PONTAJ: program + leeway ===
from datetime import datetime, timedelta
from django.utils import timezone
from django.utils.timezone import localdate


from django.http import HttpResponse
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font


# --- PONTAJ: UPDATE / DELETE sesiune + DELETE zi ---

from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.utils import timezone
from django.db import transaction
from datetime import date as _date

# --- NFC (safe import) ---
try:
    import nfc  # nfcpy
except Exception:
    nfc = None






UID_PIN_MAP = {
    "A54936EF": "2839",
    "C55A37EF": "1414",
    "053B33EF": "5541",
    "B55536EF": "5043",
    "458F33EF": "4685",
    "85BB34EF": "3304",
    "A56A37EF": "2693",
    "05B533EF": "2437",
    "E53E36EF": "7964",
    "055036EF": "1525",

    # ADDED
    "F33E91ED": "9958",
    "E3689CED": "9999",
    "D33F92ED": "9998",
    "43C99CED": "9997",
    "F34791ED": "9996",
    "33A591ED": "9995",
    "934692ED": "9994",
    "230694ED": "9992",
    "2379A4ED": "9991",
    "734091ED": "9990",
    "835CA5ED": "9989",
    "738AA4ED": "9988",
    "13E693ED": "9986",
    "735392ED": "9985",
    "C37393ED": "9984",
    "03F492ED": "9983",
    "638E91ED": "9982",
    "1367AAED": "9981",
    "1313A6ED": "9980",
    "53DC91ED": "9979",
    "73C4A5ED": "9978",
    "C3A491ED": "9977",
    "A3FD93ED": "9976",
    "93AC93ED": "9953",
    "832A91ED": "9952",
    "A3DB92ED": "9950",
    "73619CED": "9949",
    "E3CC93ED": "9948",
    "13EB92ED": "9947",
    "F37B93ED": "9966",
    "63619CED": "9964",
    "938E91ED": "9963",
    "634E91ED": "9962",
    "E3DC91ED": "9961",
    "E3FE91ED": "9960",
    "738E91ED": "9959",
    "B379A4ED": "9956",
    "139691ED": "9955",
    "83D19CED": "9954",


    "A3C692ED": "9968",
    "F3C2A5ED": "9969",
    "13FD93ED": "9970",
    "E30594ED": "9971",
    # amarjit lost tag
    "C30594ED": "9958",
    "93A390ED": "9973",
    "93B393ED": "9974",
    "53E991ED": "9957",
    "E3F392ED": "8008",
    "F3CD93ED": "9946",
    "A3D091ED": "9945",

    # NOI
    "539C91ED": "9944",
    "B319A6ED": "9943",
    "339591ED": "9942",
    "B34091ED": "9941",
    "6326A5ED": "9940",
    "D35898ED": "9939",
    "13DC92ED": "9938",
    "234EA5ED": "9937",
    "033F91ED": "9935",
    "13C99CED": "9934",
    "838D91ED": "9933",
    "834091ED": "9932",
    "734E91ED": "9931",
    "F33091ED": "9930",
    "D3F992ED": "9929",
    "637B93ED": "9928",
    "83C3A5ED": "9927",
    "D33A9CED": "9926",
    "D39AA5ED": "9925",
    "3370AAED": "9924",
# admins
"657449EF": "11111",
"057944EF": "22222",

"957F37EF": "1156",
"252F36EF": "1157",
"A5EC35EF": "1158",
"657836EF": "1159",
"B5F832EF": "1160",
"85AE33EF": "1161",
"650733EF": "1162",
"752F36EF": "1163",
"65BC33EF": "1164",
"356337EF": "1165",
"656B37EF": "1166",
"E53333EF": "1167",
"855D36EF": "1168",
"257036EF": "1169",
"D57F37EF": "1170",
"95AE33EF": "1171",
"B57837EF": "1172",
"856B37EF": "1173",
"256C36EF": "1174",
"B5C236EF": "1175",
"158833EF": "1176",
"B5FD33EF": "1177",
"253B34EF": "1178",
"957136EF": "1179",
"456B37EF": "1180",
"855C36EF": "1181",
"556D36EF": "1182",
"B53936EF": "1184",
"356D36EF": "1185",
"852E36EF": "1186",
"753636EF": "1187",
"355036EF": "1188",
"855937EF": "1189",
"054334EF": "1190",
"C5F732EF": "1191",
"A56037EF": "1192",
"C58D33EF": "1193",
"55FF32EF": "1194",
"25FF32EF": "1195",
"058137EF": "1196",
"754136EF": "1197",
"85B436EF": "1178",
"853B34EF": "1199",
"255937EF": "1200",
"255A37EF": "1202",
"353736EF": "1203",
"B5A833EF": "1204",
"A5BC33EF": "1205",
"652736EF": "1206",
"E59433EF": "1207",
"75BC33EF": "1208",

    # Raspberry ACR122U fallback map - keep server-side compatibility with boxes
    # that send only UID, or NDEF content that is not a PIN.
    "A50033EF": "1201",
    "F5BB36EF": "2882",
    "453036EF": "9844",
    "455A37EF": "7955",
    "25F635EF": "2282",
    "65BE35EF": "2319",
    "755937EF": "2340",
    "05AD35EF": "4709",
    "B5D236EF": "2992",
    "55F832EF": "3587",
    "55ED35EF": "4095",
    "25AE33EF": "6417",
    "256F36EF": "8728",
    "D56536EF": "8012",
    "B58136EF": "8665",
    "156137EF": "8804",
    "C5B535EF": "8502",
    "F53136EF": "5410",
    "95AE35EF": "1153",
}



_last_seen = {}
_DEBOUNCE_SEC = 0.7



@csrf_exempt
def check_nfc_reader(request):
    if nfc is None:
        return JsonResponse({'has_nfc_reader': False, 'error': 'nfcpy not installed'})
    try:
        with nfc.ContactlessFrontend('usb'):
            return JsonResponse({'has_nfc_reader': True})
    except Exception:
        return JsonResponse({'has_nfc_reader': False})


    
# --- AUTO CLOSE CONFIG ---
AUTO_CLOSE_HOUR = 17
AUTO_CLOSE_MIN  = 30
AUTO_CLOSE_SOURCE_TAG = "|auto1730"
MISSING_EXIT_SOURCE_TAG = "|missing_exit"
PONTAJ_MAX_SHIFT_HOURS = 14  # deja îl ai, îl folosim pentru cap durată


def _day_time_dt(local_day, hour, minute):
    tz = timezone.get_current_timezone()
    naive = datetime(local_day.year, local_day.month, local_day.day, hour, minute, 0)
    return timezone.make_aware(naive, tz)


def _source_has_tag(source_value, tag):
    return tag in str(source_value or "")


def _is_missing_exit_session(session):
    return _source_has_tag(getattr(session, "source", ""), MISSING_EXIT_SOURCE_TAG)


def _mark_session_missing_exit(session):
    source = str(session.source or "")
    if MISSING_EXIT_SOURCE_TAG not in source:
        session.source = source + MISSING_EXIT_SOURCE_TAG
    session.out_time = None
    session.duration_seconds = 0




# Create your views here.
#api angajati
@csrf_exempt
def userApi(request,id=0):
    if request.method=='GET':
        if id:
            try:
                user = Users.objects.get(UserId=id)
            except Users.DoesNotExist:
                return JsonResponse({"error": "User not found"}, status=404)

            user_serializer = UserSerializer(user)
            return JsonResponse(user_serializer.data, safe=False)

        users = Users.objects.all()
        users_serializer = UserSerializer(users, many=True)
        return JsonResponse(users_serializer.data, safe=False)

    elif request.method=='POST':
        user_data=JSONParser().parse(request)
        user_serializer = UserSerializer(data=user_data)
        if user_serializer.is_valid():
            user = user_serializer.save()
            return JsonResponse(UserSerializer(user).data , safe=False, status=201)
        return JsonResponse({"error": "Failed to Add.", "details": user_serializer.errors},safe=False, status=400)
    
    elif request.method=='PUT':
        user_data = JSONParser().parse(request)
        try:
            user=Users.objects.get(UserId=user_data['UserId'])
        except (Users.DoesNotExist, KeyError):
            return JsonResponse({"error": "User not found"}, safe=False, status=404)

        user_serializer=UserSerializer(user,data=user_data)
        if user_serializer.is_valid():
            saved = user_serializer.save()
            return JsonResponse(UserSerializer(saved).data, safe=False)
        return JsonResponse({"error": "Failed to Update.", "details": user_serializer.errors}, safe=False, status=400)

    elif request.method=='DELETE':
        try:
            user=Users.objects.get(UserId=id)
        except Users.DoesNotExist:
            return JsonResponse({"error": "User not found"}, safe=False, status=404)
        user.delete()
        return JsonResponse("Deleted Succeffully!!", safe=False)

# --- BULK USERS: import JSON sau CSV ---
@csrf_exempt
def users_bulk(request):
    """
    POST /api/users/bulk/
    - Acceptă:
        a) JSON: [ { "UserName":"...", "UserSerie":"...", "UserPin":"...", "uid":"..." }, ... ]
           (sau { "rows":[...] })
        b) CSV (Content-Type: text/csv) cu headere: UserName,UserSerie,UserPin,uid
    Returnează: { "created": N }
    """
    if request.method != "POST":
        return JsonResponse({"error": "Only POST allowed"}, status=405)

    ctype = (request.META.get("CONTENT_TYPE") or "").lower()
    body  = request.body.decode("utf-8", "ignore")

    items = []
    try:
        if "text/csv" in ctype or body.lstrip().startswith("UserName"):
            # CSV
            reader = csv.DictReader(io.StringIO(body))
            for r in reader:
                items.append({
                    "UserName": r.get("UserName") or r.get("name"),
                    "UserSerie": r.get("UserSerie") or r.get("serie"),
                    "UserPin":   r.get("UserPin")   or r.get("pin"),
                    "uid":       r.get("uid")       or r.get("UID"),
                })
        else:
            # JSON
            data = json.loads(body or "[]")
            if isinstance(data, dict) and "rows" in data:
                data = data["rows"]
            if not isinstance(data, list):
                return JsonResponse({"error": "Body trebuie să fie listă JSON sau CSV"}, status=400)
            for r in data:
                items.append({
                    "UserName": r.get("UserName") or r.get("name"),
                    "UserSerie": r.get("UserSerie") or r.get("serie"),
                    "UserPin":   r.get("UserPin")   or r.get("pin"),
                    "uid":       r.get("uid")       or r.get("UID"),
                })
    except Exception as e:
        return JsonResponse({"error": f"Parsare eșuată: {e}"}, status=400)

    if not items:
        return JsonResponse({"error": "Nu am găsit rânduri de import"}, status=400)

    ser = UserSerializer(data=items, many=True)
    if not ser.is_valid():
        return JsonResponse({"errors": ser.errors}, status=400, safe=False)

    objs = ser.save()
    return JsonResponse({"created": len(objs)})
    
# --- BULK USERS: UPDATE (PUT JSON list) ---
@csrf_exempt
def users_bulk_update(request):
    """
    PUT /api/users/bulk_update/[?create_if_missing=1][&dry_run=1]
    Body (JSON): listă sau { "rows": [...] } cu obiecte user.
    Identificare per rând (în ordinea asta): UserId | UserSerie | UserPin | uid.
    Exemplu rând: { "UserId": 12, "UserName": "..." , "hourly_rate": 27.5 }
    """
    if request.method != "PUT":
        return JsonResponse({"error": "Only PUT allowed"}, status=405)

    create_if_missing = str(request.GET.get("create_if_missing", "0")).lower() in ("1","true","yes","on")
    dry_run = str(request.GET.get("dry_run", "0")).lower() in ("1","true","yes","on")

    # parse body
    try:
        body = request.body.decode("utf-8", "ignore")
        data = json.loads(body or "[]")
        if isinstance(data, dict) and "rows" in data:
            data = data["rows"]
        if not isinstance(data, list):
            return JsonResponse({"error": "Body trebuie să fie listă JSON sau {rows:[...]}"}, status=400)
    except Exception as e:
        return JsonResponse({"error": f"JSON invalid: {e}"}, status=400)

    updated = created = errors = 0
    results = []

    for idx, row in enumerate(data, start=1):
        # cheie de căutare: UserId | UserSerie | UserPin | uid
        lookup = None
        inst_by_pin = None
        if row.get("UserId") not in (None, ""):
            lookup = {"UserId": row.get("UserId")}
        elif row.get("UserSerie"):
            lookup = {"UserSerie": row.get("UserSerie")}
        elif row.get("UserPin"):
            inst_by_pin = _find_user_by_pin(row.get("UserPin"))
        elif row.get("uid"):
            lookup = {"uid": row.get("uid")}
        else:
            errors += 1
            results.append({"row": idx, "error": "Lipsește cheie de identificare (UserId sau UserSerie sau UserPin sau uid)."})
            continue

        try:
            if inst_by_pin:
                inst = inst_by_pin
            elif lookup:
                inst = Users.objects.get(**lookup)
            else:
                raise Users.DoesNotExist()
            ser = UserSerializer(inst, data=row, partial=True)
            if ser.is_valid():
                if not dry_run:
                    ser.save()
                updated += 1
                results.append({"row": idx, "action": "updated", "UserId": ser.instance.UserId})
            else:
                errors += 1
                results.append({"row": idx, "error": ser.errors})
        except Users.DoesNotExist:
            if not create_if_missing:
                errors += 1
                results.append({"row": idx, "error": "user not found", "lookup": lookup})
                continue
            ser = UserSerializer(data=row)
            if ser.is_valid():
                if not dry_run:
                    obj = ser.save()
                    uid = obj.UserId
                else:
                    uid = None
                created += 1
                results.append({"row": idx, "action": "created", "UserId": uid})
            else:
                errors += 1
                results.append({"row": idx, "error": ser.errors})

    return JsonResponse({
        "updated": updated,
        "created": created,
        "errors": errors,
        "dry_run": dry_run,
        "results": results
    })


# --- PURGE USERS: șterge toți utilizatorii (cu confirmare) ---
@csrf_exempt
def users_purge(request):
    """
    DELETE /api/users/purge/?confirm=DELETE_ALL_USERS
    (sau POST cu { "confirm":"DELETE_ALL_USERS" })
    Opțional: &cascade=1 -> golește și pontajul (PresenceEvent/AttendanceSession)
    """
    if request.method not in ("DELETE", "POST"):
        return JsonResponse({"error": "Only DELETE or POST allowed"}, status=405)

    try:
        body = json.loads(request.body or "{}")
    except Exception:
        body = {}

    confirm = request.GET.get("confirm") or body.get("confirm")
    if confirm != "DELETE_ALL_USERS":
        return JsonResponse(
            {"error": "Confirmare lipsă", "hint": "trimite confirm=DELETE_ALL_USERS"},
            status=400,
        )

    cascade = (request.GET.get("cascade") or body.get("cascade"))
    cascade = str(cascade).lower() in ("1", "true", "yes", "on")

    # dacă vrei să golești și pontajul:
    if cascade:
        from ToolApp.models import AttendanceSession, PresenceEvent  # import local ca să evităm dubluri
        AttendanceSession.objects.all().delete()
        PresenceEvent.objects.all().delete()

    Users.objects.all().delete()
    return JsonResponse({"ok": True})










#api unelte
@csrf_exempt
def toolApi(request,id=0):
    if request.method=='GET':
        tools = Tools.objects.all()
        tools_serializer = ToolSerializer(tools, many=True)
        return JsonResponse(tools_serializer.data, safe=False)

    elif request.method=='POST':
        tool_data=JSONParser().parse(request)
        tool_serializer = ToolSerializer(data=tool_data)
        if tool_serializer.is_valid():
            tool_serializer.save()
            return JsonResponse("Added Successfully!!" , safe=False)
        return JsonResponse("Failed to Add.",safe=False)
    
    elif request.method=='PUT':
        tool_data = JSONParser().parse(request)
        tool=Tools.objects.get(ToolId=tool_data['ToolId'])
        tool_serializer=ToolSerializer(tool,data=tool_data)
        if tool_serializer.is_valid():
            tool_serializer.save()
            return JsonResponse("Updated Successfully!!", safe=False)
        return JsonResponse("Failed to Update.", safe=False)

    elif request.method=='DELETE':
        tool=Tools.objects.get(ToolId=id)
        tool.delete()
        return JsonResponse("Deleted Succeffully!!", safe=False)

#api istoric
@csrf_exempt
def historyApi(request, id=0):
    if request.method == 'GET':
        histories = Histories.objects.all()  # ordering e în Meta: ['-timestamp']
        ser = HistorySerializer(histories, many=True)
        return JsonResponse(ser.data, safe=False)

    elif request.method == 'POST':
        payload = JSONParser().parse(request)
        ser = HistorySerializer(data=payload)
        if ser.is_valid():
            # validări "soft" (nu permitem două OUT la rând)
            tool = ser.validated_data['tool_fk']
            last = Histories.objects.filter(tool_fk=tool).order_by('-timestamp').first()
            direction = ser.validated_data.get('direction', 'OUT')
            if direction == 'OUT' and last and last.direction == 'OUT':
                return JsonResponse({"error": "Unealta este deja predată (OUT)."}, status=400)
            if direction == 'IN' and (not last or last.direction != 'OUT'):
                return JsonResponse({"error": "Nu poți face IN fără un OUT activ."}, status=400)

            obj = ser.save()
            return JsonResponse(HistorySerializer(obj).data, safe=False)
        return JsonResponse(ser.errors, status=400, safe=False)

    elif request.method == 'PUT':
        data = JSONParser().parse(request)
        history = Histories.objects.get(HistoryId=data['HistoryId'])
        ser = HistorySerializer(history, data=data, partial=True)
        if ser.is_valid():
            return JsonResponse(HistorySerializer(ser.save()).data, safe=False)
        return JsonResponse(ser.errors, status=400, safe=False)

    elif request.method == 'DELETE':
        history = Histories.objects.get(HistoryId=id)
        history.delete()
        return JsonResponse("Deleted Successfully!!", safe=False)



#api materiale
@csrf_exempt
def materialApi(request,id=0):
    if request.method=='GET':
        materials = Materials.objects.all()
        materials_serializer = MaterialSerializer(materials, many=True)
        return JsonResponse(materials_serializer.data, safe=False)

    elif request.method=='POST':
        material_data=JSONParser().parse(request)
        material_serializer = MaterialSerializer(data=material_data)
        if material_serializer.is_valid():
            material_serializer.save()
            return JsonResponse("Added Successfully!!" , safe=False)
        return JsonResponse("Failed to Add.",safe=False)
    
    elif request.method=='PUT':
        material_data = JSONParser().parse(request)
        material=Materials.objects.get(MaterialId=material_data['MaterialId'])
        material_serializer=MaterialSerializer(material,data=material_data)
        if material_serializer.is_valid():
            material_serializer.save()
            return JsonResponse("Updated Successfully!!", safe=False)
        return JsonResponse("Failed to Update.", safe=False)

    elif request.method=='DELETE':
        material=Materials.objects.get(MaterialId=id)
        material.delete()
        return JsonResponse("Deleted Succeffully!!", safe=False)


#api consumabile
@csrf_exempt
def consumableApi(request,id=0):
    if request.method=='GET':
        consumables = Consumables.objects.all()
        consumables_serializer = ConsumableSerializer(consumables, many=True)
        return JsonResponse(consumables_serializer.data, safe=False)

    elif request.method=='POST':
        consumable_data=JSONParser().parse(request)
        consumable_serializer = ConsumableSerializer(data=consumable_data)
        if consumable_serializer.is_valid():
            consumable_serializer.save()
            return JsonResponse("Added Successfully!!" , safe=False)
        return JsonResponse("Failed to Add.",safe=False)
    
    elif request.method=='PUT':
        consumable_data = JSONParser().parse(request)
        consumable=Consumables.objects.get(ConsumeId=consumable_data['ConsumeId'])
        consumable_serializer=ConsumableSerializer(consumable,data=consumable_data)
        if consumable_serializer.is_valid():
            consumable_serializer.save()
            return JsonResponse("Updated Successfully!!", safe=False)
        return JsonResponse("Failed to Update.", safe=False)



    elif request.method=='DELETE':
        consumable=Consumables.objects.get(ConsumeId=id)
        consumable.delete()
        return JsonResponse("Deleted Succeffully!!", safe=False)


#api shed
@csrf_exempt
def shedApi(request,id=0):
    if request.method=='GET':
        shed = Shed.objects.all()
        shed_serializer = ShedSerializer(shed, many=True)
        return JsonResponse(shed_serializer.data, safe=False)

    elif request.method=='POST':
        shed_data=JSONParser().parse(request)
        shed_serializer = ShedSerializer(data=shed_data)
        if shed_serializer.is_valid():
            shed_serializer.save()
            return JsonResponse("Added Successfully!!" , safe=False)
        return JsonResponse("Failed to Add.",safe=False)
    
    elif request.method=='PUT':
        shed_data = JSONParser().parse(request)
        shed=Shed.objects.get(ShedId=shed_data['ShedId'])
        shed_serializer=ShedSerializer(shed,data=shed_data)
        if shed_serializer.is_valid():
            shed_serializer.save()
            return JsonResponse("Updated Successfully!!", safe=False)
        return JsonResponse("Failed to Update.", safe=False)


    elif request.method=='DELETE':
        shed=Shed.objects.get(ShedId=id)
        shed.delete()
        return JsonResponse("Deleted Succeffully!!", safe=False)


#api workfield SANTIER
@csrf_exempt
def workfieldApi(request,id=0):
    if request.method=='GET':
        workfield = WorkField.objects.all()
        workfield_serializer = WorkFieldSerializer(workfield, many=True)
        return JsonResponse(workfield_serializer.data, safe=False)

    elif request.method=='POST':
        workfield_data=JSONParser().parse(request)
        workfield_serializer = WorkFieldSerializer(data=workfield_data)
        if workfield_serializer.is_valid():
            workfield_serializer.save()
            return JsonResponse("Added Successfully!!" , safe=False)
        return JsonResponse("Failed to Add.",safe=False)
    
    elif request.method=='PUT':
        workfield_data = JSONParser().parse(request)
        workfield=WorkField.objects.get(WorkFieldId=workfield_data['WorkFieldId'])
        workfield_serializer=WorkFieldSerializer(workfield,data=workfield_data)
        if workfield_serializer.is_valid():
            workfield_serializer.save()
            return JsonResponse("Updated Successfully!!", safe=False)
        return JsonResponse("Failed to Update.", safe=False)

    elif request.method=='DELETE':
        workfield=WorkField.objects.get(WorkFieldId=id)
        workfield.delete()
        return JsonResponse("Deleted Succeffully!!", safe=False)

#api NEFUNCTIONALA 
@csrf_exempt
def unfunctionalApi(request,id=0):
    if request.method=='GET':
        unfunctional = Unfunctional.objects.all()
        unfunctional_serializer = UnfunctionalSerializer(unfunctional, many=True)
        return JsonResponse(unfunctional_serializer.data, safe=False)

    elif request.method=='POST':
        unfunctional_data=JSONParser().parse(request)
        unfunctional_serializer = UnfunctionalSerializer(data=unfunctional_data)
        if unfunctional_serializer.is_valid():
            unfunctional_serializer.save()
            return JsonResponse("Added Successfully!!" , safe=False)
        return JsonResponse("Failed to Add.",safe=False)
    
    elif request.method=='PUT':
        unfunctional_data = JSONParser().parse(request)
        unfunctional=Unfunctional.objects.get(UnfunctionalId=unfunctional_data['UnfunctionalId'])
        unfunctional_serializer=UnfunctionalSerializer(unfunctional,data=unfunctional_data)
        if unfunctional_serializer.is_valid():
            unfunctional_serializer.save()
            return JsonResponse("Updated Successfully!!", safe=False)
        return JsonResponse("Failed to Update.", safe=False)

    elif request.method=='DELETE':
        unfunctional=Unfunctional.objects.get(UnfunctionalId=id)
        unfunctional.delete()
        return JsonResponse("Deleted Succeffully!!", safe=False)

#api cofraj metalic 
@csrf_exempt
def cofrajmetalicApi(request,id=0):
    if request.method=='GET':
        cofrajmetalic = CofrajMetalics.objects.all()
        cofrajmetalic_serializer = CofrajMetalicSerializer(cofrajmetalic, many=True)
        return JsonResponse(cofrajmetalic_serializer.data, safe=False)

    elif request.method=='POST':
        cofrajmetalic_data=JSONParser().parse(request)
        cofrajmetalic_serializer = CofrajMetalicSerializer(data=cofrajmetalic_data)
        if cofrajmetalic_serializer.is_valid():
            cofrajmetalic_serializer.save()
            return JsonResponse("Added Successfully!!" , safe=False)
        return JsonResponse("Failed to Add.",safe=False)
    
    elif request.method=='PUT':
        cofrajmetalic_data = JSONParser().parse(request)
        cofrajmetalic=CofrajMetalics.objects.get(CofrajMetalicId=cofrajmetalic_data['CofrajMetalicId'])
        cofrajmetalic_serializer=CofrajMetalicSerializer(cofrajmetalic,data=cofrajmetalic_data)
        if cofrajmetalic_serializer.is_valid():
            cofrajmetalic_serializer.save()
            return JsonResponse("Updated Successfully!!", safe=False)
        return JsonResponse("Failed to Update.", safe=False)

    elif request.method=='DELETE':
        cofrajmetalic=CofrajMetalics.objects.get(CofrajMetalicId=id)
        cofrajmetalic.delete()
        return JsonResponse("Deleted Succeffully!!", safe=False)

    
#api cofraj tip doka 
@csrf_exempt
def cofrajtipdokaApi(request,id=0):
    if request.method=='GET':
        cofrajtipdoka = CofrajtTipDokas.objects.all()
        cofrajtipdoka_serializer = CofrajtTipDokaSerializer(cofrajtipdoka, many=True)
        return JsonResponse(cofrajtipdoka_serializer.data, safe=False)

    elif request.method=='POST':
        cofrajtipdoka_data=JSONParser().parse(request)
        cofrajtipdoka_serializer = CofrajtTipDokaSerializer(data=cofrajtipdoka_data)
        if cofrajtipdoka_serializer.is_valid():
            cofrajtipdoka_serializer.save()
            return JsonResponse("Added Successfully!!" , safe=False)
        return JsonResponse("Failed to Add.",safe=False)
    
    elif request.method=='PUT':
        cofrajtipdoka_data = JSONParser().parse(request)
        cofrajtipdoka=CofrajtTipDokas.objects.get(CofrajtTipDokaId=cofrajtipdoka_data['CofrajtTipDokaId'])
        cofrajtipdoka_serializer=CofrajtTipDokaSerializer(cofrajtipdoka,data=cofrajtipdoka_data)
        if cofrajtipdoka_serializer.is_valid():
            cofrajtipdoka_serializer.save()
            return JsonResponse("Updated Successfully!!", safe=False)
        return JsonResponse("Failed to Update.", safe=False)

    elif request.method=='DELETE':
        cofrajtipdoka=CofrajtTipDokas.objects.get(CofrajtTipDokaId=id)
        cofrajtipdoka.delete()
        return JsonResponse("Deleted Succeffully!!", safe=False)


#api Popi
@csrf_exempt
def popiApi(request,id=0):
    if request.method=='GET':
        popi = Popis.objects.all()
        popi_serializer = PopiSerializer(popi, many=True)
        return JsonResponse(popi_serializer.data, safe=False)

    elif request.method=='POST':
        popi_data=JSONParser().parse(request)
        popi_serializer = PopiSerializer(data=popi_data)
        if popi_serializer.is_valid():
            popi_serializer.save()
            return JsonResponse("Added Successfully!!" , safe=False)
        return JsonResponse("Failed to Add.",safe=False)
    
    elif request.method=='PUT':
        popi_data = JSONParser().parse(request)
        popi=Popis.objects.get(PopiDokaId=popi_data['PopiDokaId'])
        popi_serializer=PopiSerializer(popi,data=popi_data)
        if popi_serializer.is_valid():
            popi_serializer.save()
            return JsonResponse("Updated Successfully!!", safe=False)
        return JsonResponse("Failed to Update.", safe=False)

    elif request.method=='DELETE':
        popi=Popis.objects.get(PopiDokaId=id)
        popi.delete()
        return JsonResponse("Deleted Succeffully!!", safe=False)


#api schela usoara
@csrf_exempt
def schelausoaraApi(request,id=0):
    if request.method=='GET':
        schelausoara = SchelaUsoaras.objects.all()
        schelausoara_serializer = SchelaUsoaraSerializer(schelausoara, many=True)
        return JsonResponse(schelausoara_serializer.data, safe=False)

    elif request.method=='POST':
        schelausoara_data=JSONParser().parse(request)
        schelausoara_serializer = SchelaUsoaraSerializer(data=schelausoara_data)
        if schelausoara_serializer.is_valid():
            schelausoara_serializer.save()
            return JsonResponse("Added Successfully!!" , safe=False)
        return JsonResponse("Failed to Add.",safe=False)
    
    elif request.method=='PUT':
        schelausoara_data = JSONParser().parse(request)
        schelausoara=SchelaUsoaras.objects.get(SchelaUsoaraId=schelausoara_data['SchelaUsoaraId'])
        schelausoara_serializer=SchelaUsoaraSerializer(schelausoara,data=schelausoara_data)
        if schelausoara_serializer.is_valid():
            schelausoara_serializer.save()
            return JsonResponse("Updated Successfully!!", safe=False)
        return JsonResponse("Failed to Update.", safe=False)

    elif request.method=='DELETE':
        schelausoara=SchelaUsoaras.objects.get(SchelaUsoaraId=id)
        schelausoara.delete()
        return JsonResponse("Deleted Succeffully!!", safe=False)




#api Schela Fatada 
@csrf_exempt
def schelafatadaApi(request,id=0):
    if request.method=='GET':
        schelafatada = SchelaFatadas.objects.all()
        schelafatada_serializer = SchelaFatadaSerializer(schelafatada, many=True)
        return JsonResponse(schelafatada_serializer.data, safe=False)

    elif request.method=='POST':
        schelafatada_data=JSONParser().parse(request)
        schelafatada_serializer = SchelaFatadaSerializer(data=schelafatada_data)
        if schelafatada_serializer.is_valid():
            schelafatada_serializer.save()
            return JsonResponse("Added Successfully!!" , safe=False)
        return JsonResponse("Failed to Add.",safe=False)
    
    elif request.method=='PUT':
        schelafatada_data = JSONParser().parse(request)
        schelafatada=SchelaFatadas.objects.get(SchelaFatadaId=schelafatada_data['SchelaFatadaId'])
        schelafatada_serializer=SchelaFatadaSerializer(schelafatada,data=schelafatada_data)
        if schelafatada_serializer.is_valid():
            schelafatada_serializer.save()
            return JsonResponse("Updated Successfully!!", safe=False)
        return JsonResponse("Failed to Update.", safe=False)

    elif request.method=='DELETE':
        schelafatada=SchelaFatadas.objects.get(SchelaFatadaId=id)
        schelafatada.delete()
        return JsonResponse("Deleted Succeffully!!", safe=False)



#api Schela Fatada modulara
@csrf_exempt
def schelafatadamodularaApi(request,id=0):
    if request.method=='GET':
        schelafatadamodulara = SchelaFatadaModularas.objects.all()
        schelafatadamodulara_serializer = SchelaFatadaModularaSerializer(schelafatadamodulara, many=True)
        return JsonResponse(schelafatadamodulara_serializer.data, safe=False)

    elif request.method=='POST':
        schelafatadamodulara_data=JSONParser().parse(request)
        schelafatadamodulara_serializer = SchelaFatadaModularaSerializer(data=schelafatadamodulara_data)
        if schelafatadamodulara_serializer.is_valid():
            schelafatadamodulara_serializer.save()
            return JsonResponse("Added Successfully!!" , safe=False)
        return JsonResponse("Failed to Add.",safe=False)
    
    elif request.method=='PUT':
        schelafatadamodulara_data = JSONParser().parse(request)
        schelafatadamodulara=SchelaFatadaModularas.objects.get(SchelaFatadaModularaId=schelafatadamodulara_data['SchelaFatadaModularaId'])
        schelafatadamodulara_serializer=SchelaFatadaModularaSerializer(schelafatadamodulara,data=schelafatadamodulara_data)
        if schelafatadamodulara_serializer.is_valid():
            schelafatadamodulara_serializer.save()
            return JsonResponse("Updated Successfully!!", safe=False)
        return JsonResponse("Failed to Update.", safe=False)

    elif request.method=='DELETE':
        schelafatadamodulara=SchelaFatadaModularas.objects.get(SchelaFatadaModularaId=id)
        schelafatadamodulara.delete()
        return JsonResponse("Deleted Succeffully!!", safe=False)



#api combustibil
@csrf_exempt
def combustibilApi(request,id=0):
    if request.method=='GET':
        combustibil = Combustibils.objects.all()
        combustibil_serializer = CombustibilSerializer(combustibil, many=True)
        return JsonResponse(combustibil_serializer.data, safe=False)

    elif request.method=='POST':
        combustibil_data=JSONParser().parse(request)
        combustibil_serializer = CombustibilSerializer(data=combustibil_data)
        if combustibil_serializer.is_valid():
            combustibil_serializer.save()
            return JsonResponse("Added Successfully!!" , safe=False)
        return JsonResponse("Failed to Add.",safe=False)
    
    elif request.method=='PUT':
        combustibil_data = JSONParser().parse(request)
        combustibil=Combustibils.objects.get(CombustibilId=combustibil_data['CombustibilId'])
        combustibil_serializer=CombustibilSerializer(combustibil,data=combustibil_data)
        if combustibil_serializer.is_valid():
            combustibil_serializer.save()
            return JsonResponse("Updated Successfully!!", safe=False)
        return JsonResponse("Failed to Update.", safe=False)

    elif request.method=='DELETE':
        combustibil=Combustibils.objects.get(CombustibilId=id)
        combustibil.delete()
        return JsonResponse("Deleted Succeffully!!", safe=False)


#api istoric schela
@csrf_exempt
def istoricschelaApi(request,id=0):
    if request.method=='GET':
        istoricschele = HistorieScheles.objects.all()
        istoricschele_serializer = HistorieScheleSerializer(istoricschele, many=True)
        return JsonResponse(istoricschele_serializer.data, safe=False)

    elif request.method=='POST':
        istoricschele_data=JSONParser().parse(request)
        istoricschele_serializer = HistorieScheleSerializer(data=istoricschele_data)
        if istoricschele_serializer.is_valid():
            istoricschele_serializer.save()
            return JsonResponse("Added Successfully!!" , safe=False)
        return JsonResponse("Failed to Add.",safe=False)
    
    elif request.method=='PUT':
        istoricschele_data = JSONParser().parse(request)
        istoricschele=HistorieScheles.objects.get(HistoriesScheleId=istoricschele_data['HistoriesScheleId'])
        istoricschele_serializer=HistorieScheleSerializer(istoricschele,data=istoricschele_data)
        if istoricschele_serializer.is_valid():
            istoricschele_serializer.save()
            return JsonResponse("Updated Successfully!!", safe=False)
        return JsonResponse("Failed to Update.", safe=False)

    elif request.method=='DELETE':
        istoricschele=HistorieScheles.objects.get(HistoriesScheleId=id)
        istoricschele.delete()
        return JsonResponse("Deleted Succeffully!!", safe=False)


#api mijloace fixe
@csrf_exempt
def mijloacefixeApi(request,id=0):
    if request.method=='GET':
        mijloacefixe = MijloaceFixes.objects.all()
        mijloacefixe_serializer = MijloaceFixeSerializer(mijloacefixe, many=True)
        return JsonResponse(mijloacefixe_serializer.data, safe=False)

    elif request.method=='POST':
        mijloacefixe_data=JSONParser().parse(request)
        mijloacefixe_serializer = MijloaceFixeSerializer(data=mijloacefixe_data)
        if mijloacefixe_serializer.is_valid():
            mijloacefixe_serializer.save()
            return JsonResponse("Added Successfully!!" , safe=False)
        return JsonResponse("Failed to Add.",safe=False)
    
    elif request.method=='PUT':
        mijloacefixe_data = JSONParser().parse(request)
        mijloacefixe=MijloaceFixes.objects.get(MijloaceFixeId=mijloacefixe_data['MijloaceFixeId'])
        mijloacefixe_serializer=MijloaceFixeSerializer(mijloacefixe,data=mijloacefixe_data)
        if mijloacefixe_serializer.is_valid():
            mijloacefixe_serializer.save()
            return JsonResponse("Updated Successfully!!", safe=False)
        return JsonResponse("Failed to Update.", safe=False)

    elif request.method=='DELETE':
        mijloacefixe=MijloaceFixes.objects.get(MijloaceFixeId=id)
        mijloacefixe.delete()
        return JsonResponse("Deleted Succeffully!!", safe=False)

#test nfc reader
def nfc_tag_view(request):
    with nfc.ContactlessFrontend('usb') as clf:
        target = clf.sense(nfc.clf.RemoteTarget('iso14443a'))
        tag = nfc.tag.activate(clf, target)
        text_record = tag.ndef.records[0].text
    return JsonResponse({'text_record': text_record})



#RDIF CHIP MAGAZIE 
@csrf_exempt
def rfid_entry_exit(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body or "{}")
            event = data.get("event")
            # aici poți face mapare către issue/return dacă vrei
            return JsonResponse({"status": "ok", "received": event})
        except Exception as e:
            return JsonResponse({"status": "error", "msg": str(e)}, status=400)
    return JsonResponse({"msg": "Only POST allowed"}, status=405)



#MAGAIE UNELTE - ISSUE / RETURN
@csrf_exempt
def issue_tool(request):
    if request.method != 'POST':
        return JsonResponse({"msg": "Only POST allowed"}, status=405)
    data = JSONParser().parse(request)
    # forțăm OUT
    data = {**data, "direction": "OUT"}
    ser = HistorySerializer(data=data)
    if ser.is_valid():
        tool = ser.validated_data['tool_fk']
        last = Histories.objects.filter(tool_fk=tool).order_by('-timestamp').first()
        if last and last.direction == 'OUT':
            return JsonResponse({"error": "Unealta este deja predată (OUT)."}, status=400)
        obj = ser.save()
        return JsonResponse(HistorySerializer(obj).data, safe=False)
    return JsonResponse(ser.errors, status=400, safe=False)

#-- MAGAZIE UNELTE - RETURN ---
@csrf_exempt
def return_tool(request):
    if request.method != 'POST':
        return JsonResponse({"msg": "Only POST allowed"}, status=405)
    data = JSONParser().parse(request)
    data = {**data, "direction": "IN"}
    ser = HistorySerializer(data=data)
    if ser.is_valid():
        tool = ser.validated_data['tool_fk']
        last = Histories.objects.filter(tool_fk=tool).order_by('-timestamp').first()
        if not last or last.direction != 'OUT':
            return JsonResponse({"error": "Unealta nu este în OUT; nu poți face IN."}, status=400)
        obj = ser.save()
        return JsonResponse(HistorySerializer(obj).data, safe=False)
    return JsonResponse(ser.errors, status=400, safe=False)


#-- MAGAZIE UNELTE - STATUS TOOLS ---
@csrf_exempt
def tools_status(request):
    if request.method != 'GET':
        return JsonResponse({"msg": "Only GET allowed"}, status=405)

    data = []
    tools = Tools.objects.all()
    for t in tools:
        last = Histories.objects.filter(tool_fk=t).order_by('-timestamp').first()
        state = "IN"
        holder = None
        ts = None
        if last:
            ts = last.timestamp.isoformat()
            if last.direction == "OUT":
                state = "OUT"
                if last.user_fk:
                    holder = {
                        "UserId": last.user_fk.UserId,
                        "UserName": last.user_fk.UserName,
                        "UserSerie": last.user_fk.UserSerie,
                    }
        data.append({
            "ToolId": t.ToolId,
            "ToolSerie": t.ToolSerie,
            "ToolName": t.ToolName,
            "status": state,
            "last_movement_at": ts,
            "holder": holder
        })
    return JsonResponse(data, safe=False)




#--- RFID/Senzor intrare/ieșire magazie & preluare/predare unealtă ---
@csrf_exempt
def sensor_event(request):
    if request.method != 'POST':
        return JsonResponse({"msg": "Only POST allowed"}, status=405)

    try:
        data = json.loads(request.body or "{}")
        ev_type = (data.get("type") or "").lower()
        user_serie = (data.get("user_serie") or "").strip()
        tool_serie = (data.get("tool_serie") or "").strip() if data.get("tool_serie") else None
        qty = int(data.get("quantity") or 1)
        note = data.get("note") or "RFID/Senzor"
        ws = (data.get("worksite") or data.get("site") or data.get("santier") or "").strip() or None

        if not user_serie:
            return JsonResponse({"error": "user_serie este obligatoriu"}, status=400)

        try:
            user = Users.objects.get(UserSerie=user_serie)
        except Users.DoesNotExist:
            return JsonResponse({"error": f"User cu seria '{user_serie}' nu există"}, status=404)

        if ev_type in ("enter", "exit"):
            kind = PresenceEvent.Kind.ENTER if ev_type == "enter" else PresenceEvent.Kind.EXIT
            PresenceEvent.objects.create(user_fk=user, kind=kind, timestamp=timezone.now(), worksite=ws)
            msg = f"{user.UserName} a intrat în magazie" if ev_type == "enter" else f"{user.UserName} a ieșit din magazie"
            display = {
                "User": user.UserName,
                "Tool": "",
                "GiveRecive": "intrare" if ev_type == "enter" else "ieșire",
                "DateOfGiving": timezone.localtime().date().isoformat(),
                "timestamp": timezone.localtime().isoformat(),
                "worksite": ws,
            }
            return JsonResponse({"ok": True, "message": msg, "display": display})

        if ev_type not in ("tool_out", "tool_in"):
            return JsonResponse({"error": "type necunoscut. Folosește: enter | exit | tool_out | tool_in"}, status=400)

        if not tool_serie:
            return JsonResponse({"error": "tool_serie este obligatoriu pentru tool_out/tool_in"}, status=400)

        try:
            tool = Tools.objects.get(ToolSerie=tool_serie)
        except Tools.DoesNotExist:
            return JsonResponse({"error": f"Unealtă cu seria '{tool_serie}' nu există"}, status=404)

        direction = "OUT" if ev_type == "tool_out" else "IN"

        with transaction.atomic():
            last = Histories.objects.filter(tool_fk=tool).order_by('-timestamp').select_for_update().first()
            if direction == "OUT" and last and last.direction == "OUT":
                return JsonResponse({"error": "Unealta este deja în afara magaziei (OUT)."}, status=409)
            if direction == "IN" and (not last or last.direction != "OUT"):
                return JsonResponse({"error": "Nu poți face IN fără un OUT activ."}, status=409)

            payload = {
                "user_serie": user_serie,
                "tool_serie": tool_serie,
                "direction": direction,
                "quantity": max(qty, 1),
                "note": note,
            }
            ser = HistorySerializer(data=payload)
            if not ser.is_valid():
                return JsonResponse(ser.errors, status=400, safe=False)

            obj = ser.save()

        msg = f"{user.UserName} a {'preluat' if direction=='OUT' else 'predat'} {tool.ToolName}"
        return JsonResponse({"ok": True, "message": msg, "history": HistorySerializer(obj).data})

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)





logger = logging.getLogger(__name__)

_last_seen = {}
_DEBOUNCE_SEC = 0.7  # server-side debounce

#--- NFC SCAN: la EXIT suprascrie worksite dacă vine în payload ---
def _fmt_hms(seconds: int):
    seconds = max(0, int(seconds))
    h = seconds // 3600
    m = (seconds % 3600) // 60
    s = seconds % 60
    return f"{h:02d}:{m:02d}:{s:02d}"


# Config rapid — schimbă după programul tău
PONTAJ_SHIFT_END_HOUR = 18   # 18:00 ora locală = sfârșitul zilei
PONTAJ_MAX_SHIFT_HOURS = 14  # limită de siguranță (cap durată sesiune)


#--- NFC SCAN: la EXIT suprascrie worksite dacă vine în payload ---
def workday_close_dt(local_day):
    """Ora de închidere a zilei (tz-aware) pentru o dată locală."""
    tz = timezone.get_current_timezone()
    naive = datetime(local_day.year, local_day.month, local_day.day, PONTAJ_SHIFT_END_HOUR, 0, 0)
    return timezone.make_aware(naive, tz)


def _is_manual_attendance_scan(uid, tag_type):
    return str(uid or "").upper().strip() == "MANUAL" or str(tag_type or "").strip().lower() == "manual"


def _get_client_ip(request):
    forwarded_for = (request.META.get("HTTP_X_FORWARDED_FOR") or "").strip()
    if forwarded_for:
        return forwarded_for.split(",")[0].strip()[:64]

    real_ip = (request.META.get("HTTP_X_REAL_IP") or "").strip()
    if real_ip:
        return real_ip[:64]

    return (request.META.get("REMOTE_ADDR") or "").strip()[:64] or None


def _normalize_manual_device_key(raw_value):
    value = str(raw_value or "").strip()
    return value[:64] or None


def _extract_pin_candidate(raw_content):
    content = str(raw_content or "").strip()
    if not content:
        return ""

    if re.fullmatch(r"\d{3,8}", content):
        return content

    match = re.search(r"\b(?:pin|cod|code)\D{0,12}(\d{3,8})\b", content, flags=re.IGNORECASE)
    return match.group(1) if match else ""


PIN_RATE_LIMIT_MAX_FAILURES = 5
PIN_RATE_LIMIT_WINDOW_SECONDS = 15 * 60
PIN_RATE_LIMIT_BLOCK_SECONDS = 15 * 60


def _pin_rate_identity(request, device_key=None, uid=None):
    ip = _get_client_ip(request) or ""
    device = _normalize_manual_device_key(device_key) or ""
    uid_part = str(uid or "").upper().strip()[:64]
    return ip, device, f"pin:{ip}:{device}:{uid_part}"


def _pin_is_blocked(request, device_key=None, uid=None):
    ip, device, base_key = _pin_rate_identity(request, device_key=device_key, uid=uid)
    blocked_until = cache.get(f"{base_key}:blocked_until")
    if blocked_until and blocked_until > timezone.now().timestamp():
        PinAttemptLog.objects.create(
            ip_address=ip,
            device_key=device,
            uid=str(uid or "")[:128],
            success=False,
            blocked=True,
            reason="temporarily_blocked",
        )
        return True, int(blocked_until - timezone.now().timestamp())
    return False, 0


def _log_pin_attempt(request, *, success, reason, device_key=None, uid=None, worksite=None):
    ip, device, base_key = _pin_rate_identity(request, device_key=device_key, uid=uid)
    PinAttemptLog.objects.create(
        ip_address=ip,
        device_key=device,
        uid=str(uid or "")[:128],
        worksite=str(worksite or "")[:100],
        success=bool(success),
        blocked=False,
        reason=str(reason or "")[:128],
    )

    if success:
        cache.delete(f"{base_key}:failures")
        cache.delete(f"{base_key}:blocked_until")
        return

    failures = cache.get(f"{base_key}:failures") or 0
    failures += 1
    cache.set(f"{base_key}:failures", failures, PIN_RATE_LIMIT_WINDOW_SECONDS)
    if failures >= PIN_RATE_LIMIT_MAX_FAILURES:
        cache.set(
            f"{base_key}:blocked_until",
            timezone.now().timestamp() + PIN_RATE_LIMIT_BLOCK_SECONDS,
            PIN_RATE_LIMIT_BLOCK_SECONDS,
        )


PIN_CHECK_WORKERS = 12
PIN_LOGIN_CACHE_SECONDS = 24 * 3600


def _pin_cache_key(raw_pin):
    digest = hashlib.sha256(str(raw_pin or "").strip().encode("utf-8")).hexdigest()
    return f"pin-login-user:{digest}"


def _check_pin_row(raw_pin, row):
    user_id, pin_hash = row
    if pin_hash and check_password(raw_pin, pin_hash):
        return user_id
    return None


def _find_user_by_pin(raw_pin):
    raw_pin = str(raw_pin or "").strip()
    if not raw_pin:
        return None

    cached_user_id = cache.get(_pin_cache_key(raw_pin))
    if cached_user_id:
        cached_user = Users.objects.filter(UserId=cached_user_id).first()
        if cached_user:
            return cached_user

    legacy_user = Users.objects.filter(UserPin=raw_pin).first()
    if legacy_user:
        legacy_user.set_pin(raw_pin)
        legacy_user.save(update_fields=["UserPin", "pin_hash"])
        cache.set(_pin_cache_key(raw_pin), legacy_user.UserId, PIN_LOGIN_CACHE_SECONDS)
        return legacy_user

    rows = list(
        Users.objects
        .exclude(pin_hash="")
        .values_list("UserId", "pin_hash")
    )
    if not rows:
        return None

    executor = ThreadPoolExecutor(max_workers=min(PIN_CHECK_WORKERS, len(rows)))
    try:
        futures = [executor.submit(_check_pin_row, raw_pin, row) for row in rows]
        for future in as_completed(futures):
            user_id = future.result()
            if user_id:
                for pending in futures:
                    pending.cancel()
                cache.set(_pin_cache_key(raw_pin), user_id, PIN_LOGIN_CACHE_SECONDS)
                executor.shutdown(wait=False, cancel_futures=True)
                return Users.objects.filter(UserId=user_id).first()
    finally:
        executor.shutdown(wait=False, cancel_futures=True)

    return None


def _client_owns_manual_session(session, client_ip=None, device_key=None):
    if not session:
        return False

    if not session.manual_device_key and not session.manual_client_ip:
        return True

    if device_key and session.manual_device_key:
        return session.manual_device_key == device_key

    if client_ip and session.manual_client_ip:
        return session.manual_client_ip == client_ip

    return False


def _parse_float_or_none(value):
    if value in (None, ""):
        return None

    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _extract_gps_payload(data):
    gps = data.get("gps") if isinstance(data.get("gps"), dict) else {}

    lat = _parse_float_or_none(gps.get("lat") if gps else data.get("gps_lat") or data.get("latitude") or data.get("lat"))
    lng = _parse_float_or_none(gps.get("lng") if gps else data.get("gps_lng") or data.get("longitude") or data.get("lng"))
    accuracy = _parse_float_or_none(gps.get("accuracy") if gps else data.get("gps_accuracy") or data.get("accuracy"))

    if lat is None or lng is None:
        return None

    if lat < -90 or lat > 90 or lng < -180 or lng > 180:
        return None

    return {
        "lat": lat,
        "lng": lng,
        "accuracy": accuracy,
    }


def _extract_gps_captured_at(data):
    gps = data.get("gps") if isinstance(data.get("gps"), dict) else {}
    raw_value = gps.get("captured_at") if gps else data.get("gps_captured_at")
    return _parse_client_ts(raw_value)


def _session_gps_payload(session):
    return {
        "in_gps": (
            {
                "lat": session.in_gps_latitude,
                "lng": session.in_gps_longitude,
                "accuracy": session.in_gps_accuracy_m,
            }
            if session.in_gps_latitude is not None and session.in_gps_longitude is not None
            else None
        ),
        "out_gps": (
            {
                "lat": session.out_gps_latitude,
                "lng": session.out_gps_longitude,
                "accuracy": session.out_gps_accuracy_m,
            }
            if session.out_gps_latitude is not None and session.out_gps_longitude is not None
            else None
        ),
    }



# --- NFC SCAN: la EXIT suprascrie worksite dacă vine în payload ---

@csrf_exempt
def nfc_scan(request):
    if request.method != "POST":
        return JsonResponse({"error": "Only POST allowed"}, status=405)

    try:
        data = json.loads(request.body.decode("utf-8"))
    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    uid = str(data.get("uid") or "").upper().strip()
    tag_type = str(data.get("tag_type") or "").strip()
    content = str(data.get("content") or "").strip()  # PIN de pe tag (dacă există)

    # log brut
    print(f"[NFC] RAW data: uid={uid} type={tag_type} content='{content}'")

    # Folosim timestamp-ul clientului dacă e valid, altfel now()
    client_when = _parse_client_ts(data.get("timestamp"))
    when = client_when or timezone.now()
    # Acceptăm mai multe chei: worksite/site/santier
    ws = (data.get("worksite") or data.get("site") or data.get("santier") or "").strip() or None
    attendance_mode = (data.get("mode") or data.get("attendance_mode") or "").strip().lower() or None
    is_manual_scan = _is_manual_attendance_scan(uid, tag_type)
    manual_client_ip = _get_client_ip(request) if is_manual_scan else None
    manual_device_key = _normalize_manual_device_key(data.get("device_key")) if is_manual_scan else None
    gps_payload = _extract_gps_payload(data)
    gps_captured_at = _extract_gps_captured_at(data)

    if attendance_mode == "driver" and not gps_payload:
        return JsonResponse({
            "error": "Locatia GPS este obligatorie pentru pontajul soferilor.",
            "error_code": "GPS_REQUIRED_FOR_DRIVER"
        }, status=400)

    if gps_payload and gps_captured_at and gps_captured_at < (timezone.now() - timedelta(minutes=10)):
        return JsonResponse({
            "error": "Locatia GPS salvata a expirat. Cere din nou locatia si introdu PIN-ul in maximum 10 minute.",
            "error_code": "GPS_CAPTURE_EXPIRED"
        }, status=400)

    # --- DETERMINĂM PINUL EFECTIV: din content sau din maparea UID_PIN_MAP ---
    raw_pin = _extract_pin_candidate(content)
    mapped_pin = None
    mapped_user = None

    if raw_pin:
        # avem PIN direct de pe tag
        print(f"[NFC-MAP] UID={uid} a trimis PIN direct de pe tag: {raw_pin}")
    else:
        if content:
            print(f"[NFC-MAP] UID={uid} a trimis content non-PIN; încerc mapare UID. content='{content[:80]}'")
        # nu avem PIN pe tag -> încercăm mapare UID -> PIN
        mapped_pin = UID_PIN_MAP.get(uid)
        if mapped_pin:
            print(f"[NFC-MAP] UID={uid} mapat din UID_PIN_MAP la PIN={mapped_pin}")
        else:
            mapped_user = Users.objects.filter(uid__iexact=uid).first() if uid and uid != "MANUAL" else None
            if mapped_user:
                print(
                    f"[NFC-MAP] UID={uid} găsit direct în Users.uid pentru user={mapped_user.UserName} "
                    f"(id={mapped_user.UserId})"
                )
            else:
                print(f"[NFC-MAP] UID={uid} NU există nici în UID_PIN_MAP, nici în Users.uid")

    effective_pin = mapped_pin or raw_pin or ""

    blocked, retry_after = _pin_is_blocked(request, device_key=data.get("device_key"), uid=uid)
    if blocked:
        return JsonResponse({
            "error": "Prea multe incercari gresite. Incearca din nou mai tarziu.",
            "error_code": "PIN_TEMPORARILY_BLOCKED",
            "retry_after_seconds": retry_after,
        }, status=429)

    # Debounce identic (server-side) – folosim PINUL EFECTIV, nu contentul brut
    now_ts = timezone.now().timestamp()
    debounce_identity = effective_pin or (f"user:{mapped_user.UserId}" if mapped_user else "")
    key = (uid, debounce_identity)
    if key in _last_seen and (now_ts - _last_seen[key]) < _DEBOUNCE_SEC:
        print(f"[NFC] DEBOUNCED scan pentru UID={uid} pin={effective_pin}")
        return JsonResponse({"ok": True, "debounced": True})
    _last_seen[key] = now_ts

    # 1) Identific user după PIN EFECTIV
    if not effective_pin and not mapped_user:
        # nici PIN pe tag, nici în mapare -> nu avem cum găsi user
        print(f"[NFC] {timezone.now()} UID={uid} type={tag_type} pin_scanned='' -> no user match (no PIN, no mapping)")
        _log_pin_attempt(request, success=False, reason="missing_pin", device_key=data.get("device_key"), uid=uid, worksite=ws)
        return JsonResponse({"ok": True, "match": None, "received": data})

    user = mapped_user or _find_user_by_pin(effective_pin)
    if not user:
        print(f"[NFC] {timezone.now()} UID={uid} type={tag_type} pin_scanned={effective_pin} -> no user match")
        _log_pin_attempt(request, success=False, reason="invalid_pin", device_key=data.get("device_key"), uid=uid, worksite=ws)
        return JsonResponse({"ok": True, "match": None, "received": data})

    _log_pin_attempt(request, success=True, reason="ok", device_key=data.get("device_key"), uid=uid, worksite=ws)

    # avem user -> log ca să vezi clar
    print(
        f"[NFC] {timezone.now()} UID={uid} type={tag_type} "
        f"pin_scanned={effective_pin} -> user={user.UserName} (id={user.UserId})"
    )

    # work_date derivat din 'when', nu din 'now'
    today = localdate(when)

    with transaction.atomic():
        open_sess = (AttendanceSession.objects
                     .select_for_update()
                     .filter(user_fk=user, out_time__isnull=True)
                     .exclude(source__contains=MISSING_EXIT_SOURCE_TAG)
                     .order_by('-in_time')
                     .first())

        if is_manual_scan:
            matching_client_session = None
            if manual_device_key:
                matching_client_session = (AttendanceSession.objects
                                           .select_for_update()
                                           .filter(out_time__isnull=True, work_date=today, manual_device_key=manual_device_key)
                                           .order_by('-in_time')
                                           .first())
            elif manual_client_ip:
                matching_client_session = (AttendanceSession.objects
                                           .select_for_update()
                                           .filter(out_time__isnull=True, work_date=today, manual_client_ip=manual_client_ip)
                                           .order_by('-in_time')
                                           .first())

            if matching_client_session and matching_client_session.user_fk_id != user.UserId:
                return JsonResponse({
                    "error": "Telefonul sau browserul acesta are deja un check-in activ pentru alt muncitor. Pe acest dispozitiv poti face acum doar check-out pentru muncitorul pontat primul.",
                    "error_code": "MANUAL_DEVICE_LOCKED"
                }, status=409)

            if open_sess and open_sess.work_date == today and not _client_owns_manual_session(
                open_sess,
                client_ip=manual_client_ip,
                device_key=manual_device_key
            ):
                return JsonResponse({
                    "error": "Check-out-ul pentru acest muncitor trebuie facut de pe acelasi telefon sau browser folosit la check-in.",
                    "error_code": "MANUAL_CHECKOUT_DEVICE_MISMATCH"
                }, status=409)

        if open_sess:
            # dacă sesiunea e din altă zi, o închidem la ora de închidere a zilei anterioare
            if open_sess.work_date < today:
                _mark_session_missing_exit(open_sess)
                open_sess.save(update_fields=["out_time", "duration_seconds", "source"])

                _publish("auto_close", user, when, {
                    "closed_at_hm": None,
                    "duration_hms": _fmt_hms(open_sess.duration_seconds),
                    "work_date": str(open_sess.work_date),
                    "worksite": open_sess.worksite,
                    "missing_exit": True,
                })

                # deschidem sesiune nouă la 'when', cu worksite curent (ws)
                new_sess = AttendanceSession.objects.create(
                    user_fk=user,
                    work_date=today,
                    in_time=apply_enter_grace(when),
                    in_gps_latitude=gps_payload["lat"] if gps_payload else None,
                    in_gps_longitude=gps_payload["lng"] if gps_payload else None,
                    in_gps_accuracy_m=gps_payload["accuracy"] if gps_payload else None,
                    source="manual-driver" if attendance_mode == "driver" else ("manual-web" if is_manual_scan else "nfc"),
                    worksite=ws,
                    manual_client_ip=manual_client_ip if is_manual_scan else None,
                    manual_device_key=manual_device_key if is_manual_scan else None,
                )
                _publish("enter", user, when, {"worksite": ws})
                PresenceEvent.objects.create(
                    user_fk=user, kind=PresenceEvent.Kind.ENTER,
                    timestamp=when, worksite=ws
                )

                print(f"[NFC] AUTO-CLOSE + ENTER nou pentru {user.UserName} pe {today} la site='{ws}'")

                return JsonResponse({
                    "ok": True,
                    "state": "ENTER",
                    "auto_closed_previous": {
                        "work_date": str(open_sess.work_date),
                        "closed_at": None,
                        "duration_hms": _fmt_hms(open_sess.duration_seconds),
                        "session_id": open_sess.id,
                        "worksite": open_sess.worksite,
                        "missing_exit": True,
                    },
                    "user": {"id": user.UserId, "name": user.UserName},
                    "session": {
                        "work_date": str(new_sess.work_date),
                        "in_time": localtime(new_sess.in_time).isoformat(),
                        "worksite": ws,
                        **_session_gps_payload(new_sess),
                    },
                    "received": data,
                    "ts_used": "client" if client_when else "server"
                })

            # aceeași zi -> EXIT la 'when'
            # dacă am primit ws, îl considerăm „locația finală” a zilei/sesiunii
            if ws:
                open_sess.worksite = ws

            open_sess.out_time = apply_exit_grace(when)
            if gps_payload:
                open_sess.out_gps_latitude = gps_payload["lat"]
                open_sess.out_gps_longitude = gps_payload["lng"]
                open_sess.out_gps_accuracy_m = gps_payload["accuracy"]
            open_sess.duration_seconds = max(0, int((open_sess.out_time - open_sess.in_time).total_seconds()))
            open_sess.save()
            recompute_daily_pay(user, open_sess.work_date)

            _publish("exit", user, open_sess.out_time, {
                "duration_hms": _fmt_hms(open_sess.duration_seconds),
                "worksite": open_sess.worksite,
            })
            PresenceEvent.objects.create(
                user_fk=user, kind=PresenceEvent.Kind.EXIT,
                timestamp=open_sess.out_time, worksite=open_sess.worksite
            )

            print(f"[NFC] EXIT pentru {user.UserName} pe {open_sess.work_date} la site='{open_sess.worksite}'")

            return JsonResponse({
                "ok": True,
                "state": "EXIT",
                "user": {"id": user.UserId, "name": user.UserName},
                "session": {
                    "work_date": str(open_sess.work_date),
                    "in_time": localtime(open_sess.in_time).isoformat(),
                    "out_time": localtime(open_sess.out_time).isoformat(),
                    "duration_hms": _fmt_hms(open_sess.duration_seconds),
                    "worksite": open_sess.worksite,
                    **_session_gps_payload(open_sess),
                },
                "received": data,
                "ts_used": "client" if client_when else "server"
            })

        # ENTER: sesiune nouă la 'when' cu worksite
        sess = AttendanceSession.objects.create(
            user_fk=user,
            work_date=today,
            in_time=apply_enter_grace(when),
            in_gps_latitude=gps_payload["lat"] if gps_payload else None,
            in_gps_longitude=gps_payload["lng"] if gps_payload else None,
            in_gps_accuracy_m=gps_payload["accuracy"] if gps_payload else None,
            source="manual-driver" if attendance_mode == "driver" else ("manual-web" if is_manual_scan else "nfc"),
            worksite=ws,
            manual_client_ip=manual_client_ip if is_manual_scan else None,
            manual_device_key=manual_device_key if is_manual_scan else None,
        )
        PresenceEvent.objects.create(
            user_fk=user, kind=PresenceEvent.Kind.ENTER,
            timestamp=when, worksite=ws
        )
        _publish("enter", user, when, {"worksite": ws})

        print(f"[NFC] ENTER nou pentru {user.UserName} pe {today} la site='{ws}'")

        return JsonResponse({
            "ok": True,
            "state": "ENTER",
            "user": {"id": user.UserId, "name": user.UserName},
            "session": {
                "work_date": str(sess.work_date),
                "in_time": localtime(sess.in_time).isoformat(),
                "worksite": ws,
                **_session_gps_payload(sess),
            },
            "received": data,
            "ts_used": "client" if client_when else "server"
        })



@csrf_exempt
def app_version(request):
    """GET /api/app/version/ - public version gate used by the Android app."""
    if request.method != "GET":
        return JsonResponse({"error": "Only GET allowed"}, status=405)

    return JsonResponse({
        "minimum_version_code": int(os.environ.get("DMX_ANDROID_MIN_VERSION_CODE", "1")),
        "latest_version_code": int(os.environ.get("DMX_ANDROID_LATEST_VERSION_CODE", "1")),
        "latest_version_name": os.environ.get("DMX_ANDROID_LATEST_VERSION_NAME", "1.0"),
        "update_url": os.environ.get(
            "DMX_ANDROID_UPDATE_URL",
            "https://play.google.com/store/apps/details?id=ro.dmxconstruction.dmxclock",
        ),
        "message": os.environ.get(
            "DMX_ANDROID_UPDATE_MESSAGE",
            "Exista o versiune noua. Actualizeaza aplicatia ca sa poti continua pontajul.",
        ),
    })


@csrf_exempt
def pontaj_login(request):
    """
    POST /api/pontaj/login/
    Body Android:
    {
      "pin": "1234",
      "device_key": "android-device-id",
      "uid": "MANUAL"
    }

    Valideaza PIN-ul si returneaza userul fara sa creeze eveniment de pontaj.
    """
    if request.method != "POST":
        return JsonResponse({"error": "Only POST allowed"}, status=405)

    try:
        data = json.loads(request.body or "{}")
    except Exception:
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    uid = str(data.get("uid") or "MANUAL").upper().strip()
    pin = str(data.get("pin") or data.get("content") or "").strip()
    device_key = data.get("device_key")

    blocked, retry_after = _pin_is_blocked(request, device_key=device_key, uid=uid)
    if blocked:
        return JsonResponse({
            "ok": False,
            "error": "Prea multe incercari gresite. Incearca din nou mai tarziu.",
            "error_code": "PIN_TEMPORARILY_BLOCKED",
            "retry_after_seconds": retry_after,
        }, status=429)

    if not pin:
        _log_pin_attempt(request, success=False, reason="missing_pin", device_key=device_key, uid=uid)
        return JsonResponse({
            "ok": False,
            "error": "PIN-ul este obligatoriu.",
            "error_code": "PIN_REQUIRED",
        }, status=400)

    user = _find_user_by_pin(pin)
    if not user:
        _log_pin_attempt(request, success=False, reason="invalid_pin", device_key=device_key, uid=uid)
        return JsonResponse({
            "ok": False,
            "error": "Nu exista niciun angajat pentru PIN-ul introdus.",
            "error_code": "INVALID_PIN",
        }, status=404)

    _log_pin_attempt(request, success=True, reason="login_ok", device_key=device_key, uid=uid)
    return JsonResponse({
        "ok": True,
        "mode": "worker",
        "user": {
            "id": user.UserId,
            "name": user.UserName,
            "serie": user.UserSerie,
            "company": user.Company,
        },
    })


@csrf_exempt
def pontaj_clock(request):
    """
    POST /api/pontaj/clock/
    Body Android/web:
    {
      "pin": "1234",
      "device_key": "android-device-id",
      "gps": {"lat": 45.1, "lng": 25.1, "accuracy": 20, "captured_at": "..."},
      "worksite": "Tractorului Bloc B2",
      "mode": "manual" | "driver"
    }
    """
    if request.method != "POST":
        return JsonResponse({"error": "Only POST allowed"}, status=405)

    try:
        data = json.loads(request.body or "{}")
    except Exception:
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    payload = {
        "uid": str(data.get("uid") or "MANUAL"),
        "tag_type": "manual",
        "content": str(data.get("pin") or data.get("content") or "").strip(),
        "timestamp": data.get("timestamp") or timezone.now().isoformat(),
        "device_key": data.get("device_key"),
        "gps": data.get("gps"),
        "worksite": data.get("worksite") or data.get("site") or data.get("santier"),
        "mode": data.get("mode") or data.get("attendance_mode") or "manual",
    }
    request._body = json.dumps(payload).encode("utf-8")
    return nfc_scan(request)



#rezumat pe ziua de azi
@csrf_exempt
def attendance_today(request):
    if request.method != "GET":
        return JsonResponse({"error": "Only GET allowed"}, status=405)

    today = localdate()
    qs = (AttendanceSession.objects
          .filter(work_date=today)
          .values("user_fk__UserId", "user_fk__UserName")
          .annotate(first_in=Min("in_time"),
                    last_out=Max("out_time"),
                    total_seconds=Sum("duration_seconds"))
          .order_by("user_fk__UserName"))

    rows = []
    for row in qs:
        total = int(row.get("total_seconds") or 0)
        rows.append({
            "UserId": row["user_fk__UserId"],
            "UserName": row["user_fk__UserName"],
            "first_in": localtime(row["first_in"]).strftime("%H:%M:%S") if row["first_in"] else None,
            "last_out": localtime(row["last_out"]).strftime("%H:%M:%S") if row["last_out"] else None,
            "total_hms": _fmt_hms(total),
        })
    return JsonResponse({"date": str(today), "rows": rows})




#doua helpere transforma o durata in secunde in text, si un string yyy mm dd in date
def _fmt_hms(seconds: int) -> str:
    seconds = max(0, int(seconds or 0))
    h = seconds // 3600
    m = (seconds % 3600) // 60
    s = seconds % 60
    return f"{h:02d}:{m:02d}:{s:02d}"

def _parse_iso_date(s: str) -> date:
    try:
        return date.fromisoformat(s)
    except Exception:
        return localdate()


# --- ATTENDANCE DAY: pontajul pe o zi anume ---
@csrf_exempt
def attendance_day(request):
    """GET /api/pontaj/day/?date=YYYY-MM-DD"""
    if request.method != "GET":
        return JsonResponse({"error": "Only GET allowed"}, status=405)

    day = _parse_iso_date(request.GET.get("date") or str(localdate()))

    # NEW: map pentru concedii pe zi
    leaves_map = {x.user_fk_id: x for x in LeaveDay.objects.filter(work_date=day)}

    qs = (AttendanceSession.objects
          .filter(work_date=day)
          .select_related("user_fk")
          .order_by("user_fk__UserName", "in_time"))

    rows_by_user = {}
    for s in qs:
        missing_exit = _is_missing_exit_session(s)
        u = s.user_fk
        key = u.UserId
        if key not in rows_by_user:
            rows_by_user[key] = {
                "UserId": u.UserId,
                "UserName": u.UserName,
                "sessions": [],
                "first_in": None,
                "last_out": None,
                "total_seconds": 0,
                "status": "OUT",
                "day_worksite": None,
            }

        in_local  = localtime(s.in_time) if s.in_time else None
        out_local = localtime(s.out_time) if (s.out_time and not missing_exit) else None
        if s.out_time is None and not missing_exit:
            dur = int((timezone.now() - s.in_time).total_seconds())
            rows_by_user[key]["status"] = "IN"
        else:
            dur = 0 if missing_exit else (s.duration_seconds or int((s.out_time - s.in_time).total_seconds()))
        rows_by_user[key]["total_seconds"] += max(0, dur)

        rows_by_user[key]["sessions"].append({
            "in_time":  in_local.strftime("%H:%M:%S") if in_local else None,
            "out_time": out_local.strftime("%H:%M:%S") if out_local else None,
            "duration_hms": _fmt_hms(dur),
            "open": (s.out_time is None and not missing_exit),
            "missing_exit": missing_exit,
            "session_id": s.id,
            "worksite": s.worksite,
            **_session_gps_payload(s),
        })

        # ultima locație nenulă devine site-ul zilei
        if s.worksite:
            rows_by_user[key]["day_worksite"] = s.worksite

        if in_local:
            if (rows_by_user[key]["first_in"] is None or
                in_local < datetime.fromisoformat(rows_by_user[key]["first_in"])):
                rows_by_user[key]["first_in"] = in_local.isoformat()
        if out_local:
            if (rows_by_user[key]["last_out"] is None or
                out_local > datetime.fromisoformat(rows_by_user[key]["last_out"])):
                rows_by_user[key]["last_out"] = out_local.isoformat()

    rows = []
    for _, row in rows_by_user.items():
        ld = leaves_map.get(row["UserId"])
        rows.append({
            "UserId": row["UserId"],
            "UserName": row["UserName"],
            "first_in": row["first_in"],
            "last_out": row["last_out"],
            "total_hms": _fmt_hms(row["total_seconds"]),
            "status": row["status"],
            "sessions": row["sessions"],
            "day_worksite": row["day_worksite"],
            "leave": (
                {"reason": ld.reason, "hours": str(ld.hours), "multiplier": str(ld.multiplier)}
                if ld else None
            ),
        })

    rows.sort(key=lambda r: r["UserName"].lower())
    return JsonResponse({"date": str(day), "rows": rows}, safe=False)



# --- ATTENDANCE PRESENT: cine e acum în lucru ---
@csrf_exempt
def attendance_present(request):
    if request.method != "GET":
        return JsonResponse({"error": "Only GET allowed"}, status=405)

    open_qs = (AttendanceSession.objects
               .filter(out_time__isnull=True)
               .exclude(source__contains=MISSING_EXIT_SOURCE_TAG)
               .select_related("user_fk")
               .order_by("in_time"))

    now = timezone.now()
    data = []
    for s in open_qs:
        u = s.user_fk
        in_local = localtime(s.in_time)
        dur = int((now - s.in_time).total_seconds())
        data.append({
            "UserId": u.UserId,
            "UserName": u.UserName,
            "work_date": str(s.work_date),
            "in_time": in_local.strftime("%H:%M:%S"),
            "duration_hms": _fmt_hms(dur),
            "session_id": s.id,
            "worksite": s.worksite,  # NEW
            **_session_gps_payload(s),
        })

    return JsonResponse({"now": localtime(now).isoformat(), "present": data})





# --- ATTENDANCE RANGE (user_id): adaugă day_worksite în fiecare zi ---
@csrf_exempt
def attendance_range(request):
    """GET /api/pontaj/range/?start=YYYY-MM-DD&end=YYYY-MM-DD[&user_id=ID]"""
    if request.method != "GET":
        return JsonResponse({"error": "Only GET allowed"}, status=405)

    start = _parse_iso_date(request.GET.get("start") or str(localdate()))
    end   = _parse_iso_date(request.GET.get("end")   or str(localdate()))
    if end < start:
        start, end = end, start

    user_id = request.GET.get("user_id")

    if user_id:
        try:
            user = Users.objects.get(UserId=user_id)
        except Users.DoesNotExist:
            return JsonResponse({"start": str(start), "end": str(end), "users": []})

        qs = (AttendanceSession.objects
              .filter(user_fk=user, work_date__range=(start, end))
              .order_by("in_time"))

        from collections import OrderedDict
        days = OrderedDict()
        cur = start
        while cur <= end:
            days[cur] = {
                "date": str(cur),
                "first_in": None,
                "last_out": None,
                "total_seconds": 0,
                "entries": 0,
                "exits": 0,
                "sessions": [],
                "day_worksite": None,
            }
            cur += timedelta(days=1)

        now = timezone.now()
        for s in qs:
            bucket = days.get(s.work_date)
            if not bucket:
                continue
            missing_exit = _is_missing_exit_session(s)
            in_local = localtime(s.in_time)
            out_local = localtime(s.out_time) if (s.out_time and not missing_exit) else None
            dur = s.duration_seconds if s.out_time else int((now - s.in_time).total_seconds())
            if missing_exit:
                dur = 0
            dur = max(0, int(dur))

            bucket["sessions"].append({
                "in_time":  in_local.strftime("%H:%M:%S"),
                "out_time": out_local.strftime("%H:%M:%S") if out_local else None,
                "duration_hms": _fmt_hms(dur),
                "open": (s.out_time is None and not missing_exit),
                "missing_exit": missing_exit,
                "session_id": s.id,
                "worksite": s.worksite,
                **_session_gps_payload(s),
            })
            bucket["entries"] += 1
            if s.out_time and not missing_exit:
                bucket["exits"] += 1
            bucket["total_seconds"] += dur

            if s.worksite:
                bucket["day_worksite"] = s.worksite

            if not bucket["first_in"] or in_local.isoformat() < bucket["first_in"]:
                bucket["first_in"] = in_local.isoformat()
            if out_local and (not bucket["last_out"] or out_local.isoformat() > bucket["last_out"]):
                bucket["last_out"] = out_local.isoformat()

        # NEW: map pentru concedii în interval
        leave_by_date = {
            x.work_date: x
            for x in LeaveDay.objects.filter(user_fk=user, work_date__range=(start, end))
        }

        day_list = []
        for d in days.values():
            ld = leave_by_date.get(_date.fromisoformat(d["date"]))
            day_list.append({
                "date": d["date"],
                "first_in": d["first_in"],
                "last_out": d["last_out"],
                "total_hms": _fmt_hms(d["total_seconds"]),
                "entries": d["entries"],
                "exits": d["exits"],
                "sessions": d["sessions"],
                "day_worksite": d["day_worksite"],
                "leave": (
                    {"reason": ld.reason, "hours": str(ld.hours), "multiplier": str(ld.multiplier)}
                    if ld else None
                ),
            })

        return JsonResponse({
            "start": str(start),
            "end": str(end),
            "users": [{
                "UserId": user.UserId,
                "UserName": user.UserName,
                "days": day_list
            }]
        })

    # agregat fără user_id – rămâne ca înainte
    qs = (AttendanceSession.objects
          .filter(work_date__range=(start, end))
          .values("user_fk__UserId", "user_fk__UserName", "work_date")
          .annotate(
              first_in=Min("in_time"),
              last_out=Max("out_time"),
              total_seconds=Sum("duration_seconds"),
          )
          .order_by("user_fk__UserName", "work_date"))

    users = {}
    for r in qs:
        uid = r["user_fk__UserId"]; uname = r["user_fk__UserName"]
        if uid not in users:
            users[uid] = {"UserId": uid, "UserName": uname, "days": []}
        users[uid]["days"].append({
            "date": str(r["work_date"]),
            "first_in": localtime(r["first_in"]).strftime("%H:%M:%S") if r["first_in"] else None,
            "last_out": localtime(r["last_out"]).strftime("%H:%M:%S") if r["last_out"] else None,
            "total_hms": _fmt_hms(r["total_seconds"]),
        })

    result = list(sorted(users.values(), key=lambda x: x["UserName"].lower()))
    return JsonResponse({"start": str(start), "end": str(end), "users": result})


@csrf_exempt
def attendance_worksite_report(request):
    """GET /api/pontaj/reports/worksites/?start=YYYY-MM-DD&end=YYYY-MM-DD"""
    if request.method != "GET":
        return JsonResponse({"error": "Only GET allowed"}, status=405)

    start = _parse_iso_date(request.GET.get("start") or str(localdate()))
    end = _parse_iso_date(request.GET.get("end") or str(localdate()))
    company = (request.GET.get("company") or "").strip()
    if end < start:
        start, end = end, start

    now = timezone.now()
    qs = (
        AttendanceSession.objects
        .filter(work_date__range=(start, end))
        .select_related("user_fk")
        .order_by("worksite", "user_fk__UserName", "in_time")
    )
    if company:
        qs = qs.filter(user_fk__Company__iexact=company)

    worksites = {}
    total_users = set()
    total_days = set()
    total_person_days = set()
    total_seconds = 0
    total_sessions = 0

    for session in qs:
        missing_exit = _is_missing_exit_session(session)
        user = session.user_fk
        worksite_key = (session.worksite or "").strip()
        worksite_label = worksite_key or "Fără șantier asignat"

        duration = session.duration_seconds
        if session.out_time is None:
            duration = int((now - session.in_time).total_seconds())
        if missing_exit:
            duration = 0
        duration = max(0, int(duration or 0))

        bucket = worksites.setdefault(
            worksite_key,
            {
                "worksite": worksite_label,
                "raw_worksite": worksite_key or None,
                "total_seconds": 0,
                "total_sessions": 0,
                "days": set(),
                "person_days": set(),
                "people": {},
                "last_activity": None,
            },
        )

        bucket["total_seconds"] += duration
        bucket["total_sessions"] += 1
        bucket["days"].add(session.work_date.isoformat())
        bucket["person_days"].add((user.UserId, session.work_date.isoformat()))

        person = bucket["people"].setdefault(
            user.UserId,
            {
                "UserId": user.UserId,
                "UserName": user.UserName,
                "Company": user.Company,
                "total_seconds": 0,
                "sessions_count": 0,
                "days": set(),
            },
        )
        person["total_seconds"] += duration
        person["sessions_count"] += 1
        person["days"].add(session.work_date.isoformat())

        activity_dt = session.out_time or session.in_time
        if activity_dt and (bucket["last_activity"] is None or activity_dt > bucket["last_activity"]):
            bucket["last_activity"] = activity_dt

        total_users.add(user.UserId)
        total_days.add(session.work_date.isoformat())
        total_person_days.add((user.UserId, session.work_date.isoformat()))
        total_seconds += duration
        total_sessions += 1

    rows = []
    for item in worksites.values():
        people = []
        for person in item["people"].values():
            people.append({
                "UserId": person["UserId"],
                "UserName": person["UserName"],
                "Company": person["Company"],
                "total_seconds": person["total_seconds"],
                "total_hms": _fmt_hms(person["total_seconds"]),
                "sessions_count": person["sessions_count"],
                "days_count": len(person["days"]),
            })

        people.sort(key=lambda p: (-p["total_seconds"], p["UserName"].lower()))

        rows.append({
            "worksite": item["worksite"],
            "raw_worksite": item["raw_worksite"],
            "people_count": len(item["people"]),
            "active_days_count": len(item["days"]),
            "person_days_count": len(item["person_days"]),
            "total_sessions": item["total_sessions"],
            "total_seconds": item["total_seconds"],
            "total_hms": _fmt_hms(item["total_seconds"]),
            "last_activity": localtime(item["last_activity"]).isoformat() if item["last_activity"] else None,
            "people": people,
        })

    rows.sort(key=lambda r: (-r["people_count"], -r["total_seconds"], r["worksite"].lower()))

    return JsonResponse({
        "start": str(start),
        "end": str(end),
        "company": company or None,
        "summary": {
            "worksites_count": len(rows),
            "people_count": len(total_users),
            "active_days_count": len(total_days),
            "person_days_count": len(total_person_days),
            "total_sessions": total_sessions,
            "total_seconds": total_seconds,
            "total_hms": _fmt_hms(total_seconds),
        },
        "rows": rows,
    })



from django.shortcuts import render

def _monitor_initial_events(limit: int = 24):
    today = localdate()
    events = []

    sessions = (
        AttendanceSession.objects
        .filter(work_date=today)
        .select_related("user_fk")
        .order_by("in_time", "id")
    )

    for session in sessions:
        in_time = localtime(session.in_time) if session.in_time else None
        out_time = localtime(session.out_time) if session.out_time else None
        worksite = session.worksite or ""

        if in_time:
            events.append({
                "kind": "enter",
                "user_name": session.user_fk.UserName,
                "at": in_time.strftime("%H:%M:%S"),
                "worksite": worksite,
                "_sort_key": session.in_time,
            })

        if out_time:
            events.append({
                "kind": "exit",
                "user_name": session.user_fk.UserName,
                "at": out_time.strftime("%H:%M:%S"),
                "worksite": worksite,
                "_sort_key": session.out_time,
            })

    events.sort(key=lambda item: item["_sort_key"], reverse=True)
    trimmed = events[:limit]
    for item in trimmed:
        item.pop("_sort_key", None)
    return trimmed


# --- PAGINA MONITOR PONTAJ ---
def monitor_pontaj_page(request):
    # template-ul tău e ToolApp/templates/ToolApp/monitor_pontaj.html
    return render(request, "ToolApp/monitor_pontaj.html", {
        "initial_events": _monitor_initial_events(),
    })


# --- SSE broker minimal pentru monitor pontaj ---(Server-Sent Events), folosit ca să trimiți în timp real către browser ce se întâmplă la pontaj (enter/exit/auto_close etc.).
class SseBroker:
    def __init__(self):
        self._lock = Lock()
        self._subs = set()               # Set[Queue[str]]
        self._recent = deque(maxlen=200) # (id, raw_msg)
        self._next_id = 1

    def publish(self, event_type: str, payload: dict):
        ev_id = self._next_id
        self._next_id += 1
        body = json.dumps(payload, ensure_ascii=False)
        raw = f"id: {ev_id}\nevent: {event_type}\ndata: {body}\n\n"
        self._recent.append((ev_id, raw))
        with self._lock:
            for q in list(self._subs):
                try:
                    q.put_nowait(raw)
                except Exception:
                    pass

    def subscribe(self, last_id: Optional[int] = None):
        q: Queue[str] = Queue()
        with self._lock:
            self._subs.add(q)

        def stream():
            try:
                # replay după Last-Event-ID
                if last_id is not None:
                    for ev_id, raw in list(self._recent):
                        if ev_id > last_id:
                            yield raw
                # buclă principală, cu ping keep-alive
                from queue import Empty as QEmpty
                while True:
                    try:
                        raw = q.get(timeout=15)
                        yield raw
                    except QEmpty:
                        yield "event: ping\ndata: {}\n\n"
            finally:
                with self._lock:
                    self._subs.discard(q)

        return stream()

sse = SseBroker()

# “anunță live” monitorul de pontaj ce s-a întâmplat (enter/exit/etc.)
def _publish(kind: str, user, when=None, extra: Optional[dict] = None):
    sse.publish(kind, {
        "user_id": user.UserId,
        "user_name": user.UserName,
        "at": localtime(when).strftime("%H:%M:%S") if when else "",
        **(extra or {})
    })

#endpoint-ul SSE la care se conectează pagina de monitor (browserul)
@csrf_exempt
def pontaj_stream(request):
    """SSE: /api/pontaj/stream/"""
    last_id = None
    lei = request.META.get("HTTP_LAST_EVENT_ID")
    if lei:
        try:
            last_id = int(lei)
        except Exception:
            last_id = None

    resp = StreamingHttpResponse(sse.subscribe(last_id), content_type="text/event-stream")
    resp["Cache-Control"] = "no-cache"
    resp["X-Accel-Buffering"] = "no"
    return resp





def _parse_client_ts(ts_str: str) -> Optional[datetime]:
    """
    Acceptă ISO 8601 cu 'Z' sau cu offset (+02:00). Dacă e naiv, îl tratăm ca UTC.
    Returnează tz-aware în timezone curent.
    """
    if not ts_str:
        return None
    try:
        ts_norm = ts_str.strip().replace('Z', '+00:00')
        dt = datetime.fromisoformat(ts_norm)
        if timezone.is_naive(dt):
            dt = timezone.make_aware(dt, timezone.utc)
        return dt.astimezone(timezone.get_current_timezone())
    except Exception:
        return None


def close_open_sessions_for_day_at_1730(target_day):
    """
    Închide toate sesiunile cu out_time NULL pentru 'target_day' la 17:30 (ora locală).
    Durata e plafonată și nu devine negativă.
    Creează și PresenceEvent(EXIT). Trimite SSE (dacă există _publish).
    Returnează: numărul de sesiuni închise.
    """
    from ToolApp.models import AttendanceSession, PresenceEvent  # import local ca să evit cicluri
    _, close_at = schedule_bounds_for_day(target_day)
    now = timezone.now()
    # nu setăm niciodată o oră în viitor
    if close_at > now:
        close_at = now

    count = 0
    with transaction.atomic():
        qs = (AttendanceSession.objects
              .select_for_update()
              .filter(work_date=target_day, out_time__isnull=True)
              .exclude(source__contains=MISSING_EXIT_SOURCE_TAG)
              .order_by('in_time'))
        for s in qs:
            # out = max(in_time, 17:30) ca să evităm durată negativă
            out_dt = max(s.in_time, close_at)
            # plafon de siguranță
            max_close = s.in_time + timedelta(hours=PONTAJ_MAX_SHIFT_HOURS)
            if out_dt > max_close:
                out_dt = max_close

            s.out_time = out_dt
            s.duration_seconds = max(0, int((s.out_time - s.in_time).total_seconds()))
            s.source = (s.source or '') + AUTO_CLOSE_SOURCE_TAG
            s.save()

            # PresenceEvent EXIT
            PresenceEvent.objects.create(
                user_fk=s.user_fk,
                kind=PresenceEvent.Kind.EXIT,
                timestamp=out_dt,
                worksite=s.worksite
            )

            # SSE (best effort)
            try:
                _publish("auto_1730", s.user_fk, out_dt, {
                    "work_date": str(s.work_date),
                    "duration_hms": _fmt_hms(s.duration_seconds),
                    "worksite": s.worksite,
                })
            except Exception:
                pass

            count += 1
    return count


@csrf_exempt
def attendance_force_close_1730(request):
    """
    GET /api/pontaj/force_close_1730/?date=YYYY-MM-DD
    Dacă lipsește data -> folosește azi (localdate()).
    """
    if request.method != "GET":
        return JsonResponse({"error": "Only GET allowed"}, status=405)

    d_str = request.GET.get("date")
    if d_str:
        try:
            from datetime import date as _date
            target_day = _date.fromisoformat(d_str)
        except Exception:
            return JsonResponse({"error": "date invalid (YYYY-MM-DD)"}, status=400)
    else:
        target_day = localdate()

    closed = close_open_sessions_for_day_at_1730(target_day)
    return JsonResponse({"ok": True, "date": str(target_day), "closed_sessions": closed})


#un concediu pentru un angajat într-o zi.
@csrf_exempt
def leave_upsert(request):
    """
    POST /api/leave/upsert/
    Body:
    {
      "user_id" | "user_pin" | "user_serie",
      "date": "YYYY-MM-DD",
      "reason": "CO"|"CM"|"ALT",
      "hours": 8,
      "hourly_rate": 25.0,   # opțional; default = Users.hourly_rate
      "multiplier": 1.0,     # ex. 0.75 pt. CM
      "note": "text"
    }
    """
    if request.method != "POST":
        return JsonResponse({"error": "Only POST allowed"}, status=405)
    try:
        data = json.loads(request.body or "{}")
    except Exception:
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    # user lookup
    user = None
    if data.get("user_id"):
        user = Users.objects.filter(UserId=data["user_id"]).first()
    elif data.get("user_pin"):
        user = _find_user_by_pin(data["user_pin"])
    elif data.get("user_serie"):
        user = Users.objects.filter(UserSerie=str(data["user_serie"])).first()
    if not user:
        return JsonResponse({"error": "user not found"}, status=404)

    # date
    try:
        d = _date.fromisoformat(str(data.get("date")))
    except Exception:
        return JsonResponse({"error": "Bad 'date' (YYYY-MM-DD)"}, status=400)

    reason = (data.get("reason") or "CO").upper()
    if reason not in ("CO", "CM", "ALT"):
        return JsonResponse({"error": "reason must be CO|CM|ALT"}, status=400)

    def _snap_leave(x):
        if not x:
            return None
        return {
            "id": x.id,
            "date": str(x.work_date),
            "reason": x.reason,
            "hours": str(x.hours),
            "multiplier": str(x.multiplier),
            "hourly_rate_snapshot": str(x.hourly_rate_snapshot),
            "pay_amount": str(x.pay_amount),
            "note": x.note or "",
        }

    existing = LeaveDay.objects.filter(user_fk=user, work_date=d).first()
    before = _snap_leave(existing)

    from decimal import Decimal, ROUND_HALF_UP
    hours = Decimal(str(data.get("hours") or "8.00"))
    multiplier = Decimal(str(data.get("multiplier") or "1.00"))
    rate = (
        Decimal(str(data.get("hourly_rate")))
        if data.get("hourly_rate") not in (None, "")
        else (user.hourly_rate or Decimal("0.00"))
    )

    pay = (rate * hours * multiplier).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    obj, _ = LeaveDay.objects.get_or_create(
        user_fk=user,
        work_date=d,
        defaults={
            "reason": reason,
            "hours": hours,
            "multiplier": multiplier,
            "hourly_rate_snapshot": rate,
            "pay_amount": pay,
            "note": data.get("note") or "",
        },
    )
    obj.reason = reason
    obj.hours = hours
    obj.multiplier = multiplier
    obj.hourly_rate_snapshot = rate
    obj.pay_amount = pay
    if data.get("note") is not None:
        obj.note = data.get("note") or ""
    obj.save()

    snap = recompute_daily_pay(user, d)

    after = _snap_leave(obj)

    # AUDIT
    try:
        msg = (f"s-a modificat lipsa (leave) la angajatul {user.UserName} pe data {d}: "
               f"{after['reason']} {after['hours']}h ×{after['multiplier']}")
        _audit_line(request, msg, {"user_id": user.UserId, "user_name": user.UserName, "before": before, "after": after})
    except Exception:
        pass

    return JsonResponse({
        "ok": True,
        "user_id": user.UserId,
        "date": str(d),
        "leave": {
            "reason": obj.reason,
            "hours": str(obj.hours),
            "multiplier": str(obj.multiplier),
            "hourly_rate": str(obj.hourly_rate_snapshot),
            "pay": str(obj.pay_amount),
            "note": obj.note or "",
        },
        "pay_snapshot": {
            "total_seconds": snap.total_seconds,
            "day_pay": str(snap.day_pay),
        },
    })


#Citește concediile unui user
@csrf_exempt
def leave_get(request):
    """
    GET /api/leave/get/?user_id=ID&date=YYYY-MM-DD
    GET /api/leave/get/?user_id=ID&start=YYYY-MM-DD&end=YYYY-MM-DD
    """
    if request.method != "GET":
        return JsonResponse({"error": "Only GET allowed"}, status=405)

    try:
        user = Users.objects.get(UserId=int(request.GET.get("user_id")))
    except Exception:
        return JsonResponse({"error":"user not found"}, status=404)

    if request.GET.get("date"):
        try:
            d = _date.fromisoformat(request.GET.get("date"))
        except Exception:
            return JsonResponse({"error":"Bad 'date'"}, status=400)
        ld = LeaveDay.objects.filter(user_fk=user, work_date=d).first()
        if not ld:
            return JsonResponse({"user_id": user.UserId, "date": str(d), "leave": None, "hourly_rate": str(user.hourly_rate or "0.00")})
        return JsonResponse({
            "user_id": user.UserId,
            "date": str(d),
            "leave": {
                "reason": ld.reason,
                "hours": str(ld.hours),
                "multiplier": str(ld.multiplier),
                "hourly_rate": str(ld.hourly_rate_snapshot),
                "pay": str(ld.pay_amount),
                "note": ld.note or ""
            },
            "hourly_rate": str(user.hourly_rate or "0.00")
        })

    # range
    try:
        start = _date.fromisoformat(request.GET.get("start"))
        end   = _date.fromisoformat(request.GET.get("end"))
        if end < start:
            start, end = end, start
    except Exception:
        return JsonResponse({"error":"Provide date or start&end"}, status=400)

    qs = LeaveDay.objects.filter(user_fk=user, work_date__range=(start, end)).order_by("work_date")
    rows = [{
        "date": str(x.work_date),
        "reason": x.reason,
        "hours": str(x.hours),
        "multiplier": str(x.multiplier),
        "hourly_rate": str(x.hourly_rate_snapshot),
        "pay": str(x.pay_amount),
        "note": x.note or ""
    } for x in qs]
    return JsonResponse({"user_id": user.UserId, "start": str(start), "end": str(end), "rows": rows})

#Șterge concediul (LeaveDay) unui user într-o zi
@csrf_exempt
def leave_delete(request):
    """
    POST /api/leave/delete/
    Body: { "user_id"| "user_pin"|"user_serie", "date":"YYYY-MM-DD" }
    """
    if request.method != "POST":
        return JsonResponse({"error": "Only POST allowed"}, status=405)
    try:
        data = json.loads(request.body or "{}")
    except Exception:
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    user = None
    if data.get("user_id"):
        user = Users.objects.filter(UserId=data["user_id"]).first()
    elif data.get("user_pin"):
        user = _find_user_by_pin(data["user_pin"])
    elif data.get("user_serie"):
        user = Users.objects.filter(UserSerie=str(data["user_serie"])).first()
    if not user:
        return JsonResponse({"error": "user not found"}, status=404)

    try:
        d = _date.fromisoformat(str(data.get("date")))
    except Exception:
        return JsonResponse({"error": "Bad 'date'"}, status=400)

    existing = LeaveDay.objects.filter(user_fk=user, work_date=d).first()
    before = None
    if existing:
        before = {
            "id": existing.id,
            "date": str(existing.work_date),
            "reason": existing.reason,
            "hours": str(existing.hours),
            "multiplier": str(existing.multiplier),
            "pay_amount": str(existing.pay_amount),
            "note": existing.note or "",
        }

    LeaveDay.objects.filter(user_fk=user, work_date=d).delete()
    snap = recompute_daily_pay(user, d)

    # AUDIT
    try:
        msg = f"s-a șters lipsa (leave) la angajatul {user.UserName} pe data {d}"
        _audit_line(request, msg, {"user_id": user.UserId, "user_name": user.UserName, "before": before})
    except Exception:
        pass

    return JsonResponse({"ok": True, "date": str(d), "pay_snapshot": {"total_seconds": snap.total_seconds, "day_pay": str(snap.day_pay)}})




#situația de plată a unui user într-o zi
@csrf_exempt
def pay_day(request):
    """GET /api/pay/day/?user_id=ID&date=YYYY-MM-DD"""
    if request.method != "GET":
        return JsonResponse({"error": "Only GET allowed"}, status=405)
    try:
        user_id = int(request.GET.get("user_id"))
        user = Users.objects.get(UserId=user_id)
    except Exception:
        return JsonResponse({"error": "user not found"}, status=404)

    day_str = request.GET.get("date")
    day = _date.fromisoformat(day_str) if day_str else localdate()

    # Asigură snapshot up-to-date (în implementarea ta, recompute_daily_pay poate include și concediul)
    obj = recompute_daily_pay(user, day)

    # Breakdown
    work_seconds = (AttendanceSession.objects
                    .filter(user_fk=user, work_date=day, out_time__isnull=False)
                    .aggregate(total=Sum('duration_seconds'))['total'] or 0)

    leave_qs = LeaveDay.objects.filter(user_fk=user, work_date=day)
    leave_hours = sum([float(x.hours) for x in leave_qs]) if leave_qs else 0.0
    leave_pay = sum([float(x.pay_amount) for x in leave_qs]) if leave_qs else 0.0
    last_reason = leave_qs.first().reason if leave_qs else None

    return JsonResponse({
        "user_id": user.UserId,
        "date": str(day),
        "hourly_rate": str(obj.hourly_rate_snapshot),
        "total_seconds": work_seconds,           # doar sesiuni reale
        "day_pay": str(obj.day_pay),             # include și concediul dacă recompute_daily_pay îl adună
        "breakdown": {
            "work_hours": round(work_seconds/3600, 2),
            "leave_hours": leave_hours,
            "leave_pay": round(leave_pay, 2),
            "leave_reason": last_reason
        }
    })


# situația de plată a unui user într-o lună
@csrf_exempt
def pay_month(request):
    """GET /api/pay/month/?user_id=ID&month=YYYY-MM"""
    if request.method != "GET":
        return JsonResponse({"error": "Only GET allowed"}, status=405)
    try:
        user_id = int(request.GET.get("user_id"))
        user = Users.objects.get(UserId=user_id)
    except Exception:
        return JsonResponse({"error": "user not found"}, status=404)

    month = (request.GET.get("month") or str(localdate())[:7])  # "YYYY-MM"
    y, m = month.split("-"); y = int(y); m = int(m)

    from calendar import monthrange
    last_day = monthrange(y, m)[1]
    start = _date(y, m, 1); end = _date(y, m, last_day)

    # zile cu sesiuni închise
    days_s = set(AttendanceSession.objects
                 .filter(user_fk=user, work_date__range=(start, end), out_time__isnull=False)
                 .values_list('work_date', flat=True).distinct())

    # zile cu concediu (fără să conteze dacă au sesiuni)
    days_l = set(LeaveDay.objects
                 .filter(user_fk=user, work_date__range=(start, end))
                 .values_list('work_date', flat=True).distinct())

    # unire: recalculăm snapshot-ul pentru toate zilele
    all_days = days_s | days_l
    for d in all_days:
        recompute_daily_pay(user, d)

    total = (DailyPay.objects
             .filter(user_fk=user, work_date__range=(start, end))
             .aggregate(s=Sum('day_pay'))['s'] or Decimal('0.00'))

    return JsonResponse({
        "user_id": user.UserId,
        "month": f"{y:04d}-{m:02d}",
        "month_total": str(total)
    })




#Generează un token
def _make_token():
    return make_admin_token()

#Verifică un token
def _check_token(token: str) -> bool:
    return check_admin_token(token)

#Ia parola “corectă” de login
def _expected_password():
    return (
        os.environ.get("PONTAJ_PASSWORD")
        or os.environ.get("PONTAJ_LOGIN_PASSWORD")
        or getattr(settings, "PONTAJ_PASSWORD", "")
        or getattr(settings, "PONTAJ_LOGIN_PASSWORD", "")
    )

# --- AUTH LOGIN / VERIFY ---
@csrf_exempt
def auth_login(request):
    """POST /api/auth/login/  body: { "password": "..." }"""
    if request.method != "POST":
        return JsonResponse({"error": "Only POST allowed"}, status=405)
    try:
        data = json.loads(request.body or "{}")
    except Exception:
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    got = str(data.get("password") or "")
    exp = _expected_password()
    if not exp:
        return JsonResponse({"error": "Server not configured (PONTAJ_PASSWORD missing)"}, status=500)

    if hmac.compare_digest(got, exp):
        tok = _make_token()
        resp = JsonResponse({"ok": True, "role": "admin", "expires_in": TOKEN_AGE})
        resp.set_cookie(
            "ptj",
            tok,
            max_age=TOKEN_AGE,
            httponly=True,
            samesite="Lax",
            secure=getattr(settings, "SESSION_COOKIE_SECURE", True),
        )
        return resp
    return JsonResponse({"error": "Invalid password"}, status=401)

# --- AUTH VERIFY ---
@csrf_exempt
def auth_verify(request):
    """POST /api/auth/verify/ folosind cookie HttpOnly 'ptj' sau Authorization Bearer."""
    tok = get_token_from_request(request)

    ok = _check_token(tok or "")
    if ok:
        return JsonResponse({"ok": True, "role": "admin"})
    return JsonResponse({"ok": False}, status=401)





GRACE_MINUTES = 10  # leeway ±10 min
# orele de program și închidere automată
def schedule_bounds_for_day(day):
    """
    Returnează (start_dt, end_dt) tz-aware pentru ziua 'day'.
    Luni–Vineri: 07:30–17:30
    Sâmbătă:     07:30–14:00
    Duminică:    fără reguli speciale -> tratăm ca Luni–Vineri (dacă nu dorești, schimbă aici)
    """
    wd = day.weekday()  # 0=Mon ... 6=Sun
    start_h, start_m = 7, 30
    if wd == 5:  # Saturday
        end_h, end_m = 14, 0
    else:        # Mon–Fri (+ Sun dacă vrei altfel modifică aici)
        end_h, end_m = 17, 30

    tz = timezone.get_current_timezone()
    start = timezone.make_aware(datetime(day.year, day.month, day.day, start_h, start_m), tz)
    end   = timezone.make_aware(datetime(day.year, day.month, day.day, end_h, end_m), tz)
    return start, end
# aplică leeway pentru intrare/ieșire
def apply_enter_grace(ts):
    """
    Dacă intrarea e între [start, start+10min], o rotunjim la start (ex: 07:35 -> 07:30).
    Altfel păstrăm ora reală.
    """
    day = localdate(ts)
    start, _ = schedule_bounds_for_day(day)
    if start <= ts <= start + timedelta(minutes=GRACE_MINUTES):
        return start
    return ts
# aplică leeway pentru ieșire
def apply_exit_grace(ts):
    """
    Dacă ieșirea e între [end-10min, end], o rotunjim la end (ex: 17:25 -> 17:30).
    Altfel păstrăm ora reală.
    """
    day = localdate(ts)
    _, end = schedule_bounds_for_day(day)
    if end - timedelta(minutes=GRACE_MINUTES) <= ts <= end:
        return end
    return ts
# începerea automată a zilei
def workday_close_dt(local_day):
    """Închidere automată la ora de final a zilei (folosește programul de mai sus)."""
    _, end = schedule_bounds_for_day(local_day)
    return end


# --- PAGINA MONITOR PONTAJ WHITE ---
def monitor_pontaj_page_white(request):
    # template-ul tău e ToolApp/templates/ToolApp/monitor_pontaj.html
    return render(request, "ToolApp/monitor_pontaj_white.html", {
        "initial_events": _monitor_initial_events(),
    })



#--- EDITARE ZI PONTAJ (ADMIN) ---
def _parse_hm_for_day(day: _date, hm: str):
    hh, mm = map(int, str(hm).strip().split(':')[:2])
    tz = timezone.get_current_timezone()
    return timezone.make_aware(datetime(day.year, day.month, day.day, hh, mm, 0), tz)
# formatează secunde ca HH:MM:SS
def _fmt_hms(sec: int) -> str:
    sec = max(0, int(sec or 0))
    h = sec // 3600; m = (sec % 3600) // 60; s = sec % 60
    return f"{h:02d}:{m:02d}:{s:02d}"

# recalculare DailyPay pentru un user într-o zi
def recompute_daily_pay(user, day):
    from decimal import Decimal, ROUND_HALF_UP
    from ToolApp.models import DailyPay
    agg = (AttendanceSession.objects
           .filter(user_fk=user, work_date=day, out_time__isnull=False)
           .aggregate(total=Sum('duration_seconds')))
    total_seconds = int(agg['total'] or 0)
    rate = user.hourly_rate or Decimal('0.00')
    hours = (Decimal(total_seconds) / Decimal(3600))
    day_pay = (rate * hours).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    obj, created = DailyPay.objects.get_or_create(
        user_fk=user, work_date=day,
        defaults={'total_seconds': total_seconds, 'hourly_rate_snapshot': rate, 'day_pay': day_pay}
    )
    if not created:
        obj.total_seconds = total_seconds
        obj.hourly_rate_snapshot = rate
        obj.day_pay = day_pay
        obj.save()
    return obj

#rescrie manual pontajul unei zi anume pt un user anume
@csrf_exempt
def attendance_edit_day(request):
    """
    POST /api/pontaj/day/edit/
    {
      "user_id": 53 | "user_pin": "2882" | "user_serie":"AA11",
      "date": "YYYY-MM-DD",
      "sessions": [
        {"in": "07:30", "out": "12:00", "worksite":"Bloc A"},
        {"in": "12:30", "out": "17:10", "worksite":"Bloc F"}
      ],
      "replace": true,
      "rewrite_presence": true,
      "apply_grace": false
    }
    """
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])

    try:
        data = json.loads(request.body or "{}")
    except Exception:
        return HttpResponseBadRequest("Invalid JSON")

    # user by id/serie/pin
    user = None
    if data.get("user_id"):
        user = Users.objects.filter(UserId=data["user_id"]).first()
    elif data.get("user_pin"):
        user = _find_user_by_pin(data["user_pin"])
    elif data.get("user_serie"):
        user = Users.objects.filter(UserSerie=str(data["user_serie"])).first()
    if not user:
        return JsonResponse({"error": "user not found"}, status=404)

    # day
    try:
        day = _date.fromisoformat(str(data.get("date")))
    except Exception:
        return HttpResponseBadRequest("Bad 'date' (YYYY-MM-DD)")

    sessions = data.get("sessions") or []
    if not isinstance(sessions, list) or not sessions:
        return HttpResponseBadRequest("Provide non-empty 'sessions' list")

    replace = str(data.get("replace", "1")).lower() in ("1", "true", "yes", "on")
    rewrite_presence = str(data.get("rewrite_presence", "1")).lower() in ("1", "true", "yes", "on")
    apply_grace = str(data.get("apply_grace", "0")).lower() in ("1", "true", "yes", "on")

    # build, validate, order, non-overlap
    new_items = []
    try:
        for s in sessions:
            inn = s.get("in") or s.get("in_time")
            out = s.get("out") or s.get("out_time")
            if not inn or not out:
                return HttpResponseBadRequest("Each session needs both 'in' and 'out'")

            in_dt = _parse_hm_for_day(day, inn)
            out_dt = _parse_hm_for_day(day, out)
            if out_dt <= in_dt:
                return HttpResponseBadRequest("A session has out <= in")

            if apply_grace:
                in_dt = apply_enter_grace(in_dt)
                out_dt = apply_exit_grace(out_dt)

            ws = (s.get("worksite") or "").strip() or None
            new_items.append((in_dt, out_dt, ws))
    except Exception as e:
        return HttpResponseBadRequest(f"Invalid HH:MM in sessions ({e})")

    new_items.sort(key=lambda t: t[0])
    for i in range(1, len(new_items)):
        if new_items[i][0] < new_items[i - 1][1]:
            return HttpResponseBadRequest("Overlapping sessions")

    def _snap_sessions(qs):
        out = []
        for ss in qs:
            out.append({
                "id": ss.id,
                "in": localtime(ss.in_time).strftime("%H:%M") if ss.in_time else None,
                "out": localtime(ss.out_time).strftime("%H:%M") if ss.out_time else None,
                "worksite": ss.worksite,
                "source": ss.source,
            })
        return out

    # BEFORE snapshot
    try:
        before_qs = (AttendanceSession.objects
                     .filter(user_fk=user, work_date=day)
                     .order_by("in_time"))
        before_list = _snap_sessions(before_qs)
    except Exception:
        before_list = []

    # commit
    with transaction.atomic():
        if replace:
            AttendanceSession.objects.filter(user_fk=user, work_date=day).delete()
            if rewrite_presence:
                PresenceEvent.objects.filter(user_fk=user, timestamp__date=day).delete()

        created = []
        for in_dt, out_dt, ws in new_items:
            dur = int((out_dt - in_dt).total_seconds())
            ss = AttendanceSession.objects.create(
                user_fk=user,
                work_date=day,
                in_time=in_dt,
                out_time=out_dt,
                duration_seconds=max(0, dur),
                source="manual_edit",
                worksite=ws,
            )
            created.append(ss)

            if rewrite_presence:
                PresenceEvent.objects.create(
                    user_fk=user, kind=PresenceEvent.Kind.ENTER, timestamp=in_dt, worksite=ws
                )
                PresenceEvent.objects.create(
                    user_fk=user, kind=PresenceEvent.Kind.EXIT, timestamp=out_dt, worksite=ws
                )

        recompute_daily_pay(user, day)

    # AFTER snapshot
    try:
        after_qs = (AttendanceSession.objects
                    .filter(user_fk=user, work_date=day)
                    .order_by("in_time"))
        after_list = _snap_sessions(after_qs)
    except Exception:
        after_list = []

    # AUDIT log (fișier)
    try:
        summary = "; ".join(
            [f"{x.get('in')}-{x.get('out')}" + (f" ({x.get('worksite')})" if x.get("worksite") else "")
             for x in after_list]
        )
        msg = f"s-a modificat la angajatul {user.UserName} pontajul pe data {day}: {summary}"
        _audit_line(request, msg, {
            "user_id": user.UserId,
            "user_name": user.UserName,
            "date": str(day),
            "before": before_list,
            "after": after_list,
        })
    except Exception:
        pass

    first_in = min(c.in_time for c in created)
    last_out = max(c.out_time for c in created)
    total = sum(int(c.duration_seconds or 0) for c in created)

    return JsonResponse({
        "ok": True,
        "user": {"UserId": user.UserId, "UserName": user.UserName},
        "date": str(day),
        "first_in": localtime(first_in).isoformat(),
        "last_out": localtime(last_out).isoformat(),
        "total_hms": _fmt_hms(total),
        "sessions": [{
            "session_id": s.id,
            "in_time": localtime(s.in_time).isoformat(),
            "out_time": localtime(s.out_time).isoformat(),
            "duration_hms": _fmt_hms(s.duration_seconds or 0),
            "worksite": s.worksite
        } for s in created]
    })


# șterge o sesiune de pontaj
@csrf_exempt
def attendance_session_delete(request):
    """
    POST /api/pontaj/session/delete/
    Body: { "session_id": 123, "rewrite_presence": true }
    Șterge o singură sesiune și recalculează plata pe zi.
    """
    if request.method != "POST":
        return JsonResponse({"error": "Only POST allowed"}, status=405)

    try:
        data = json.loads(request.body or "{}")
    except Exception:
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    sid = data.get("session_id")
    if not sid:
        return JsonResponse({"error": "session_id required"}, status=400)

    rewrite_presence = str(data.get("rewrite_presence", "1")).lower() in ("1", "true", "yes", "on")

    try:
        with transaction.atomic():
            s = AttendanceSession.objects.select_for_update().get(id=sid)
            user = s.user_fk
            day = s.work_date
            in_t = s.in_time
            out_t = s.out_time
            ws = s.worksite

            before = {
                "id": s.id,
                "in": localtime(in_t).strftime("%H:%M") if in_t else None,
                "out": localtime(out_t).strftime("%H:%M") if out_t else None,
                "worksite": ws,
                "date": str(day),
            }

            s.delete()

            if rewrite_presence:
                if in_t:
                    PresenceEvent.objects.filter(user_fk=user, timestamp=in_t).delete()
                if out_t:
                    PresenceEvent.objects.filter(user_fk=user, timestamp=out_t).delete()

            recompute_daily_pay(user, day)

        # AUDIT
        try:
            msg = (f"s-a șters sesiunea #{before['id']} la angajatul {user.UserName} "
                   f"pe data {day}: {before['in']}-{before['out']}" + (f" ({ws})" if ws else ""))
            _audit_line(request, msg, {"user_id": user.UserId, "user_name": user.UserName, "before": before})
        except Exception:
            pass

        return JsonResponse({"ok": True, "deleted_session_id": sid, "date": str(day)})
    except AttendanceSession.DoesNotExist:
        return JsonResponse({"error": "session not found"}, status=404)





# șterge toate sesiunile de pontaj pentru o zi a unui user
@csrf_exempt
def attendance_day_delete(request):
    """
    DELETE /api/pontaj/day/delete/?user_id=ID&date=YYYY-MM-DD[&rewrite_presence=1]
    sau
    POST   /api/pontaj/day/delete/  body: { "user_id":ID, "date":"YYYY-MM-DD", "rewrite_presence":true }
    Golește toate sesiunile utilizatorului în ziua respectivă.
    """
    method = request.method
    if method not in ("DELETE", "POST"):
        return JsonResponse({"error": "Only DELETE or POST allowed"}, status=405)

    if method == "DELETE":
        user_id = request.GET.get("user_id")
        d_str = request.GET.get("date")
        rewrite_presence = str(request.GET.get("rewrite_presence", "1")).lower() in ("1", "true", "yes", "on")
    else:
        try:
            body = json.loads(request.body or "{}")
        except Exception:
            return JsonResponse({"error": "Invalid JSON"}, status=400)
        user_id = body.get("user_id")
        d_str = body.get("date")
        rewrite_presence = str(body.get("rewrite_presence", "1")).lower() in ("1", "true", "yes", "on")

    try:
        user_id = int(user_id)
        user = Users.objects.get(UserId=user_id)
    except Exception:
        return JsonResponse({"error": "user not found"}, status=404)

    try:
        day = _date.fromisoformat(d_str)
    except Exception:
        return JsonResponse({"error": "Invalid 'date' (YYYY-MM-DD)"}, status=400)

    # BEFORE snapshot
    before_list = []
    try:
        qs0 = (AttendanceSession.objects
               .filter(user_fk=user, work_date=day)
               .order_by("in_time"))
        for s in qs0:
            before_list.append({
                "id": s.id,
                "in": localtime(s.in_time).strftime("%H:%M") if s.in_time else None,
                "out": localtime(s.out_time).strftime("%H:%M") if s.out_time else None,
                "worksite": s.worksite,
            })
    except Exception:
        pass

    with transaction.atomic():
        qs = AttendanceSession.objects.select_for_update().filter(user_fk=user, work_date=day)
        deleted = qs.count()
        qs.delete()

        if rewrite_presence:
            PresenceEvent.objects.filter(user_fk=user, timestamp__date=day).delete()

        try:
            obj = DailyPay.objects.get(user_fk=user, work_date=day)
            obj.delete()
        except DailyPay.DoesNotExist:
            pass

    # AUDIT
    try:
        msg = f"s-a șters pontajul COMPLET la angajatul {user.UserName} pe data {day} (sesiuni={deleted})"
        _audit_line(request, msg, {
            "user_id": user.UserId,
            "user_name": user.UserName,
            "date": str(day),
            "deleted_sessions": deleted,
            "before": before_list
        })
    except Exception:
        pass

    return JsonResponse({"ok": True, "date": str(day), "deleted_sessions": deleted})





#ransformă o valoare de timp într-un datetime tz-aware, acceptând două formate diferite
def _parse_hm_or_iso_for_day(day: _date, value: str):
    """
    Acceptă fie 'HH:MM', fie un ISO datetime complet (ex: '2025-12-02T07:30:00+02:00').
    Dacă e ISO -> folosim direct; dacă nu -> tratăm ca HH:MM în ziua 'day'.
    """
    if not value:
        raise ValueError("Empty time value")

    s = str(value).strip()

    # 1) Încercăm ISO 8601
    try:
        if ("T" in s) or (len(s) > 5 and "-" in s):
            s_norm = s.replace("Z", "+00:00")
            dt = datetime.fromisoformat(s_norm)
            if timezone.is_naive(dt):
                dt = timezone.make_aware(dt, timezone.get_current_timezone())
            return dt
    except Exception:
        pass

    # 2) Dacă nu e ISO, tratăm ca 'HH:MM'
    return _parse_hm_for_day(day, s)

#modifică o singură sesiune de pontaj
@csrf_exempt
def attendance_session_update(request):
    """
    POST /api/pontaj/session/update/
    Body: { "session_id": 123, "in":"07:45", "out":"17:10", "worksite":"Bloc E", "apply_grace": false }
    """
    if request.method != "POST":
        return JsonResponse({"error": "Only POST allowed"}, status=405)

    try:
        data = json.loads(request.body or "{}")
    except Exception:
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    sid = data.get("session_id")
    if not sid:
        return JsonResponse({"error": "session_id required"}, status=400)

    try:
        with transaction.atomic():
            s = AttendanceSession.objects.select_for_update().get(id=sid)
            day = s.work_date

            # Folosim helperul deja definit in fișier: _parse_hm_or_iso_for_day
            in_dt  = _parse_hm_or_iso_for_day(day, data.get("in"))  if data.get("in")  else s.in_time
            out_dt = _parse_hm_or_iso_for_day(day, data.get("out")) if data.get("out") else s.out_time
            if out_dt and out_dt < in_dt:
                return JsonResponse({"error": "out < in"}, status=400)

            if str(data.get("apply_grace","0")).lower() in ("1","true","yes","on"):
                in_dt  = apply_enter_grace(in_dt)
                if out_dt:
                    out_dt = apply_exit_grace(out_dt)

            s.in_time = in_dt
            s.out_time = out_dt
            s.worksite = data.get("worksite") or s.worksite
            s.duration_seconds = max(0, int((out_dt - in_dt).total_seconds())) if out_dt else None
            s.source = (s.source or "") + "|manual_edit"
            s.save()

            if out_dt:
                recompute_daily_pay(s.user_fk, day)

        return JsonResponse({"ok": True, "session_id": sid})
    except AttendanceSession.DoesNotExist:
        return JsonResponse({"error": "session not found"}, status=404)


@csrf_exempt
def attendance_session_delete(request):
    """
    POST /api/pontaj/session/delete/
    Body: { "session_id": 123, "rewrite_presence": true }
    Șterge o singură sesiune și recalculează plata pe zi.
    """
    if request.method != "POST":
        return JsonResponse({"error": "Only POST allowed"}, status=405)

    try:
        data = json.loads(request.body or "{}")
    except Exception:
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    sid = data.get("session_id")
    if not sid:
        return JsonResponse({"error": "session_id required"}, status=400)

    rewrite_presence = str(data.get("rewrite_presence","1")).lower() in ("1","true","yes","on")

    try:
        with transaction.atomic():
            s = AttendanceSession.objects.select_for_update().get(id=sid)
            user = s.user_fk
            day  = s.work_date
            in_t  = s.in_time
            out_t = s.out_time

            s.delete()

            # Opțional: curăț evenimentele ENTER/EXIT cu timestamp fix pe in/out-ul sesiunii
            if rewrite_presence:
                if in_t:
                    PresenceEvent.objects.filter(user_fk=user, timestamp=in_t).delete()
                if out_t:
                    PresenceEvent.objects.filter(user_fk=user, timestamp=out_t).delete()

            # Recalc salariu pe zi (dacă mai există alte sesiuni cu out_time setat)
            recompute_daily_pay(user, day)

        return JsonResponse({"ok": True, "deleted_session_id": sid, "date": str(day)})
    except AttendanceSession.DoesNotExist:
        return JsonResponse({"error": "session not found"}, status=404)


@csrf_exempt
def attendance_day_delete(request):
    """
    DELETE /api/pontaj/day/delete/?user_id=ID&date=YYYY-MM-DD[&rewrite_presence=1]
    sau
    POST   /api/pontaj/day/delete/  body: { "user_id":ID, "date":"YYYY-MM-DD", "rewrite_presence":true }
    Golește toate sesiunile utilizatorului în ziua respectivă.
    """
    method = request.method
    if method not in ("DELETE", "POST"):
        return JsonResponse({"error": "Only DELETE or POST allowed"}, status=405)

    if method == "DELETE":
        user_id = request.GET.get("user_id")
        d_str   = request.GET.get("date")
        rewrite_presence = str(request.GET.get("rewrite_presence","1")).lower() in ("1","true","yes","on")
    else:
        try:
            body = json.loads(request.body or "{}")
        except Exception:
            return JsonResponse({"error": "Invalid JSON"}, status=400)
        user_id = body.get("user_id")
        d_str   = body.get("date")
        rewrite_presence = str(body.get("rewrite_presence","1")).lower() in ("1","true","yes","on")

    # Validare parametri
    try:
        user_id = int(user_id)
        user = Users.objects.get(UserId=user_id)
    except Exception:
        return JsonResponse({"error": "user not found"}, status=404)

    try:
        day = _date.fromisoformat(d_str)
    except Exception:
        return JsonResponse({"error": "Invalid 'date' (YYYY-MM-DD)"}, status=400)

    with transaction.atomic():
        qs = AttendanceSession.objects.select_for_update().filter(user_fk=user, work_date=day)
        deleted = qs.count()
        qs.delete()

        if rewrite_presence:
            PresenceEvent.objects.filter(user_fk=user, timestamp__date=day).delete()

        # Șterg snapshot-ul DailyPay (dacă există). Alternativ ai putea seta 0.
        try:
            obj = DailyPay.objects.get(user_fk=user, work_date=day)
            obj.delete()
        except DailyPay.DoesNotExist:
            pass

    return JsonResponse({"ok": True, "date": str(day), "deleted_sessions": deleted})


#variabile din  din openpyxl.styles pentru formatare Excel
HEADER_FONT = Font(bold=True, size=14)
CENTER = Alignment(horizontal="center", vertical="center")


@csrf_exempt
def generate_excel(request):
    """
    Generează Excel de pontaj pentru o lună, opțional filtrat pe firmă.

    Body (POST JSON) – exemple:
      { "month": "2025-11" }
      { "month": 11, "year": 2025 }
      { "month": "Noiembrie", "year": 2025 }
      { "month": 11, "year": 2025, "company": "VB-ROM" }
      { "month": 11, "year": 2025, "companyName": "VB-ROM" }
    """
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request method. Use POST."}, status=400)

    try:
        data = json.loads(request.body or "{}")
    except Exception as e:
        return JsonResponse({"error": f"Invalid JSON: {e}"}, status=400)

    month_val = data.get("month")
    year_val = data.get("year")
    # acceptăm și "company" și "companyName", ca în vechiul proiect
    company_name = (data.get("company") or data.get("companyName") or "").strip()

    if not month_val:
        return JsonResponse({
            "error": "Field 'month' is required. Exemple: '2025-11', 11, 'Noiembrie'"
        }, status=400)

    # --- Parse month/year ---
    month_map = {
        "ianuarie": 1, "februarie": 2, "martie": 3, "aprilie": 4,
        "mai": 5, "iunie": 6, "iulie": 7, "august": 8,
        "septembrie": 9, "octombrie": 10, "noiembrie": 11, "decembrie": 12,
    }
    now_year = datetime.now().year

    month_idx = None
    year = None
    label_month = None   
    #daca luna vine ca string
    if isinstance(month_val, str):
        mv = month_val.strip()
        # daca contine - in string presupunem format "YYYY-MM"
        if "-" in mv:
            parts = mv.split("-")
            if len(parts) != 2:
                return JsonResponse({"error": "month 'YYYY-MM' trebuie să aibă exact 2 părți"}, status=400)
            try:
                year = int(parts[0])
                month_idx = int(parts[1])
                label_month = f"{year:04d}-{month_idx:02d}"
            except Exception:
                return JsonResponse({"error": "Nu pot parsa month='YYYY-MM'"}, status=400)
        #Dacă NU e "YYYY-MM" → încearcă nume de lună sau număr string
        else:
            mv_lower = mv.lower()
            if mv_lower in month_map:
                month_idx = month_map[mv_lower]
                year = int(year_val or now_year)
                label_month = f"{mv.capitalize()} {year}"
            else:
                # poate e "11"
                try:
                    month_idx = int(mv)
                    year = int(year_val or now_year)
                    label_month = f"{year:04d}-{month_idx:02d}"
                except Exception:
                    return JsonResponse({"error": "Nu pot parsa câmpul 'month'."}, status=400)
    else:
        # numeric direct: month=11, year=2025
        try:
            month_idx = int(month_val)
            year = int(year_val or now_year)
            label_month = f"{year:04d}-{month_idx:02d}"
        except Exception:
            return JsonResponse({"error": "Nu pot parsa month/year (numeric)."}, status=400)

    if not (1 <= month_idx <= 12):
        return JsonResponse({"error": "month trebuie să fie între 1 și 12"}, status=400)

    from calendar import monthrange
    days_in_month = monthrange(year, month_idx)[1]

    # --- intervalul de dată ---
    start_day = _date(year, month_idx, 1)
    end_day   = _date(year, month_idx, days_in_month)

    # --- toți angajații din DB, apoi filtrăm manual pe firmă ---
    all_users_qs = Users.objects.all().order_by("UserName")
    all_users = list(all_users_qs)

    # filtrare pe firmă (dacă e cazul)
    if company_name:
        filtered_users = []
        for u in all_users:
            comp_val = getattr(u, "Company", None)
            if not comp_val:
                continue
            if str(comp_val).strip().lower() == company_name.lower():
                filtered_users.append(u)
        users_list = filtered_users
    else:
        users_list = all_users

    if not users_list:
        return JsonResponse({
            "error": f"Nu există utilizatori pentru compania '{company_name}'" if company_name
                     else "Nu există utilizatori în baza de date"
        }, status=400)

    # per_user[UserId] = {
    #   "user": obj,
    #   "per_day_seconds": {1: sec, 2: sec, ...},
    #   "per_day_site":    {1: 'Bloc A', 2: 'Bloc B2', ...},
    #   "per_day_leave":   {1: {"code": "CO", "hours": 8}, ...}
    # }

    #creează “tabelul” pe care îl completam ulterior din DB și apoi îl punem în Excel.
    per_user: dict[int, dict] = {}
    for u in users_list:
        per_user[u.UserId] = {
            "user": u,
            "per_day_seconds": {day: 0 for day in range(1, days_in_month + 1)},
            "per_day_site":    {day: None for day in range(1, days_in_month + 1)},
            "per_day_leave":   {day: None for day in range(1, days_in_month + 1)},
        }

    # --- funcție de mapare a șantierului la cod (T-A, T-B2 etc.) ---
    def _site_label(raw_ws):
        """
        Transformă denumirea de șantier în cod scurt pentru Excel.
        Exemple dorite:
          "Tractorului Bloc A"   -> "T BlA"
          "Tractorului Bloc B2"  -> "T BlB2"
          "Bloc A"               -> "T BlA"
          "Depozit Tractorului"  -> "T Depozit"
        """
        if not raw_ws:
            return None

        text = str(raw_ws).strip()
        if not text:
            return None

        # Căutăm cuvântul 'bloc' oriunde în text și luăm primul token după el
        # ex: "Tractorului Bloc B2" -> group(1) = "B2"
        m = re.search(r"\bbloc\b\s*([A-Za-z0-9]+)", text, flags=re.IGNORECASE)
        if m:
            code = m.group(1).upper()
            return f"T Bl{code}"

        # fallback: dacă nu găsim 'bloc', folosim primul cuvânt ca identificator scurt
        first = text.split()[0]
        if not first:
            return None
        return f"T {first}"

    # --- funcție de mapare LeaveDay -> cod (CO, CM, ALT etc.) ---
    def _leave_code_from_lv(lv) -> Optional[str]:
        raw = getattr(lv, "reason", None) or getattr(lv, "leave_type", None)
        if not raw:
            return None
        t = str(raw).strip()
        if not t:
            return None

        up = t.upper()
        if up in ("CO", "CM", "ALT"):
            return up

        low = t.lower()
        if "odihn" in low:      # Concediu odihnă
            return "CO"
        if "medic" in low or "boal" in low:  # Concediu medical
            return "CM"
        return None

    # --- sesiuni din luna respectivă (PENTRU TOȚI), dar ignorăm userii din alte firme ---
    sessions_qs = (
        AttendanceSession.objects
        .filter(work_date__range=(start_day, end_day))
        .select_related("user_fk")
        .order_by("user_fk__UserName", "work_date", "in_time")
    )
    #“umple” tabelul per_user cu ore lucrate și șantierul
    for s in sessions_qs:
        user = s.user_fk
        uid = user.UserId

        # nu adăugăm useri noi aici; dacă nu e în per_user => altă firmă
        if uid not in per_user:
            continue
        #Determină ziua din lună pentru sesiunea respectivă
        day = s.work_date.day
        if not (1 <= day <= days_in_month):
            continue

        #Calculează durata sesiunii (în secunde)
        dur = s.duration_seconds
        if dur is None and s.in_time and s.out_time:
            dur = int((s.out_time - s.in_time).total_seconds())
        dur = int(dur or 0)
        # dună durata sesiunii în totalul pe zi al muncitorului
        per_user[uid]["per_day_seconds"][day] += max(0, dur)

        #Salvează șantierul/locația pentru ziua aia 
        if s.worksite:
            per_user[uid]["per_day_site"][day] = s.worksite

    # --- concedii / absențe din LeaveDay pentru luna respectivă ---
    leaves_qs = (
        LeaveDay.objects
        .filter(work_date__range=(start_day, end_day), user_fk__in=users_list)
        .select_related("user_fk")
        .order_by("user_fk__UserName", "work_date")
    )
    #Iterează prin toate concediile din luna selectată
    for lv in leaves_qs:
        user = lv.user_fk
        uid = user.UserId
        if uid not in per_user:
            continue

        day = lv.work_date.day
        if not (1 <= day <= days_in_month):
            continue

        code = _leave_code_from_lv(lv)
        raw_hours = getattr(lv, "hours", 0)  # Decimal / int / float
        try:
            hours = float(raw_hours or 0)
        except Exception:
            hours = 0.0

        per_user[uid]["per_day_leave"][day] = {
            "code": code,
            "hours": hours,
        }

    # --- Template Excel ---
    template_path = os.path.join(os.path.dirname(__file__), "..", "media", "model_xcel_pontaj.xlsx")
    if not os.path.exists(template_path):
        return JsonResponse({"error": f"Excel template file not found at {template_path}"}, status=500)

    wb = load_workbook(template_path, data_only=False)
    ws = wb.active

    # Header info
    if company_name:
        ws["B6"].value = company_name
    else:
        ws["B6"].value = data.get("title") or "Pontaj angajați – toți muncitorii"
    ws["A5"].value = label_month

    # Header zile (rând 8, de la coloana D)
    start_column = 4  # D
    for day in range(1, days_in_month + 1):
        col_letter = get_column_letter(start_column + day - 1)
        cell = ws[f"{col_letter}8"]
        cell.value = day
        cell.font = HEADER_FONT
        cell.alignment = CENTER

    # Rânduri per angajat
    start_row = 9
    row_increment = 3  # ore, (rând gol), șantiere
    total_ll_hours_column = "AJ"  # coloana pentru total ore libere legale
    total_cm_hours_column = "AR"  # coloana pentru total ore concediu medical
    hourly_salary_column = "AP" # coloana pentru salariu orar
    total_salary_column = "AQ" # coloana pentru salariu total

    # sortăm după nume
    sorted_users = sorted(per_user.values(), key=lambda x: x["user"].UserName.lower())
    #Parcurge fiecare muncitor și extrage “pachetele” lui
    for idx, payload in enumerate(sorted_users):
        user = payload["user"]
        per_day_seconds = payload["per_day_seconds"]
        per_day_site    = payload["per_day_site"]
        per_day_leave   = payload["per_day_leave"]

        row_hours = start_row + (idx * row_increment)
        row_sites = row_hours + 2  # aici merg codurile T-A, T-B2, CO, CM etc.
        total_ll_hours = 0.0
        total_cm_hours = 0.0
        base_salary_total = 0.0
        total_worked_hours = 0.0
        rate = Decimal(str(data.get("hourly_rate"))) if data.get("hourly_rate") not in (None, "") else (user.hourly_rate or Decimal('0.00'))
      
        # numele muncitorului în col B
        ws[f"B{row_hours}"].value = user.UserName
        
        # salariu orar în col AP
        ws[f"{hourly_salary_column}{row_hours}"].value = float(rate or 0.0)
       
        #Parcurge zilele lunii și decide ce pune în fiecare celulă
        for day in range(1, days_in_month + 1):
            day_date = _date(year, month_idx, day)
            col_letter = get_column_letter(start_column + day - 1)
            #Ia concediul din ziua respectivă (dacă există)
            leave_info = per_day_leave.get(day)


            # Total ore CM (concediu medical)
            if leave_info:
                leave_code = leave_info.get("code")
                if leave_code == "CM":
                    leave_hours = leave_info.get("hours") or 0.0
                    total_cm_hours += float(leave_hours)
            ws[f"{total_cm_hours_column}{row_hours}"].value = float(total_cm_hours)

            # Total ore LL (ALT)
            if leave_info:
                leave_code = leave_info.get("code")
                if leave_code == "ALT":
                    leave_hours = leave_info.get("hours") or 0.0
                    total_ll_hours += float(leave_hours)
            ws[f"{total_ll_hours_column}{row_hours}"].value = float(total_ll_hours)

            # Duminica o sărim doar dacă NU există concediu
            if day_date.weekday() == 6 and not leave_info:
                continue

            # 1) Concediu are prioritate față de orele reale
            if leave_info:
                leave_hours = leave_info.get("hours") or 0.0
                leave_code  = leave_info.get("code")

                if leave_hours > 0:
                    ws[f"{col_letter}{row_hours}"].value = float(leave_hours)
                if leave_code:
                    ws[f"{col_letter}{row_sites}"].value = leave_code
                continue

            # 2) Fără concediu, folosim orele din AttendanceSession
            seconds = per_day_seconds.get(day, 0)
            if seconds <= 0:
                continue
            
            hours = round(seconds / 3600.0, 2)  # ore cu 2 zecimale
            ws[f"{col_letter}{row_hours}"].value = float(hours)            

            raw_ws = per_day_site.get(day)
            label_ws = _site_label(raw_ws)
            if label_ws:
                ws[f"{col_letter}{row_sites}"].value = label_ws

    # --- trimitem fișierul la download ---
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    wb.close()
    excel_buffer.seek(0)

    if company_name:
        safe_company = company_name.replace(" ", "_")
        filename = f"pontaj_{safe_company}_{year:04d}-{month_idx:02d}.xlsx"
    else:
        filename = f"pontaj_{year:04d}-{month_idx:02d}.xlsx"

    resp = HttpResponse(
        excel_buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


import os
from threading import Lock
from django.utils import timezone
from django.utils.timezone import localtime

_PONTAJ_AUDIT_LOCK = Lock()
_PONTAJ_AUDIT_PATH = os.path.join(os.path.dirname(__file__), "pontaj_audit.log")

def _client_ip(request):
    xff = request.META.get("HTTP_X_FORWARDED_FOR")
    if xff:
        return xff.split(",")[0].strip()
    return request.META.get("REMOTE_ADDR", "")

def _audit_line(request, msg: str, extra: Optional[dict] = None):
    ts = localtime(timezone.now()).strftime("%Y-%m-%d %H:%M:%S")
    ip = _client_ip(request)
    ua = (request.META.get("HTTP_USER_AGENT") or "")[:120]
    line = f"[{ts}] ip={ip} ua={ua} | {msg}"
    if extra:
        try:
            line += " | " + json.dumps(extra, ensure_ascii=False)
        except Exception:
            pass

    # scriere safe (best-effort)
    try:
        with _PONTAJ_AUDIT_LOCK:
            with open(_PONTAJ_AUDIT_PATH, "a", encoding="utf-8") as f:
                f.write(line + "\n")
    except Exception:
        logger.exception("Cannot write pontaj audit log")
