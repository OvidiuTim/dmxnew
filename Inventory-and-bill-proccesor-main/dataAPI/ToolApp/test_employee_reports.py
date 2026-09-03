import io
import json
from datetime import date
from unittest.mock import patch

from django.test import Client, TestCase
from django.urls import reverse
from openpyxl import load_workbook

from ToolApp.models import (
    AppModuleAccess,
    AppUser,
    EmployeeDocument,
    EmployeeDocumentType,
    EmployeeTeam,
    EmployeeTeamMember,
    LeaveRequest,
    MobileDevice,
    Tools,
    Users,
)
from ToolApp.security import make_admin_token, make_app_user_token


class EmployeeExportTests(TestCase):
    def setUp(self):
        self.admin = Client()
        self.admin.cookies["ptj"] = make_admin_token()
        self.employee = Users.objects.create(
            UserName="Ana Angajat",
            UserSerie="EXP-001",
            UserPin="secret-pin",
            uid="UID-001",
            Company="DMX",
            trade="Inginer",
            phone_number="0700000000",
            email="ana@example.test",
            hire_date=date(2025, 9, 20),
            total_salary_ron="6000.00",
            ticket_benefit_enabled=True,
        )
        self.app_user = AppUser.objects.create(
            employee=self.employee,
            username="ana.angajat",
            pin_hash="secret-password-hash",
        )
        MobileDevice.objects.create(
            employee=self.employee,
            device_key="secret-device-key",
            push_token="secret-push-token",
        )
        self.team = EmployeeTeam.objects.create(name="Echipa Verde", leader=self.employee)
        EmployeeTeamMember.objects.create(team=self.team, employee=self.employee)
        Tools.objects.create(ToolName="Bormașină", ToolSerie="TOOL-1", AssignedTo=self.employee)
        document_type = EmployeeDocumentType.objects.create(
            name="Contract",
            category=EmployeeDocumentType.Category.EMPLOYMENT,
        )
        EmployeeDocument.objects.create(
            employee=self.employee,
            document_type=document_type,
            file="employee_documents/contract.pdf",
            original_file_name="contract.pdf",
            has_expiry=True,
            expiry_date=date(2027, 1, 2),
        )
        LeaveRequest.objects.create(
            employee=self.employee,
            team=self.team,
            leave_type=LeaveRequest.LeaveType.PAID_LEAVE,
            start_date=date(2026, 9, 10),
            end_date=date(2026, 9, 11),
        )
        self.dismissed_employee = Users.objects.create(
            UserName="Zed Demis",
            UserSerie="EXP-002",
            Company="=FORMULA_NEUTRALIZATA",
            employment_status=Users.EmploymentStatus.DISMISSED,
            dismissed_at=date(2026, 8, 31),
            active=False,
        )

    @patch("ToolApp.employee_reports.timezone.localdate", return_value=date(2026, 9, 3))
    def test_export_builds_selected_sheets_and_never_exports_secrets(self, _localdate):
        response = self.admin.post(
            reverse("employee_export"),
            data=json.dumps({"fields": [
                "employee_id", "name", "series", "company", "employment_status", "hire_date",
                "total_salary_ron", "ticket_enabled", "team", "app_username", "app_modules",
                "leave_requests", "tools", "documents", "team_records", "transfer_requests",
                "uid", "UserPin", "pin_hash", "push_token", "device_key",
            ]}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200, response.content)
        self.assertEqual(
            response["Content-Disposition"],
            'attachment; filename="angajati_03-09-2026.xlsx"',
        )
        self.assertEqual(response["Cache-Control"], "no-store")
        workbook = load_workbook(io.BytesIO(response.content))
        self.assertEqual(
            workbook.sheetnames,
            ["Angajați", "Concedii", "Scule", "Documente", "Echipe", "Cereri transfer"],
        )
        main = workbook["Angajați"]
        headers = [cell.value for cell in main[1]]
        self.assertEqual(main.max_row, 3)
        self.assertEqual(headers[:4], ["ID angajat", "Nume", "Serie angajat", "Companie"])
        self.assertEqual(main.cell(2, headers.index("Activ/demis") + 1).value, "Activ")
        self.assertEqual(main.cell(2, headers.index("Eligibil pentru ajutor bilet acasă") + 1).value, "Da")
        self.assertEqual(main.cell(2, headers.index("Data angajării") + 1).number_format, "dd.mm.yyyy")
        dismissed_row = next(
            row for row in range(2, main.max_row + 1)
            if main.cell(row, headers.index("Serie angajat") + 1).value == "EXP-002"
        )
        self.assertEqual(main.cell(dismissed_row, headers.index("Activ/demis") + 1).value, "Demis")
        self.assertEqual(
            main.cell(dismissed_row, headers.index("Companie") + 1).value,
            "'=FORMULA_NEUTRALIZATA",
        )
        exported_text = "\n".join(
            str(cell.value)
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
            if cell.value is not None
        )
        for secret in ("UID-001", "secret-pin", "secret-password-hash", "secret-push-token", "secret-device-key"):
            self.assertNotIn(secret, exported_text)

    def test_non_admin_cannot_export_all_employees(self):
        AppModuleAccess.objects.create(app_user=self.app_user, module_code="attendance")
        client = Client()
        client.cookies["appj"] = make_app_user_token(self.app_user)

        response = client.post(
            reverse("employee_export"),
            data=json.dumps({"fields": ["employee_id", "name"]}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 403)


class IndiaTicketEligibilityReportTests(TestCase):
    start_date = "2026-09-01"
    end_date = "2026-10-31"

    def setUp(self):
        self.admin = Client()
        self.admin.cookies["ptj"] = make_admin_token()
        self.never_used_already = self.employee("Nefolosit Eligibil", "IND-1", date(2024, 1, 1))
        self.never_used_during = self.employee("Nefolosit Devine", "IND-2", date(2025, 9, 20))
        self.used_during = self.employee(
            "Folosit Devine", "IND-3", date(2023, 1, 1), last_trip=date(2025, 10, 15),
        )
        self.used_already = self.employee(
            "Folosit Eligibil", "IND-4", date(2023, 1, 1), last_trip=date(2025, 8, 20),
        )
        self.not_yet = self.employee(
            "Nu Devine", "IND-5", date(2023, 1, 1), last_trip=date(2026, 1, 1),
        )
        self.disabled = self.employee("Beneficiu Oprit", "IND-6", date(2024, 1, 1), enabled=False)
        self.dismissed = self.employee("Demis", "IND-7", date(2024, 1, 1), employment_status="dismissed")
        self.inactive = self.employee("Inactiv", "IND-8", date(2024, 1, 1), active=False)

    def employee(self, name, series, hire_date, *, last_trip=None, enabled=True, active=True, employment_status="active"):
        return Users.objects.create(
            UserName=name,
            UserSerie=series,
            Company="DMX",
            trade="Montator",
            hire_date=hire_date,
            ticket_benefit_enabled=enabled,
            last_home_trip_date=last_trip,
            active=active,
            employment_status=employment_status,
            dismissed_at=date(2026, 8, 1) if employment_status == "dismissed" else None,
        )

    def test_report_uses_backend_eligibility_and_filters_statuses(self):
        response = self.admin.get(reverse("india_ticket_report"), {
            "start_date": self.start_date,
            "end_date": self.end_date,
        })

        self.assertEqual(response.status_code, 200, response.content)
        rows = response.json()["employees"]
        self.assertEqual(
            {row["series"] for row in rows},
            {"IND-1", "IND-2", "IND-3", "IND-4"},
        )
        by_series = {row["series"]: row for row in rows}
        self.assertFalse(by_series["IND-2"]["already_used"])
        self.assertEqual(by_series["IND-2"]["next_eligibility_date"], "2026-09-20")
        self.assertEqual(by_series["IND-2"]["situation"], "Devine eligibil la data de 20.09.2026")
        self.assertTrue(by_series["IND-3"]["already_used"])
        self.assertEqual(by_series["IND-4"]["situation"], "Deja eligibil la începutul perioadei")

    def test_excel_contains_exactly_the_displayed_report_rows(self):
        params = {"start_date": self.start_date, "end_date": self.end_date}
        displayed = self.admin.get(reverse("india_ticket_report"), params).json()["employees"]
        response = self.admin.get(reverse("india_ticket_report_excel"), params)

        self.assertEqual(response.status_code, 200, response.content)
        workbook = load_workbook(io.BytesIO(response.content))
        self.assertEqual(workbook.sheetnames, ["Eligibili bilet India"])
        sheet = workbook.active
        exported_names = [sheet.cell(row, 1).value for row in range(2, sheet.max_row + 1)]
        self.assertEqual(exported_names, [item["name"] for item in displayed])
        self.assertEqual(sheet["E2"].number_format, "dd.mm.yyyy")
        row_by_series = {
            sheet.cell(row, 2).value: row
            for row in range(2, sheet.max_row + 1)
        }
        self.assertEqual(sheet.cell(row_by_series["IND-2"], 7).value, "Nu")
        self.assertEqual(sheet.cell(row_by_series["IND-3"], 7).value, "Da")

    def test_empty_period_returns_no_rows(self):
        response = self.admin.get(reverse("india_ticket_report"), {
            "start_date": "2024-01-01",
            "end_date": "2024-01-31",
        })

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["employees"], [])

    def test_invalid_period_is_rejected(self):
        response = self.admin.get(reverse("india_ticket_report"), {
            "start_date": "2026-10-31",
            "end_date": "2026-09-01",
        })

        self.assertEqual(response.status_code, 400)
