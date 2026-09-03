import io
import json
from collections import OrderedDict
from datetime import date

from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ToolApp.mobile_services import (
    build_leave_summary,
    build_ticket_benefit,
    employee_effective_hire_date,
    seniority_months,
)
from ToolApp.models import Users
from ToolApp.security import request_has_admin


DATE_FORMAT = "dd.mm.yyyy"
HEADER_FILL = PatternFill("solid", fgColor="123B2B")
HEADER_FONT = Font(color="FFFFFF", bold=True)
MONEY_FORMAT = '#,##0.00'


def _yes_no(value):
    return "Da" if bool(value) else "Nu"


def _text(value):
    return "" if value is None else str(value)


def _money(value):
    return None if value is None else float(value)


def _safe_excel_value(value):
    """Keep exported text from being evaluated as a spreadsheet formula."""
    if isinstance(value, str) and value.lstrip().startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def _seniority_label(months):
    months = max(0, int(months or 0))
    years, remaining_months = divmod(months, 12)
    if not years:
        return f"{remaining_months} {'lună' if remaining_months == 1 else 'luni'}"
    if not remaining_months:
        return f"{years} {'an' if years == 1 else 'ani'}"
    return f"{years} {'an' if years == 1 else 'ani'} și {remaining_months} luni"


def _active_membership(employee):
    return next(
        (membership for membership in employee.team_memberships.all() if membership.active and membership.team.active),
        None,
    )


def _team_name(employee):
    membership = _active_membership(employee)
    return membership.team.name if membership else ""


def _team_roles(employee):
    roles = []
    if _active_membership(employee):
        roles.append("Membru")
    if any(team.active for team in employee.led_employee_teams.all()):
        roles.append("Șef de echipă")
    if any(team.active for team in employee.supervised_employee_teams.all()):
        roles.append("Supervisor")
    return ", ".join(roles)


def _organization_value(employee, attribute):
    values = []
    for member in employee.organization_members.all():
        if attribute == "department":
            value = member.department.name
        elif attribute == "role":
            value = member.role
        else:
            value = member.reports_to.name if member.reports_to_id else ""
        if value and value not in values:
            values.append(value)
    return ", ".join(values)


def _app_user(employee):
    return getattr(employee, "app_user", None)


def _salary_profile(employee):
    return getattr(employee, "salary_profile", None)


def _employee_context(employee, as_of):
    effective_hire_date = employee_effective_hire_date(employee, as_of)
    return {
        "effective_hire_date": effective_hire_date,
        "seniority_months": seniority_months(effective_hire_date, as_of),
        "leave": build_leave_summary(employee, as_of),
        "ticket": build_ticket_benefit(employee, as_of),
    }


EXPORT_GROUPS = OrderedDict([
    ("Identificare", OrderedDict([
        ("employee_id", "ID angajat"),
        ("name", "Nume"),
        ("series", "Serie angajat"),
        ("person_type", "Categorie persoană"),
        ("employment_status", "Activ/demis"),
        ("dismissed_at", "Data demiterii"),
    ])),
    ("Contact și angajare", OrderedDict([
        ("company", "Companie"),
        ("trade", "Meserie"),
        ("phone", "Telefon"),
        ("email", "E-mail"),
        ("hire_date", "Data angajării"),
        ("seniority", "Vechime"),
    ])),
    ("Acces în aplicație", OrderedDict([
        ("app_username", "Utilizator aplicație"),
        ("app_account_active", "Cont aplicație activ"),
        ("is_storekeeper", "Magazioner"),
        ("app_modules", "Module permise"),
        ("page_permissions", "Pagini permise"),
    ])),
    ("Salarizare", OrderedDict([
        ("hourly_rate", "Tarif orar (RON)"),
        ("total_salary_ron", "Salariu total (RON)"),
        ("salary_advance_ron", "Avans (RON)"),
        ("salary_remainder_ron", "Rest (RON)"),
        ("meal_vouchers_ron", "Bonuri de masă (RON)"),
        ("net_salary_eur", "Salariu net (EUR)"),
        ("net_salary_ron", "Salariu net profil (RON)"),
        ("food_money_enabled", "Bani de mâncare activați"),
        ("food_money_ron", "Bani de mâncare (RON)"),
    ])),
    ("Concedii", OrderedDict([
        ("leave_accrued", "Zile de concediu acumulate"),
        ("leave_used", "Zile de concediu folosite"),
        ("leave_remaining", "Zile de concediu rămase"),
        ("leave_extra", "Zile suplimentare luate"),
        ("leave_requests", "Cereri de concediu (sheet separat)"),
    ])),
    ("Ajutor bilet acasă", OrderedDict([
        ("ticket_enabled", "Eligibil pentru ajutor bilet acasă"),
        ("ticket_already_used", "A beneficiat deja"),
        ("last_home_trip", "Data ultimei plecări"),
        ("next_ticket_eligibility", "Următoarea dată de eligibilitate"),
        ("ticket_amount_eur", "Sumă maximă ajutor (EUR)"),
    ])),
    ("Cazare", OrderedDict([
        ("accommodation", "Cazare"),
        ("accommodation_address", "Adresă cazare"),
        ("accommodation_room", "Cameră"),
    ])),
    ("Echipament personal", OrderedDict([
        ("equipment_received", "A primit echipament"),
        ("equipment_size", "Mărime echipament"),
    ])),
    ("Scule", OrderedDict([
        ("tools", "Scule atribuite (sheet separat)"),
    ])),
    ("Documente", OrderedDict([
        ("documents", "Documente (sheet separat)"),
    ])),
    ("Echipe", OrderedDict([
        ("team", "Echipă"),
        ("team_role", "Rol în echipă"),
        ("team_records", "Apartenențe și roluri (sheet separat)"),
        ("transfer_requests", "Cereri transfer (sheet separat)"),
    ])),
    ("Organigramă", OrderedDict([
        ("organization_department", "Departament"),
        ("organization_role", "Funcție în organigramă"),
        ("organization_manager", "Superior direct"),
    ])),
])


DEFAULT_EXPORT_FIELDS = [
    "employee_id", "name", "series", "company", "trade", "phone", "email",
    "employment_status", "hire_date", "seniority", "total_salary_ron",
    "leave_remaining", "ticket_enabled", "ticket_already_used", "last_home_trip",
    "next_ticket_eligibility", "accommodation", "team",
]


MAIN_FIELD_VALUES = {
    "employee_id": lambda employee, context: employee.pk,
    "name": lambda employee, context: employee.UserName,
    "series": lambda employee, context: employee.UserSerie,
    "person_type": lambda employee, context: employee.get_person_type_display(),
    "employment_status": lambda employee, context: employee.get_employment_status_display(),
    "dismissed_at": lambda employee, context: employee.dismissed_at,
    "company": lambda employee, context: employee.Company or "",
    "trade": lambda employee, context: employee.trade or "",
    "phone": lambda employee, context: employee.phone_number or "",
    "email": lambda employee, context: employee.email or "",
    "hire_date": lambda employee, context: context["effective_hire_date"],
    "seniority": lambda employee, context: _seniority_label(context["seniority_months"]),
    "app_username": lambda employee, context: _app_user(employee).username if _app_user(employee) else "",
    "app_account_active": lambda employee, context: _yes_no(_app_user(employee).is_active) if _app_user(employee) else "Nu",
    "is_storekeeper": lambda employee, context: _yes_no(_app_user(employee).is_storekeeper) if _app_user(employee) else "Nu",
    "app_modules": lambda employee, context: ", ".join(
        access.get_module_code_display()
        for access in (_app_user(employee).module_accesses.all() if _app_user(employee) else [])
        if access.can_access
    ),
    "page_permissions": lambda employee, context: ", ".join(
        permission.route
        for permission in (_app_user(employee).page_permissions.all() if _app_user(employee) else [])
        if permission.can_access
    ),
    "hourly_rate": lambda employee, context: _money(employee.hourly_rate),
    "total_salary_ron": lambda employee, context: _money(employee.total_salary_ron),
    "salary_advance_ron": lambda employee, context: _money(employee.salary_advance_ron),
    "salary_remainder_ron": lambda employee, context: _money(employee.salary_remainder_ron),
    "meal_vouchers_ron": lambda employee, context: _money(employee.meal_vouchers_ron),
    "net_salary_eur": lambda employee, context: _money(_salary_profile(employee).net_salary_eur) if _salary_profile(employee) else None,
    "net_salary_ron": lambda employee, context: _money(_salary_profile(employee).net_salary_ron) if _salary_profile(employee) else None,
    "food_money_enabled": lambda employee, context: _yes_no(_salary_profile(employee).food_money_enabled) if _salary_profile(employee) else "Nu",
    "food_money_ron": lambda employee, context: _money(_salary_profile(employee).food_money_ron) if _salary_profile(employee) else None,
    "leave_accrued": lambda employee, context: float(context["leave"]["total_accrued_days"]),
    "leave_used": lambda employee, context: context["leave"]["total_used_days"],
    "leave_remaining": lambda employee, context: float(context["leave"]["remaining_days"]),
    "leave_extra": lambda employee, context: float(context["leave"]["extra_days_taken"]),
    "ticket_enabled": lambda employee, context: _yes_no(context["ticket"]["ticket_benefit_enabled"]),
    "ticket_already_used": lambda employee, context: _yes_no(bool(employee.last_home_trip_date)),
    "last_home_trip": lambda employee, context: employee.last_home_trip_date,
    "next_ticket_eligibility": lambda employee, context: date.fromisoformat(context["ticket"]["next_eligibility_date"]) if context["ticket"]["next_eligibility_date"] else None,
    "ticket_amount_eur": lambda employee, context: float(context["ticket"]["ticket_benefit_amount_eur"]),
    "accommodation": lambda employee, context: employee.accommodation.name if employee.accommodation_id else employee.housing_location,
    "accommodation_address": lambda employee, context: employee.accommodation.address if employee.accommodation_id else "",
    "accommodation_room": lambda employee, context: employee.accommodation_room.name if employee.accommodation_room_id else "",
    "equipment_received": lambda employee, context: "" if employee.received_equipment is None else _yes_no(employee.received_equipment),
    "equipment_size": lambda employee, context: employee.equipment_size or "",
    "team": lambda employee, context: _team_name(employee),
    "team_role": lambda employee, context: _team_roles(employee),
    "organization_department": lambda employee, context: _organization_value(employee, "department"),
    "organization_role": lambda employee, context: _organization_value(employee, "role"),
    "organization_manager": lambda employee, context: _organization_value(employee, "manager"),
}


def _all_export_field_labels():
    return {key: label for fields in EXPORT_GROUPS.values() for key, label in fields.items()}


def _employee_queryset():
    return (
        Users.objects.filter(person_type=Users.PersonType.EMPLOYEE)
        .select_related("accommodation", "accommodation_room", "app_user", "salary_profile")
        .prefetch_related(
            "app_user__module_accesses",
            "app_user__page_permissions",
            "assigned_tools",
            "documents__document_type",
            "leave_requests__team",
            "team_memberships__team",
            "led_employee_teams",
            "supervised_employee_teams",
            "temporary_team_requests__requester_team",
            "temporary_team_requests__source_team",
            "portal_transfer_requests__source_team",
            "portal_transfer_requests__destination_team",
            "organization_members__department",
            "organization_members__reports_to",
        )
        .order_by("UserName", "UserId")
    )


def _configure_sheet(sheet, date_columns=(), money_columns=(), decimal_columns=()):
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 32
    for column_index, cells in enumerate(sheet.columns, start=1):
        values = [_text(cell.value) for cell in cells]
        width = min(42, max(12, max((len(value) for value in values), default=0) + 2))
        sheet.column_dimensions[get_column_letter(column_index)].width = width
    for column_index in date_columns:
        for cell in sheet[get_column_letter(column_index)][1:]:
            if cell.value:
                cell.number_format = DATE_FORMAT
    for column_index in money_columns:
        for cell in sheet[get_column_letter(column_index)][1:]:
            if cell.value is not None:
                cell.number_format = MONEY_FORMAT
    for column_index in decimal_columns:
        for cell in sheet[get_column_letter(column_index)][1:]:
            if cell.value is not None:
                cell.number_format = "0.00"


def _add_sheet(workbook, title, headers, rows, *, date_headers=(), money_headers=(), decimal_headers=()):
    sheet = workbook.create_sheet(title)
    sheet.append([_safe_excel_value(value) for value in headers])
    for row in rows:
        sheet.append([_safe_excel_value(value) for value in row])
    _configure_sheet(
        sheet,
        date_columns=[headers.index(header) + 1 for header in date_headers if header in headers],
        money_columns=[headers.index(header) + 1 for header in money_headers if header in headers],
        decimal_columns=[headers.index(header) + 1 for header in decimal_headers if header in headers],
    )
    return sheet


def _identity(employee):
    return [employee.pk, employee.UserName, employee.UserSerie]


def build_employee_export_workbook(selected_fields, as_of=None):
    as_of = as_of or timezone.localdate()
    employees = list(_employee_queryset())
    labels = _all_export_field_labels()
    main_fields = [field for field in selected_fields if field in MAIN_FIELD_VALUES]
    if not main_fields:
        main_fields = ["employee_id", "name", "series"]
    workbook = Workbook()
    workbook.remove(workbook.active)

    main_headers = [labels[field] for field in main_fields]
    main_rows = []
    for employee in employees:
        context = _employee_context(employee, as_of)
        main_rows.append([MAIN_FIELD_VALUES[field](employee, context) for field in main_fields])
    _add_sheet(
        workbook,
        "Angajați",
        main_headers,
        main_rows,
        date_headers={"Data demiterii", "Data angajării", "Data ultimei plecări", "Următoarea dată de eligibilitate"},
        money_headers={label for key, label in labels.items() if key in {
            "hourly_rate", "total_salary_ron", "salary_advance_ron", "salary_remainder_ron",
            "meal_vouchers_ron", "net_salary_eur", "net_salary_ron", "food_money_ron", "ticket_amount_eur",
        }},
        decimal_headers={"Zile de concediu acumulate", "Zile de concediu rămase", "Zile suplimentare luate"},
    )

    if "leave_requests" in selected_fields:
        rows = []
        for employee in employees:
            for item in employee.leave_requests.all():
                rows.append(_identity(employee) + [
                    item.pk, item.team.name if item.team_id else "", item.get_leave_type_display(),
                    item.start_date, item.end_date, item.reason, item.get_status_display(),
                    item.created_at.date() if item.created_at else None,
                    item.reviewed_at.date() if item.reviewed_at else None,
                    item.approved_at.date() if item.approved_at else None,
                ])
        _add_sheet(
            workbook, "Concedii",
            ["ID angajat", "Nume", "Serie angajat", "ID cerere", "Echipă", "Tip concediu", "Data început", "Data sfârșit", "Motiv", "Status", "Data cererii", "Data soluționării", "Data aprobării"],
            rows,
            date_headers={"Data început", "Data sfârșit", "Data cererii", "Data soluționării", "Data aprobării"},
        )

    if "tools" in selected_fields:
        rows = []
        for employee in employees:
            for tool in employee.assigned_tools.all():
                rows.append(_identity(employee) + [
                    tool.ToolId, tool.ToolName, tool.ToolSerie or tool.SerialNumber or "", tool.Category or "",
                    tool.Brand or "", tool.Model or "", tool.Pieces, _yes_no(tool.IsSSM),
                    tool.get_Status_display(), tool.DateOfGiving, tool.ExpiryDate, tool.MainLocation or "",
                    tool.Provider or "", _yes_no(tool.IsReturned), tool.DateReturned,
                    _yes_no(tool.IsLost), tool.DateLost, tool.Detail or "",
                ])
        _add_sheet(
            workbook, "Scule",
            ["ID angajat", "Nume", "Serie angajat", "ID sculă", "Denumire", "Serie sculă", "Categorie", "Marcă", "Model", "Cantitate", "SSM", "Stare", "Data predării", "Data expirării", "Locație", "Furnizor", "Returnată", "Data returnării", "Pierdută", "Data pierderii", "Detalii"],
            rows,
            date_headers={"Data predării", "Data expirării", "Data returnării", "Data pierderii"},
        )

    if "documents" in selected_fields:
        rows = []
        for employee in employees:
            for document in employee.documents.all():
                rows.append(_identity(employee) + [
                    document.pk, document.document_type.get_category_display(), document.document_type.name,
                    document.original_file_name, _yes_no(document.has_expiry), document.expiry_date,
                    document.uploaded_at.date() if document.uploaded_at else None,
                ])
        _add_sheet(
            workbook, "Documente",
            ["ID angajat", "Nume", "Serie angajat", "ID document", "Categorie", "Tip document", "Fișier", "Are expirare", "Data expirării", "Data încărcării"],
            rows,
            date_headers={"Data expirării", "Data încărcării"},
        )

    if "team_records" in selected_fields:
        rows = []
        for employee in employees:
            team_records = {}
            for membership in employee.team_memberships.all():
                team_records.setdefault(membership.team_id, {"team": membership.team, "roles": set(), "active": False})
                team_records[membership.team_id]["roles"].add("Membru")
                team_records[membership.team_id]["active"] = membership.active and membership.team.active
            for team in employee.led_employee_teams.all():
                team_records.setdefault(team.pk, {"team": team, "roles": set(), "active": team.active})
                team_records[team.pk]["roles"].add("Șef de echipă")
            for team in employee.supervised_employee_teams.all():
                team_records.setdefault(team.pk, {"team": team, "roles": set(), "active": team.active})
                team_records[team.pk]["roles"].add("Supervisor")
            for record in team_records.values():
                team = record["team"]
                rows.append(_identity(employee) + [
                    team.pk, team.name, ", ".join(sorted(record["roles"])), _yes_no(record["active"]), team.default_worksite,
                ])
        _add_sheet(
            workbook, "Echipe",
            ["ID angajat", "Nume", "Serie angajat", "ID echipă", "Echipă", "Rol", "Activ", "Șantier implicit"],
            rows,
        )

    if "transfer_requests" in selected_fields:
        rows = []
        for employee in employees:
            for item in employee.temporary_team_requests.all():
                rows.append(_identity(employee) + [
                    "Cerere personal", item.pk, item.get_request_type_display(), item.source_team.name,
                    item.requester_team.name, item.start_date, item.end_date, item.reason,
                    item.get_status_display(), "", "", item.created_at.date() if item.created_at else None,
                    item.resolved_at.date() if item.resolved_at else None,
                ])
            for item in employee.portal_transfer_requests.all():
                rows.append(_identity(employee) + [
                    "Transfer portal", item.pk, "Permanentă", item.source_team.name if item.source_team_id else "",
                    item.destination_team.name, None, None, item.reason, item.get_status_display(),
                    item.get_source_approval_display(), item.get_destination_approval_display(),
                    item.created_at.date() if item.created_at else None,
                    item.completed_at.date() if item.completed_at else None,
                ])
        _add_sheet(
            workbook, "Cereri transfer",
            ["ID angajat", "Nume", "Serie angajat", "Sursă cerere", "ID cerere", "Tip", "Echipă sursă", "Echipă destinație", "Data început", "Data sfârșit", "Motiv", "Status", "Aprobare sursă", "Aprobare destinație", "Data cererii", "Data finalizării"],
            rows,
            date_headers={"Data început", "Data sfârșit", "Data cererii", "Data finalizării"},
        )

    return workbook


def _workbook_response(workbook, filename):
    buffer = io.BytesIO()
    workbook.save(buffer)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    response["X-Content-Type-Options"] = "nosniff"
    response["Cache-Control"] = "no-store"
    return response


def _admin_required(request):
    return getattr(request, "dmx_role", None) == "admin" or request_has_admin(request)


@csrf_exempt
def employee_export(request):
    if request.method != "POST":
        return JsonResponse({"error": "Metodă nepermisă."}, status=405)
    if not _admin_required(request):
        return JsonResponse({"error": "Nu ai dreptul să exporți datele tuturor angajaților."}, status=403)
    try:
        body = json.loads(request.body or "{}")
    except (TypeError, ValueError, UnicodeDecodeError):
        return JsonResponse({"error": "Corpul cererii nu este JSON valid."}, status=400)
    selected_fields = body.get("fields") or []
    if not isinstance(selected_fields, list):
        return JsonResponse({"error": "Selecția câmpurilor nu este validă."}, status=400)
    allowed_fields = set(_all_export_field_labels())
    selected_fields = list(dict.fromkeys(field for field in selected_fields if field in allowed_fields))
    if not selected_fields:
        return JsonResponse({"error": "Selectează cel puțin un câmp pentru export."}, status=400)
    workbook = build_employee_export_workbook(selected_fields)
    today = timezone.localdate()
    return _workbook_response(workbook, f"angajati_{today:%d-%m-%Y}.xlsx")


def india_ticket_report_rows(start_date, end_date):
    employees = Users.objects.filter(
        person_type=Users.PersonType.EMPLOYEE,
        active=True,
        employment_status=Users.EmploymentStatus.ACTIVE,
        ticket_benefit_enabled=True,
    ).order_by("UserName", "UserId")
    rows = []
    for employee in employees:
        ticket = build_ticket_benefit(employee, start_date)
        next_date_raw = ticket.get("next_eligibility_date")
        if not next_date_raw:
            continue
        next_date = date.fromisoformat(next_date_raw)
        if next_date > end_date:
            continue
        effective_hire_date = employee_effective_hire_date(employee, start_date)
        already_eligible = next_date <= start_date
        rows.append({
            "employee_id": employee.pk,
            "name": employee.UserName,
            "series": employee.UserSerie,
            "company": employee.Company or "",
            "trade": employee.trade or "",
            "hire_date": effective_hire_date.isoformat(),
            "seniority_months": seniority_months(effective_hire_date, start_date),
            "seniority": _seniority_label(seniority_months(effective_hire_date, start_date)),
            "already_used": bool(employee.last_home_trip_date),
            "last_home_trip_date": employee.last_home_trip_date.isoformat() if employee.last_home_trip_date else None,
            "next_eligibility_date": next_date.isoformat(),
            "situation": (
                "Deja eligibil la începutul perioadei"
                if already_eligible
                else f"Devine eligibil la data de {next_date:%d.%m.%Y}"
            ),
        })
    return rows


def _parse_report_period(request):
    try:
        start_date = date.fromisoformat(str(request.GET.get("start_date") or ""))
        end_date = date.fromisoformat(str(request.GET.get("end_date") or ""))
    except ValueError:
        return None, None, JsonResponse({"error": "Data de început și data de sfârșit sunt obligatorii."}, status=400)
    if end_date < start_date:
        return None, None, JsonResponse({"error": "Data de sfârșit nu poate fi înaintea datei de început."}, status=400)
    return start_date, end_date, None


def india_ticket_report(request):
    if request.method != "GET":
        return JsonResponse({"error": "Metodă nepermisă."}, status=405)
    if not _admin_required(request):
        return JsonResponse({"error": "Nu ai dreptul să vezi raportul tuturor angajaților."}, status=403)
    start_date, end_date, error = _parse_report_period(request)
    if error:
        return error
    rows = india_ticket_report_rows(start_date, end_date)
    return JsonResponse({
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "employees": rows,
        "count": len(rows),
    })


def india_ticket_report_excel(request):
    if request.method != "GET":
        return JsonResponse({"error": "Metodă nepermisă."}, status=405)
    if not _admin_required(request):
        return JsonResponse({"error": "Nu ai dreptul să exporți raportul tuturor angajaților."}, status=403)
    start_date, end_date, error = _parse_report_period(request)
    if error:
        return error
    report_rows = india_ticket_report_rows(start_date, end_date)
    workbook = Workbook()
    workbook.remove(workbook.active)
    rows = [[
        item["name"], item["series"], item["company"], item["trade"],
        date.fromisoformat(item["hire_date"]), item["seniority"], _yes_no(item["already_used"]),
        date.fromisoformat(item["last_home_trip_date"]) if item["last_home_trip_date"] else None,
        date.fromisoformat(item["next_eligibility_date"]), item["situation"],
    ] for item in report_rows]
    _add_sheet(
        workbook,
        "Eligibili bilet India",
        ["Nume", "Serie angajat", "Companie", "Meserie", "Data angajării", "Vechime", "A beneficiat deja", "Data ultimei plecări", "Următoarea dată de eligibilitate", "Situația în perioada selectată"],
        rows,
        date_headers={"Data angajării", "Data ultimei plecări", "Următoarea dată de eligibilitate"},
    )
    return _workbook_response(
        workbook,
        f"eligibili_bilet_india_{start_date:%d-%m-%Y}_{end_date:%d-%m-%Y}.xlsx",
    )
