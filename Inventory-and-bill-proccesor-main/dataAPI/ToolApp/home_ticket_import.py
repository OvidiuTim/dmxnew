import math
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

from django.db import transaction
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


REQUIRED_COLUMNS = (
    "employee_name",
    "hire_date",
    "ticket_bonus_eligible",
    "company_paid_ticket_before",
    "last_paid_ticket_departure_date",
)
IMPORTED_FIELDS = (
    "hire_date",
    "ticket_benefit_enabled",
    "last_home_trip_date",
)
FIELD_LABELS = {
    "hire_date": "Date de angajare",
    "ticket_benefit_enabled": "Eligibilități",
    "last_home_trip_date": "Istorice de bilet",
}


class HomeTicketImportError(ValueError):
    pass


@dataclass(frozen=True)
class EmployeeIdentifiers:
    cnp: tuple[str, ...] = ()
    passports: tuple[str, ...] = ()


@dataclass(frozen=True)
class ImportIssue:
    sheet: str
    row_number: int
    employee_name: str
    reason: str
    field_name: str = ""
    raw_value: object = None

    @property
    def location(self):
        return f"{self.sheet}!r{self.row_number}"


@dataclass(frozen=True)
class FieldOperation:
    field_name: str
    source_column: str
    raw_value: object
    value: object


@dataclass(frozen=True)
class HomeTicketRow:
    sheet: str
    row_number: int
    raw_name: str
    normalized_name: str
    sorted_name: str
    identifiers: EmployeeIdentifiers
    operations: tuple[FieldOperation, ...]

    @property
    def location(self):
        return f"{self.sheet}!r{self.row_number}"


def _field_counter():
    return {field_name: 0 for field_name in IMPORTED_FIELDS}


@dataclass
class HomeTicketImportReport:
    source: str
    total_rows: int = 0
    employees_updated: int = 0
    employees_unchanged: int = 0
    source_valid_fields: dict[str, int] = field(default_factory=_field_counter)
    updated_fields: dict[str, int] = field(default_factory=_field_counter)
    unchanged_fields: dict[str, int] = field(default_factory=_field_counter)
    not_found: list[ImportIssue] = field(default_factory=list)
    ambiguous: list[ImportIssue] = field(default_factory=list)
    ignored_values: list[ImportIssue] = field(default_factory=list)

    @property
    def updated(self):
        return self.employees_updated

    @property
    def unchanged(self):
        return self.employees_unchanged


def normalize_text(value):
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(character for character in text if not unicodedata.combining(character))
    return " ".join(re.findall(r"[a-z0-9]+", text.casefold()))


def normalize_identifier(value):
    text = unicodedata.normalize("NFKD", str(value or "")).upper()
    return re.sub(r"[^A-Z0-9]", "", text)


def clean_source_name(value):
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    text = re.split(r"_", text, maxsplit=1)[0]
    text = re.split(
        r",\s*(?:\d{6,}|CIM\b|CNP\b)|\s+CNP\b|\s+CIM\b|\s*/\s*AA\b|\s+salar\b",
        text,
        maxsplit=1,
        flags=re.IGNORECASE,
    )[0]
    return re.sub(r"\s+", " ", text).strip(" ,._-")


def clean_database_name(value):
    text = str(value or "").split("//", 1)[0]
    text = re.sub(r"\b[A-Z]{1,3}\s*\d{5,9}\b", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+\d+$", "", text)
    return re.sub(r"\s+", " ", text).strip(" ,._-")


def _sorted_name(value):
    return " ".join(sorted(normalize_text(value).split()))


def extract_identifiers(value):
    text = unicodedata.normalize("NFKD", str(value or "")).upper()
    text = re.split(r"\bCIM\b", text, maxsplit=1)[0]
    cnp = []
    passports = []
    for raw_part in re.split(r"[_/,;|]+", text):
        part = re.sub(r"\s+", " ", raw_part).strip(" .-")
        cnp_match = re.fullmatch(r"(?:CNP\s*)?(\d{13})", part)
        if cnp_match:
            value = cnp_match.group(1)
            if value not in cnp:
                cnp.append(value)
            continue

        passport_match = re.fullmatch(r"([A-Z]{1,3})[\s.-]*(\d{6,9})", part)
        if passport_match and passport_match.group(1) not in {"CIM", "CNP"}:
            value = normalize_identifier("".join(passport_match.groups()))
            if value not in passports:
                passports.append(value)
            continue

        numeric_passport = re.fullmatch(r"\d{6,9}", part)
        if numeric_passport:
            value = numeric_passport.group(0)
            if value not in passports:
                passports.append(value)

    return EmployeeIdentifiers(tuple(cnp), tuple(passports))


def parse_yes_no(value, field_name):
    normalized = re.sub(r"\s+", "", str(value or "")).casefold()
    if normalized == "da":
        return True
    if normalized == "nu":
        return False
    shown = repr(value) if value not in (None, "") else "valoare lipsă"
    raise HomeTicketImportError(f"{field_name}: se acceptă numai da/nu; primit {shown}")


def parse_excel_date(value, field_name, epoch):
    parsed = None
    if isinstance(value, datetime):
        parsed = value.date()
    elif isinstance(value, date):
        parsed = value
    elif isinstance(value, bool):
        parsed = None
    elif isinstance(value, (int, float)) and math.isfinite(value):
        try:
            excel_value = from_excel(value, epoch=epoch)
            parsed = excel_value.date() if isinstance(excel_value, datetime) else excel_value
        except (TypeError, ValueError, OverflowError):
            parsed = None
    elif isinstance(value, str):
        raw = value.strip()
        try:
            parsed = datetime.strptime(raw, "%d.%m.%Y").date()
        except ValueError:
            parsed = None

    if not isinstance(parsed, date) or not 1900 <= parsed.year <= 2200:
        shown = repr(value) if value not in (None, "") else "valoare lipsă"
        raise HomeTicketImportError(
            f"{field_name}: dată invalidă; folosește o dată Excel sau ZZ.LL.AAAA; primit {shown}"
        )
    return parsed


def _header_positions(sheet):
    positions = {}
    for index, value in enumerate(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))):
        normalized = normalize_text(value).replace(" ", "_")
        if normalized in REQUIRED_COLUMNS:
            positions[normalized] = index
    missing = [column for column in REQUIRED_COLUMNS if column not in positions]
    if missing:
        raise HomeTicketImportError(
            f"Foaia {sheet.title!r} nu conține coloanele obligatorii: {', '.join(missing)}"
        )
    return positions


def _cell(values, index):
    return values[index] if index < len(values) else None


def _is_blank(value):
    return value is None or (isinstance(value, str) and not value.strip())


def _ignored(sheet, row_number, raw_name, field_name, raw_value, reason):
    return ImportIssue(
        sheet=sheet,
        row_number=row_number,
        employee_name=raw_name,
        field_name=field_name,
        raw_value=raw_value,
        reason=reason,
    )


def read_home_ticket_workbook(path):
    path = Path(path)
    if not path.is_file():
        raise HomeTicketImportError(f"Fișierul nu există: {path}")
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise HomeTicketImportError(f"Nu pot deschide fișierul XLSX {path!s}: {exc}") from exc

    parsed_rows = []
    ignored_values = []
    total_rows = 0
    try:
        for sheet in workbook.worksheets:
            positions = _header_positions(sheet)
            for row_number, values in enumerate(
                sheet.iter_rows(min_row=2, values_only=True), start=2
            ):
                if not any(value not in (None, "") for value in values):
                    continue
                total_rows += 1
                raw_name_value = _cell(values, positions["employee_name"])
                raw_name = re.sub(r"\s+", " ", str(raw_name_value or "")).strip()
                if not raw_name:
                    ignored_values.append(_ignored(
                        sheet.title,
                        row_number,
                        raw_name,
                        "employee_name",
                        raw_name_value,
                        "employee_name lipsește; rândul nu poate fi asociat",
                    ))
                    continue

                operations = []
                row_issues = []

                raw_hire_date = _cell(values, positions["hire_date"])
                if not _is_blank(raw_hire_date):
                    try:
                        parsed_hire_date = parse_excel_date(
                            raw_hire_date, "hire_date", workbook.epoch
                        )
                        operations.append(FieldOperation(
                            "hire_date", "hire_date", raw_hire_date, parsed_hire_date
                        ))
                    except HomeTicketImportError as exc:
                        row_issues.append(_ignored(
                            sheet.title,
                            row_number,
                            raw_name,
                            "hire_date",
                            raw_hire_date,
                            str(exc),
                        ))

                raw_eligible = _cell(values, positions["ticket_bonus_eligible"])
                if not _is_blank(raw_eligible):
                    try:
                        eligible = parse_yes_no(raw_eligible, "ticket_bonus_eligible")
                        operations.append(FieldOperation(
                            "ticket_benefit_enabled",
                            "ticket_bonus_eligible",
                            raw_eligible,
                            eligible,
                        ))
                    except HomeTicketImportError as exc:
                        row_issues.append(_ignored(
                            sheet.title,
                            row_number,
                            raw_name,
                            "ticket_bonus_eligible",
                            raw_eligible,
                            str(exc),
                        ))

                raw_paid_before = _cell(values, positions["company_paid_ticket_before"])
                raw_departure = _cell(values, positions["last_paid_ticket_departure_date"])
                if _is_blank(raw_paid_before):
                    if not _is_blank(raw_departure):
                        try:
                            departure_date = parse_excel_date(
                                raw_departure,
                                "last_paid_ticket_departure_date",
                                workbook.epoch,
                            )
                            operations.append(FieldOperation(
                                "last_home_trip_date",
                                "last_paid_ticket_departure_date",
                                raw_departure,
                                departure_date,
                            ))
                        except HomeTicketImportError as exc:
                            row_issues.append(_ignored(
                                sheet.title,
                                row_number,
                                raw_name,
                                "last_paid_ticket_departure_date",
                                raw_departure,
                                str(exc),
                            ))
                else:
                    try:
                        paid_before = parse_yes_no(
                            raw_paid_before, "company_paid_ticket_before"
                        )
                    except HomeTicketImportError as exc:
                        row_issues.append(_ignored(
                            sheet.title,
                            row_number,
                            raw_name,
                            "company_paid_ticket_before",
                            raw_paid_before,
                            str(exc),
                        ))
                        if not _is_blank(raw_departure):
                            row_issues.append(_ignored(
                                sheet.title,
                                row_number,
                                raw_name,
                                "last_paid_ticket_departure_date",
                                raw_departure,
                                "data nu poate stabili istoricul deoarece "
                                "company_paid_ticket_before este invalid",
                            ))
                    else:
                        if paid_before is False:
                            operations.append(FieldOperation(
                                "last_home_trip_date",
                                "company_paid_ticket_before",
                                raw_paid_before,
                                None,
                            ))
                            if not _is_blank(raw_departure):
                                row_issues.append(_ignored(
                                    sheet.title,
                                    row_number,
                                    raw_name,
                                    "last_paid_ticket_departure_date",
                                    raw_departure,
                                    "valoare ignorată deoarece "
                                    "company_paid_ticket_before = nu golește istoricul",
                                ))
                        elif _is_blank(raw_departure):
                            row_issues.append(_ignored(
                                sheet.title,
                                row_number,
                                raw_name,
                                "company_paid_ticket_before",
                                raw_paid_before,
                                "valoarea da nu poate fi aplicată fără o "
                                "last_paid_ticket_departure_date validă",
                            ))
                        else:
                            try:
                                departure_date = parse_excel_date(
                                    raw_departure,
                                    "last_paid_ticket_departure_date",
                                    workbook.epoch,
                                )
                                operations.append(FieldOperation(
                                    "last_home_trip_date",
                                    "last_paid_ticket_departure_date",
                                    raw_departure,
                                    departure_date,
                                ))
                            except HomeTicketImportError as exc:
                                row_issues.append(_ignored(
                                    sheet.title,
                                    row_number,
                                    raw_name,
                                    "last_paid_ticket_departure_date",
                                    raw_departure,
                                    str(exc),
                                ))

                ignored_values.extend(row_issues)
                if not operations and not row_issues:
                    continue

                source_name = clean_source_name(raw_name)
                normalized_name = normalize_text(source_name)
                if not normalized_name:
                    ignored_values.append(_ignored(
                        sheet.title,
                        row_number,
                        raw_name,
                        "employee_name",
                        raw_name_value,
                        "employee_name nu conține un nume utilizabil după normalizare",
                    ))
                    continue
                parsed_rows.append(HomeTicketRow(
                    sheet=sheet.title,
                    row_number=row_number,
                    raw_name=raw_name,
                    normalized_name=normalized_name,
                    sorted_name=_sorted_name(source_name),
                    identifiers=extract_identifiers(raw_name),
                    operations=tuple(operations),
                ))
    finally:
        workbook.close()
    return total_rows, parsed_rows, ignored_values


def _employee_indexes(employees):
    by_pk = {employee.pk: employee for employee in employees}
    cnp = {}
    passports = {}
    names = {}
    sorted_names = {}

    def add(index, key, employee):
        if key:
            index.setdefault(key, set()).add(employee.pk)

    for employee in employees:
        for value in (employee.UserSerie, employee.NameAndSerie, employee.UserName):
            identifiers = extract_identifiers(value)
            for identifier in identifiers.cnp:
                add(cnp, identifier, employee)
            for identifier in identifiers.passports:
                add(passports, identifier, employee)

        cleaned_name = clean_database_name(employee.UserName)
        add(names, normalize_text(cleaned_name), employee)
        add(sorted_names, _sorted_name(cleaned_name), employee)
    return by_pk, cnp, passports, names, sorted_names


def _candidate_description(candidates):
    return ", ".join(
        f"{employee.UserName} [ID {employee.pk}]"
        for employee in sorted(candidates, key=lambda item: item.pk)
    )


def match_home_ticket_row(row, indexes):
    by_pk, cnp_index, passport_index, names, sorted_names = indexes
    identifier_ids = set()
    matched_identifiers = []
    for identifier in row.identifiers.cnp:
        matches = cnp_index.get(identifier, set())
        if matches:
            identifier_ids.update(matches)
            matched_identifiers.append(f"CNP {identifier}")
    for identifier in row.identifiers.passports:
        matches = passport_index.get(identifier, set())
        if matches:
            identifier_ids.update(matches)
            matched_identifiers.append(f"pașaport {identifier}")

    if len(identifier_ids) == 1:
        return by_pk[identifier_ids.pop()], None
    if len(identifier_ids) > 1:
        candidates = [by_pk[pk] for pk in identifier_ids]
        reason = (
            f"identificatorii {', '.join(matched_identifiers)} indică mai mulți angajați: "
            f"{_candidate_description(candidates)}"
        )
        return None, reason

    candidate_ids = names.get(row.normalized_name, set())
    method = "numele normalizat"
    if not candidate_ids:
        candidate_ids = sorted_names.get(row.sorted_name, set())
        method = "aceleași cuvinte din numele normalizat"
    if len(candidate_ids) == 1:
        return by_pk[next(iter(candidate_ids))], None
    if len(candidate_ids) > 1:
        candidates = [by_pk[pk] for pk in candidate_ids]
        return None, (
            f"{method} corespunde mai multor angajați: {_candidate_description(candidates)}"
        )

    attempted = list(row.identifiers.cnp) + list(row.identifiers.passports)
    identifier_note = (
        f"; identificatori fără corespondent: {', '.join(attempted)}" if attempted else ""
    )
    return None, f"niciun angajat cu numele normalizat {row.normalized_name!r}{identifier_note}"


def _issue_from_row(row, reason):
    return ImportIssue(row.sheet, row.row_number, row.raw_name, reason)


def _issue_from_operation(row, operation, reason):
    return ImportIssue(
        sheet=row.sheet,
        row_number=row.row_number,
        employee_name=row.raw_name,
        reason=reason,
        field_name=operation.source_column,
        raw_value=operation.raw_value,
    )


def import_home_ticket_benefits(path, Users, *, using="default", apply=True, writer=print):
    path = Path(path)
    report = HomeTicketImportReport(source=str(path))
    with transaction.atomic(using=using):
        total_rows, rows, ignored_values = read_home_ticket_workbook(path)
        report.total_rows = total_rows
        report.ignored_values.extend(ignored_values)
        for row in rows:
            for operation in row.operations:
                report.source_valid_fields[operation.field_name] += 1

        employee_query = Users.objects.using(using).filter(person_type="employee").order_by("pk")
        if apply:
            employee_query = employee_query.select_for_update()
        employees = list(employee_query)
        indexes = _employee_indexes(employees)

        matched = {}
        for row in rows:
            employee, reason = match_home_ticket_row(row, indexes)
            if employee is None:
                issue = _issue_from_row(row, reason)
                if reason.startswith("niciun angajat"):
                    report.not_found.append(issue)
                else:
                    report.ambiguous.append(issue)
                continue
            if not row.operations:
                continue
            item = matched.setdefault(employee.pk, {
                "employee": employee,
                "operations": {field_name: [] for field_name in IMPORTED_FIELDS},
            })
            for operation in row.operations:
                item["operations"][operation.field_name].append((row, operation))

        for item in matched.values():
            employee = item["employee"]
            resolved = {}
            for field_name, field_operations in item["operations"].items():
                if not field_operations:
                    continue
                distinct_values = {operation.value for _row, operation in field_operations}
                if len(distinct_values) > 1:
                    locations = ", ".join(row.location for row, _operation in field_operations)
                    for row, operation in field_operations:
                        report.ignored_values.append(_issue_from_operation(
                            row,
                            operation,
                            f"valoare conflictuală pentru {FIELD_LABELS[field_name].lower()}; "
                            f"același angajat apare pe rândurile {locations}",
                        ))
                    continue
                resolved[field_name] = (
                    next(iter(distinct_values)),
                    field_operations,
                )

            if "last_home_trip_date" in resolved:
                history_value, history_operations = resolved["last_home_trip_date"]
                effective_hire_date = (
                    resolved["hire_date"][0]
                    if "hire_date" in resolved
                    else employee.hire_date
                )
                if (
                    history_value is not None
                    and effective_hire_date is not None
                    and history_value < effective_hire_date
                ):
                    for row, operation in history_operations:
                        report.ignored_values.append(_issue_from_operation(
                            row,
                            operation,
                            "data ultimei plecări este anterioară datei angajării",
                        ))
                    del resolved["last_home_trip_date"]

            update_values = {}
            for field_name, (value, _operations) in resolved.items():
                if getattr(employee, field_name) == value:
                    report.unchanged_fields[field_name] += 1
                else:
                    report.updated_fields[field_name] += 1
                    update_values[field_name] = value

            if update_values:
                if apply:
                    Users.objects.using(using).filter(pk=employee.pk).update(**update_values)
                report.employees_updated += 1
            elif resolved:
                report.employees_unchanged += 1

    write_home_ticket_report(report, writer=writer, applied=apply)
    return report


def write_home_ticket_report(report, *, writer=print, applied=True):
    writer("Import Ajutor bilet acasă")
    writer(f"Fișier: {report.source}")
    writer(f"Rânduri citite: {report.total_rows}")
    writer(f"Angajați cu cel puțin un câmp actualizat: {report.employees_updated}")
    writer(f"Angajați cu valori valide deja nemodificate: {report.employees_unchanged}")
    for field_name in IMPORTED_FIELDS:
        label = FIELD_LABELS[field_name]
        writer(f"{label} valide în Excel: {report.source_valid_fields[field_name]}")
        writer(f"{label} actualizate: {report.updated_fields[field_name]}")
        writer(f"{label} deja nemodificate: {report.unchanged_fields[field_name]}")
    writer(f"Angajați negăsiți în aplicație: {len(report.not_found)}")
    writer(f"Potriviri ambigue: {len(report.ambiguous)}")
    writer(f"Valori ignorate: {len(report.ignored_values)}")

    sections = (
        ("Angajați negăsiți în aplicație", report.not_found),
        ("Potriviri ambigue (nu au fost actualizate)", report.ambiguous),
    )
    for title, issues in sections:
        if not issues:
            continue
        writer("")
        writer(title + ":")
        for issue in issues:
            writer(
                f"  - {issue.location} | employee_name={issue.employee_name!r} | "
                f"motiv: {issue.reason}"
            )
    if report.ignored_values:
        writer("")
        writer("Valori ignorate:")
        for issue in report.ignored_values:
            writer(
                f"  - {issue.location} | employee_name={issue.employee_name!r} | "
                f"câmp={issue.field_name} | valoare={issue.raw_value!r} | "
                f"motiv: {issue.reason}"
            )
    if not applied:
        writer("")
        writer("DRY-RUN: baza de date nu a fost modificată.")
